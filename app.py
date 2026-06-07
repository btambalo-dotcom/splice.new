from flask import Flask, render_template, request, redirect, url_for, flash, send_file, abort, session, make_response, jsonify, current_app
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from datetime import datetime, date
from sqlalchemy import text, case, or_, inspect, func
from sqlalchemy.orm import deferred
import os
import json
import os
from fpdf import FPDF
from werkzeug.utils import secure_filename
from io import BytesIO
import urllib.parse
import urllib.request
import io
import zipfile
import hashlib
from PIL import Image, ImageOps
from functools import wraps
import csv
import threading
import time
import traceback
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import re
import xml.etree.ElementTree as ET

import base64
import requests

import uuid
import boto3
from botocore.client import Config as BotoConfig
from botocore.exceptions import ClientError
# --------- Util: compress/resize uploaded photos ---------
def process_uploaded_photo(file_storage, max_size=(1600, 1600), quality=80, thumb_max_size=(480, 480), thumb_quality=65):
    """Resize/compress an uploaded image to keep DB small and fast.

    Returns (filename, content_type, data_bytes, thumb_bytes, thumb_content_type)
    or (None, None, None, None, None) on failure.
    """
    if not file_storage or not getattr(file_storage, "filename", None):
        return None, None, None, None, None

    try:
        raw = file_storage.read()
        if not raw:
            return None, None, None, None, None

        img = Image.open(BytesIO(raw))
        # Corrige orientação baseada no EXIF (fotos de celular)
        try:
            img = ImageOps.exif_transpose(img)
        except Exception:
            pass

        # garante formato compatível
        img = img.convert("RGB")

        max_side = int(os.environ.get("UPLOAD_IMAGE_MAX_SIDE", "1600"))
        jpg_quality = int(os.environ.get("UPLOAD_IMAGE_JPEG_QUALITY", "80"))
        thumb_side = int(os.environ.get("UPLOAD_THUMB_MAX_SIDE", "420"))
        thumb_quality = int(os.environ.get("UPLOAD_THUMB_JPEG_QUALITY", "70"))

        # FULL (já comprimida)
        full = img.copy()
        full.thumbnail((max_side, max_side))
        buf = BytesIO()
        full.save(buf, format="JPEG", quality=jpg_quality, optimize=True, progressive=True)
        full_bytes = buf.getvalue()

        # THUMB
        thumb = img.copy()
        thumb.thumbnail((thumb_side, thumb_side))
        buf2 = BytesIO()
        thumb.save(buf2, format="JPEG", quality=thumb_quality, optimize=True, progressive=True)
        thumb_bytes = buf2.getvalue()

        base_name, _ = os.path.splitext(file_storage.filename or "foto.jpg")
        safe_name = secure_filename(base_name) or "foto"
        filename = f"{safe_name}.jpg"

        return filename[:255], "image/jpeg", full_bytes, thumb_bytes, "image/jpeg"

    except Exception:
        # Se der qualquer erro ao processar a imagem, tenta salvar o arquivo original (sem thumb)
        try:
            try:
                file_storage.seek(0)
            except Exception:
                pass
            data = file_storage.read()
            if not data:
                return None, None, None, None, None
            filename = secure_filename(file_storage.filename) or "arquivo.bin"
            return filename[:255], (file_storage.mimetype or "application/octet-stream"), data, None, None
        except Exception:
            return None, None, None, None, None







def extract_photo_date_from_exif(raw_bytes):
    """Tenta extrair a data da foto a partir do EXIF.

    Retorna um datetime.date ou None.
    """
    if not raw_bytes:
        return None
    try:
        img = Image.open(BytesIO(raw_bytes))
        exif = getattr(img, "_getexif", lambda: None)() or {}
        # 36867 = DateTimeOriginal, 306 = DateTime
        dt_str = exif.get(36867) or exif.get(306)
        if not dt_str:
            return None
        # Formato típico: 'YYYY:MM:DD HH:MM:SS'
        parts = str(dt_str).split()
        if not parts:
            return None
        date_part = parts[0].replace(":", "-")
        try:
            dt = datetime.strptime(date_part, "%Y-%m-%d")
            return dt.date()
        except Exception:
            return None
    except Exception:
        return None



def _extract_kmz_meta(name: str | None, description: str | None):
    name = (name or '').strip()
    description = (description or '').strip()
    port_label = None
    pon_name = None
    ote_label = None
    splitter_name = None

    m = re.search(r"\b(P\d+)\b", name, re.I)
    if m:
        port_label = m.group(1).upper()

    m = re.search(r"(?:^|[;|\n\r ]+)PON\s*[:= -]*\s*(\d+)\b", description, re.I)
    if m:
        pon_name = f"PON {m.group(1)}"

    if not pon_name:
        m = re.search(r"(?:^|[;|\n\r ]+)P(\d+)\b", description, re.I)
        if m:
            pon_name = f"PON {m.group(1)}"

    m = re.search(r"(?:OTE|FIBRAS?)\s*[:= -]*\s*([0-9]+(?:\s*[-/]\s*[0-9]+)?)", description, re.I)
    if m:
        ote_label = re.sub(r"\s+", '', m.group(1))
    else:
        m = re.search(r"\b([0-9]+\s*[-/]\s*[0-9]+)\b", description)
        if m:
            ote_label = re.sub(r"\s+", '', m.group(1))

    m = re.search(r"SPLITTER\s*[:= -]*\s*([^;|\n\r]+)", description, re.I)
    if m:
        splitter_name = m.group(1).strip()

    return {
        'port_label': port_label,
        'pon_name': pon_name,
        'ote_label': ote_label,
        'splitter_name': splitter_name,
    }


def _get_syscfg():
    cfg = SystemConfig.query.first()
    if not cfg:
        cfg = SystemConfig()
        db.session.add(cfg)
        db.session.commit()
    return cfg


def geoapify_reverse_geocode(lat, lng):
    try:
        cfg = _get_syscfg()
        api_key = (getattr(cfg, 'geoapify_api_key', None) or os.environ.get('GEOAPIFY_API_KEY') or '').strip()
        if not api_key:
            return None
        params = urllib.parse.urlencode({
            'lat': f'{float(lat):.8f}',
            'lon': f'{float(lng):.8f}',
            'apiKey': api_key,
        })
        url = f'https://api.geoapify.com/v1/geocode/reverse?{params}'
        with urllib.request.urlopen(url, timeout=12) as resp:
            payload = json.loads(resp.read().decode('utf-8', errors='ignore'))
        feats = payload.get('features') or []
        if not feats:
            return None
        props = feats[0].get('properties') or {}
        return (props.get('formatted') or '').strip() or None
    except Exception:
        return None



def build_network_for_map(map_obj):
    records = Record.query.filter(Record.map == map_obj.name)
    if map_obj.company:
        records = records.filter(Record.company == map_obj.company)
    records = records.filter(Record.latitude.isnot(None), Record.longitude.isnot(None)).all()

    def dist(a, b):
        ax, ay = float(a.latitude or 0), float(a.longitude or 0)
        bx, by = float(b.latitude or 0), float(b.longitude or 0)
        return ((ax - bx) ** 2 + (ay - by) ** 2) ** 0.5

    def norm(v):
        return (v or '').strip().upper()

    updated = 0
    groups = {}
    for rec in records:
        pon = norm(getattr(rec, 'pon_name', None))
        if not pon:
            continue
        groups.setdefault(pon, []).append(rec)

    for pon, items in groups.items():
        splitters = [r for r in items if norm(r.type) == 'SPLITTER']
        devices = [r for r in items if norm(r.type or 'OTE') != 'SPLITTER']
        if not splitters or not devices:
            continue

        splitter = next((r for r in splitters if (r.device or '').strip()), splitters[0])

        # Regra correta:
        # - o splitter é sempre a raiz do PON
        # - o sinal sempre parte dele
        # - FROM/OUT são calculados pela árvore mais econômica (MST),
        #   orientada a partir do splitter
        nodes = [splitter] + devices

        # Prim MST
        connected = {splitter}
        parent = {}
        children = {n: [] for n in nodes}

        while len(connected) < len(nodes):
            best_u = None
            best_v = None
            best_d = None
            for u in list(connected):
                for v in nodes:
                    if v in connected:
                        continue
                    d = dist(u, v)
                    if best_d is None or d < best_d:
                        best_d = d
                        best_u = u
                        best_v = v
            if best_v is None:
                break
            connected.add(best_v)
            parent[best_v] = best_u
            children.setdefault(best_u, []).append(best_v)
            children.setdefault(best_v, [])

        # Ordena filhos de cada nó por proximidade para ficar consistente
        for u, childs in list(children.items()):
            childs.sort(key=lambda c: dist(u, c))

        # Grava FROM/OUT:
        # - splitter tem FROM vazio e OUT como lista das primeiras saídas
        # - cada dispositivo recebe FROM = pai
        # - OUT = filhos, quando houver; se não houver, vazio
        splitter_from = ''
        splitter_out = ', '.join([(c.device or '').strip() for c in children.get(splitter, []) if (c.device or '').strip()])
        if getattr(splitter, 'source_from', None) != (splitter_from or None):
            splitter.source_from = splitter_from or None
            updated += 1
        if getattr(splitter, 'source_out', None) != (splitter_out or None):
            splitter.source_out = splitter_out or None
            updated += 1

        for rec in devices:
            p = parent.get(rec)
            new_from = ((p.device or p.splitter_name or 'SPLITTER').strip() if p is not None else '')
            child_names = [(c.device or '').strip() for c in children.get(rec, []) if (c.device or '').strip()]
            # Se houver apenas um filho, OUT é esse próximo device.
            # Se houver bifurcação, lista todos.
            new_out = ', '.join(child_names) if child_names else ''

            if getattr(rec, 'source_from', None) != (new_from or None):
                rec.source_from = new_from or None
                updated += 1
            if getattr(rec, 'source_out', None) != (new_out or None):
                rec.source_out = new_out or None
                updated += 1

    if updated:
        db.session.commit()
    return updated



def import_kmz_for_map(company_map, file_storage):
    """Importa um arquivo KMZ e cria/atualiza Records com coordenadas
    para o mapa/projeto informado.

    Cada Placemark -> Point vira um dispositivo (Record) com latitude/longitude.
    O nome do Placemark é usado como nome do device.
    """
    if not file_storage or not getattr(file_storage, "filename", None):
        return 0

    try:
        raw = file_storage.read()
        if not raw:
            return 0

        # KMZ é um ZIP com um ou mais arquivos .kml dentro
        with zipfile.ZipFile(BytesIO(raw)) as z:
            kml_name = None
            for name in z.namelist():
                if name.lower().endswith(".kml"):
                    kml_name = name
                    break
            if not kml_name:
                return 0
            kml_bytes = z.read(kml_name)

        root = ET.fromstring(kml_bytes)
    except Exception:
        # KMZ inválido
        return 0

    ns = {"k": "http://www.opengis.net/kml/2.2"}
    created_or_updated = 0

    from sqlalchemy import and_

    # Mapeia cada Placemark para uma seção (nome da pasta ou ExtendedData)
    placemark_sections = {}

    # 1) Folders -> seções
    for folder in root.findall(".//k:Folder", ns):
        fname_el = folder.find("k:name", ns)
        sec_name = (fname_el.text or "").strip() if fname_el is not None else None
        if not sec_name:
            continue
        for pm in folder.findall(".//k:Placemark", ns):
            placemark_sections[id(pm)] = sec_name

    # 2) percorre todos os Placemarks com Point/coordinates
    for pm in root.findall(".//k:Placemark", ns):
        # seção padrão herdada da pasta (se houver)
        section_name = placemark_sections.get(id(pm))

        # tenta sobrescrever pela ExtendedData (campo section/secao/seção/setor/sector)
        # tenta sobrescrever pela ExtendedData (seção, nome do device, observações)
        ext = pm.find("k:ExtendedData", ns)
        extra_name = None
        extra_info = None
        if ext is not None:
            for data_el in ext.findall("k:Data", ns):
                key = (data_el.get("name") or "").lower()
                val_el = data_el.find("k:value", ns)
                val_text = (val_el.text or "").strip() if val_el is not None and val_el.text else ""
                if not val_text:
                    continue
                # seção
                if key in ("section", "secao", "seção", "setor", "sector"):
                    section_name = val_text
                    continue
                # possíveis nomes de campo para o device
                if key in ("device", "nome", "name", "cto", "caixa", "equip", "equipamento"):
                    extra_name = extra_name or val_text
                    continue
                # possíveis nomes de campo para observações
                if key in ("info", "obs", "observacao", "observação", "notes", "descricao", "descrição", "description"):
                    extra_info = extra_info or val_text
                    continue

        name_el = pm.find("k:name", ns)
        name = (name_el.text or "").strip() if name_el is not None else ""
        if extra_name:
            name = extra_name

        # descrição padrão do Placemark vira observação se nada vier no ExtendedData
        desc_el = pm.find("k:description", ns)
        description_text = (desc_el.text or "").strip() if desc_el is not None and desc_el.text else ""
        device_info_val = extra_info or description_text or None
        kmz_meta = _extract_kmz_meta(name, device_info_val)
        coord_el = pm.find(".//k:Point/k:coordinates", ns)
        if coord_el is None or not coord_el.text:
            continue

        coord_text = coord_el.text.strip().split()[0]
        parts = coord_text.split(",")
        if len(parts) < 2:
            continue

        try:
            lon = float(parts[0])
            lat = float(parts[1])
        except ValueError:
            continue

        # procura Record existente para mesma empresa+device+projeto,
        # independente do mapa antigo. Assim, ao importar o KMZ em um
        # novo mapa, reaproveitamos os lançamentos já feitos (splices, testes, fotos),
        # apenas movendo o dispositivo para o mapa atual.
        filters = [
            Record.company == company_map.company,
            Record.device == name,
            Record.map == company_map.name,
        ]
        if company_map.project_id is not None:
            filters.append(Record.project_id == company_map.project_id)

        # Agora só reaproveitamos um Record se ele já for deste mesmo mapa.
        # Se existir apenas em outro mapa, vamos criar um novo registro,
        # para não "roubar" o dispositivo de outro mapa diferente.
        rec = Record.query.filter(and_(*filters)).order_by(Record.id.desc()).first()

        if rec is None:
            rec = Record(
                company=company_map.company,
                project_id=company_map.project_id,
                map=company_map.name,
                device=name,
                type='OTE',
                splices=0,
                splicer="",
                created_date=None,
                section=section_name,
                device_info=device_info_val,
                port_label=kmz_meta.get('port_label'),
                pon_name=kmz_meta.get('pon_name'),
                ote_label=kmz_meta.get('ote_label'),
                splitter_name=kmz_meta.get('splitter_name'),
            )
            db.session.add(rec)

        if not rec:
            rec = Record(
                map=company_map.name,
                company=company_map.company,
                project_id=company_map.project_id,
                device=name or f"DEVICE-{created_or_updated+1}",
                type='OTE',
                splices=0,
                splicer="",
                created_date=None,
                section=section_name,
                device_info=device_info_val,
                port_label=kmz_meta.get('port_label'),
                pon_name=kmz_meta.get('pon_name'),
                ote_label=kmz_meta.get('ote_label'),
                splitter_name=kmz_meta.get('splitter_name'),
            )
            db.session.add(rec)

        # Atualiza seção (se veio do KMZ); não sobrescreve se já houver uma definida manualmente
        if section_name and not getattr(rec, "section", None):
            rec.section = section_name

        # Atualiza coordenadas sempre que importar
        rec.latitude = lat
        rec.longitude = lon
        if not (rec.type or '').strip():
            rec.type = 'OTE'
        if kmz_meta.get('port_label') and not getattr(rec, 'port_label', None):
            rec.port_label = kmz_meta.get('port_label')
        if kmz_meta.get('pon_name') and not getattr(rec, 'pon_name', None):
            rec.pon_name = kmz_meta.get('pon_name')
        if kmz_meta.get('ote_label') and not getattr(rec, 'ote_label', None):
            rec.ote_label = kmz_meta.get('ote_label')
        if kmz_meta.get('splitter_name') and not getattr(rec, 'splitter_name', None):
            rec.splitter_name = kmz_meta.get('splitter_name')
        if not getattr(rec, 'geo_address', None):
            rec.geo_address = geoapify_reverse_geocode(lat, lon)
        created_or_updated += 1


    if created_or_updated:
        db.session.commit()

    return created_or_updated


# --------- App & DB setup ---------
app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "dev-secret-key-change-me")

@app.template_filter('from_json')
def from_json_filter(value):
    """Jinja2 filter: converte JSON string para objeto Python."""
    if not value:
        return []
    try:
        return json.loads(value)
    except Exception:
        return []
app.config["SECRET_KEY"] = os.environ.get("FLASK_SECRET_KEY", "dev-key")

# Database configuration: prefer DATABASE_URL/RENDER_DATABASE_URL (e.g. Render PostgreSQL),
# fallback to local SQLite for development.
db_url = os.environ.get("DATABASE_URL") or os.environ.get("RENDER_DATABASE_URL") or "sqlite:///data.db"
if db_url.startswith("postgres://"):
    db_url = db_url.replace("postgres://", "postgresql://", 1)

app.config["SQLALCHEMY_DATABASE_URI"] = db_url
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False


db = SQLAlchemy(app)
# --- Auto-migrate SQLite schemas on startup (Render/local) ---
try:
    from auto_migrate_all_dbs import main as _auto_migrate_all_dbs_main
    _auto_migrate_all_dbs_main()
except Exception as _e:
    # Log but do not crash the app; migrations are additive and idempotent.
    print('DB auto-migration failed:', _e)

login_manager = LoginManager(app)
login_manager.login_view = "login"

# --- Performance: ensure important DB indexes exist (idempotent) ---
from sqlalchemy import text as _sql_text

def ensure_db_indexes() -> None:
    """Create indexes used by filters (company/map/device, date, etc.)."""
    with app.app_context():
        try:
            db.session.execute(_sql_text(
                "CREATE INDEX IF NOT EXISTS idx_records_company_map_device ON record (company, map, device)"
            ))
            db.session.execute(_sql_text(
                "CREATE INDEX IF NOT EXISTS idx_records_company_date ON record (company, created_date)"
            ))
            db.session.execute(_sql_text(
                "CREATE INDEX IF NOT EXISTS idx_invoices_created_at ON invoice (created_at)"
            ))
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            print('ensure_db_indexes skipped:', e)

ensure_db_indexes()

# --------- R2 (S3-compatible) storage helpers ---------
_r2_client = None

def r2_bucket_name() -> str:
    """Bucket name for Cloudflare R2.

    The Render dashboard screenshots show env var name `R2_BUCKET_NAME`, while older versions used `R2_BUCKET`.
    We support both.
    """
    return (os.environ.get("R2_BUCKET_NAME") or os.environ.get("R2_BUCKET") or "").strip()

def r2_enabled() -> bool:
    return bool(
        os.environ.get("R2_ACCESS_KEY_ID")
        and os.environ.get("R2_SECRET_ACCESS_KEY")
        and os.environ.get("R2_ENDPOINT")
        and r2_bucket_name()
    )

def r2_client():
    global _r2_client
    if _r2_client is not None:
        return _r2_client
    if not r2_enabled():
        return None
    _r2_client = boto3.client(
        "s3",
        endpoint_url=os.environ.get("R2_ENDPOINT"),
        aws_access_key_id=os.environ.get("R2_ACCESS_KEY_ID"),
        aws_secret_access_key=os.environ.get("R2_SECRET_ACCESS_KEY"),
        region_name=os.environ.get("R2_REGION", "auto"),
        config=BotoConfig(signature_version="s3v4", s3={"addressing_style": "path"}),
    )
    return _r2_client

def r2_bucket() -> str:
    # Backwards compatible wrapper
    return r2_bucket_name()

def r2_put_bytes(key: str, data: bytes, content_type: str | None = None) -> dict:
    c = r2_client()
    if c is None:
        raise RuntimeError("R2 is not enabled")
    extra = {}
    if content_type:
        extra["ContentType"] = content_type
    resp = c.put_object(Bucket=r2_bucket(), Key=key, Body=data, **extra)
    return resp or {}

def r2_get_bytes(key: str) -> bytes:
    c = r2_client()
    if c is None:
        raise RuntimeError("R2 is not enabled")
    obj = c.get_object(Bucket=r2_bucket(), Key=key)
    body = obj.get("Body")
    if body is None:
        return b""
    if hasattr(body, "read"):
        return body.read()
    if isinstance(body, (bytes, bytearray)):
        return bytes(body)
    return str(body).encode("utf-8", errors="ignore")

def r2_key_for_record_photo(record_id: int, filename: str) -> str:
    safe = re.sub(r"[^a-zA-Z0-9._-]+", "_", (filename or "photo"))
    return f"records/{record_id}/{uuid.uuid4().hex}_{safe}"


def optimize_upload_bytes(file_bytes: bytes, original_content_type: str) -> tuple[bytes, str]:
    """Reduce payload size for uploads.

    - If image: auto-rotate (EXIF), downscale, and re-encode to JPEG.
    - Otherwise: return as-is.
    """
    try:
        img = Image.open(io.BytesIO(file_bytes))
        img = ImageOps.exif_transpose(img)
        if img.mode != "RGB":
            img = img.convert("RGB")

        max_size = int(os.environ.get("UPLOAD_IMAGE_MAX_SIDE", "1600"))
        img.thumbnail((max_size, max_size))

        out = io.BytesIO()
        quality = int(os.environ.get("UPLOAD_IMAGE_JPEG_QUALITY", "75"))
        img.save(out, format="JPEG", quality=quality, optimize=True, progressive=True)
        return out.getvalue(), "image/jpeg"
    except Exception:
        return file_bytes, (original_content_type or "application/octet-stream")


def enqueue_r2_upload(photo_id: int, key: str, data: bytes, content_type: str, thumb_key=None, thumb_bytes=None, thumb_content_type=None):
    """Upload to R2 in a background thread (NOLOSS).

    IMPORTANT:
    - On the create-photo request we may enqueue before the DB transaction commits.
      On Render/Gunicorn, the background thread uses a new DB connection, so it may not
      see the photo row immediately.
    - We retry a few times to wait for commit, then upload and update r2_key.
    - We ALWAYS log errors to Render logs (stdout) to avoid silent failures.
    """
    if not r2_enabled():
        print("[R2] Disabled (missing env vars). Skipping upload.")
        return

    def _worker():
        try:
            with app.app_context():
                # Wait a bit for the photo row to become visible (transaction commit)
                p = None
                for _ in range(int(os.environ.get("R2_DB_RETRY_ATTEMPTS", "40"))):
                    p = RecordPhoto.query.get(photo_id)
                    if p is not None:
                        break
                    time.sleep(float(os.environ.get("R2_DB_RETRY_SLEEP", "0.25")))
                if not p:
                    print(f"[R2] Photo id={photo_id} not found after retries. Skipping.")
                    return
                if p.r2_key:
                    return

                # Upload bytes to R2 (original + thumbnail)
                r2_put_bytes(key, data, content_type=content_type)

                if thumb_key and thumb_bytes:
                    # thumbnail always jpeg by default
                    r2_put_bytes(thumb_key, thumb_bytes, content_type=(thumb_content_type or "image/jpeg"))
                    p.r2_thumb_key = thumb_key

                # Mark uploaded
                p.r2_key = key
                if os.environ.get("CLEAR_DB_AFTER_R2", "0") == "1":
                    p.data = b""
                    try:
                        p.thumb_data = None
                    except Exception:
                        pass
                db.session.commit()
                if thumb_key and thumb_bytes:
                    print(f"[R2] Uploaded photo id={photo_id} -> key={key} (thumb={thumb_key})")
                else:
                    print(f"[R2] Uploaded photo id={photo_id} -> key={key}")
        except Exception as e:
            print(f"[R2] Upload failed for photo id={photo_id} key={key}: {e}")
            traceback.print_exc()
            try:
                with app.app_context():
                    db.session.rollback()
            except Exception:
                pass

    threading.Thread(target=_worker, daemon=True).start()


# --------- Models ---------
class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password = db.Column(db.String(120), nullable=False)  # simples, sem hash, uso local
    is_admin = db.Column(db.Boolean, default=False)
    splicer_name = db.Column(db.String(120), nullable=True)  # nome que aparece como Splicer nos lançamentos
    is_company_owner = db.Column(db.Boolean, default=False, nullable=False)  # dono de empresa: vê registros da própria empresa
    company_name = db.Column(db.String(120), nullable=True)  # nome da empresa a que o usuário pertence
    # Permissões granulares
    is_active = db.Column(db.Boolean, default=True, nullable=False)  # se False, login bloqueado
    can_access_expenses = db.Column(db.Boolean, default=False, nullable=False)  # pode acessar módulo de despesas
    can_view_values = db.Column(db.Boolean, default=True, nullable=False)  # pode ver valores financeiros nos lançamentos

    # Mapas interativos aos quais o usuário (splicer) tem acesso explícito.
    maps_with_access = db.relationship(
        "CompanyMap",
        secondary="map_splicer_access",
        back_populates="allowed_splicers",
        lazy="select",
    )


class CompanyConfig(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), unique=True, nullable=False)
    included_splices = db.Column(db.Integer, default=1, nullable=False)  # fusões inclusas por lançamento
    invoice_address = db.Column(db.Text, nullable=True)  # nome + endereço p/ usar na invoice


class Project(db.Model):
    """Projeto dentro de uma empresa, com regras e valores próprios."""
    id = db.Column(db.Integer, primary_key=True)
    company = db.Column(db.String(120), nullable=False, index=True)
    name = db.Column(db.String(200), nullable=False)
    # Se None, usa o valor da empresa (CompanyConfig.included_splices)
    included_splices = db.Column(db.Integer, nullable=True)
    # Prazo de pagamento em dias após a criação do payroll
    payment_days = db.Column(db.Integer, nullable=True, default=30)

    __table_args__ = (
        db.UniqueConstraint('company', 'name', name='uq_project_company_name'),
    )


class SystemConfig(db.Model):
    """Configurações gerais do sistema (dados da sua empresa para sair na invoice)."""
    id = db.Column(db.Integer, primary_key=True)
    my_company_name = db.Column(db.String(200), nullable=True)
    my_company_address = db.Column(db.Text, nullable=True)
    my_company_tax_id = db.Column(db.String(120), nullable=True)
    my_company_email = db.Column(db.String(120), nullable=True)
    my_company_phone = db.Column(db.String(60), nullable=True)
    geoapify_api_key = db.Column(db.String(255), nullable=True)
    board_header = db.Column(db.String(120), nullable=True)


class Invoice(db.Model):
    """Invoices geradas para controle contábil."""
    id = db.Column(db.Integer, primary_key=True)
    number = db.Column(db.String(50), nullable=False, unique=True)
    company = db.Column(db.String(120), nullable=False)
    start_date = db.Column(db.Date, nullable=True)
    end_date = db.Column(db.Date, nullable=True)
    total_usd = db.Column(db.Float, nullable=False, default=0.0)
    status = db.Column(db.String(20), nullable=False, default="pending")  # pending / paid
    created_at = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    pdf_filename = db.Column(db.String(255), nullable=True)
    pdf_content_type = db.Column(db.String(100), nullable=True)
    pdf_data = db.Column(db.LargeBinary, nullable=True)

    created_by_user = db.relationship('User', foreign_keys=[created_by], backref=db.backref('invoices_created', lazy=True))

class DeviceType(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    value_usd = db.Column(db.Float, default=0.0, nullable=False)
    # Valores específicos para MEIO/PONTA (quando mid_end_enabled no mapa).
    # Se NULL, cai no value_usd padrão.
    value_meio_usd = db.Column(db.Float, nullable=True)
    value_ponta_usd = db.Column(db.Float, nullable=True)
    # Ribbon: quando True, o lancamento usa fitas (ribbon_count) em vez de fusoes.
    is_ribbon = db.Column(db.Boolean, default=False, nullable=False)
    ribbon_price_usd = db.Column(db.Float, nullable=True)  # valor fixo por fita
    # Codigos de cobranca do dispositivo
    # billing_code       = codigo padrao (quando nao ha MEIO/PONTA ou sem distincao)
    # billing_code_meio  = codigo especifico quando lancamento for MEIO
    # billing_code_ponta = codigo especifico quando lancamento for PONTA
    billing_code       = db.Column(db.String(30), nullable=True)
    billing_code_meio  = db.Column(db.String(30), nullable=True)
    billing_code_ponta = db.Column(db.String(30), nullable=True)
    company = db.Column(db.String(120), nullable=True)  # se None = valor padrao
    project_id = db.Column(db.Integer, db.ForeignKey('project.id'), nullable=True, index=True)
    project = db.relationship('Project', backref=db.backref('device_types', lazy=True))

class SpliceTier(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    min_splices = db.Column(db.Integer, nullable=False)
    max_splices = db.Column(db.Integer, nullable=True)
    price_per_splice_usd = db.Column(db.Float, default=0.0, nullable=False)
    # Codigo de cobranca (opcional): codigo unico por faixa de fusoes (ex: FS01)
    code_splice = db.Column(db.String(30), nullable=True)
    company = db.Column(db.String(120), nullable=True)
    project_id = db.Column(db.Integer, db.ForeignKey('project.id'), nullable=True, index=True)
    project = db.relationship('Project', backref=db.backref('splice_tiers', lazy=True))


class HourlyRate(db.Model):
    """Valor por hora configurado por projeto/empresa para lancamentos de horas."""
    id = db.Column(db.Integer, primary_key=True)
    rate_usd = db.Column(db.Float, default=0.0, nullable=False)
    billing_code = db.Column(db.String(30), nullable=True)
    description = db.Column(db.String(120), nullable=True)  # ex: "Reparo / Conserto"
    company = db.Column(db.String(120), nullable=True)
    project_id = db.Column(db.Integer, db.ForeignKey('project.id'), nullable=True, index=True)
    project = db.relationship('Project', backref=db.backref('hourly_rates', lazy=True))


class HourRecord(db.Model):
    """Lancamento de horas trabalhadas (conserto, reparo, etc.)."""
    id = db.Column(db.Integer, primary_key=True)
    company = db.Column(db.String(120), nullable=True)
    project_id = db.Column(db.Integer, db.ForeignKey('project.id'), nullable=True, index=True)
    project = db.relationship('Project', backref=db.backref('hour_records', lazy=True))
    splicer = db.Column(db.String(120), nullable=True)
    created_date = db.Column(db.DateTime, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    hours = db.Column(db.Float, nullable=False, default=0.0)
    description = db.Column(db.String(255), nullable=True)
    rate_usd = db.Column(db.Float, default=0.0)
    total_usd = db.Column(db.Float, default=0.0)
    billing_code = db.Column(db.String(30), nullable=True)
    map_name = db.Column(db.String(200), nullable=True)


# Associação entre mapas interativos e splicers que podem acessá-los.
map_splicer_access = db.Table(
    "map_splicer_access",
    db.Column("map_id", db.Integer, db.ForeignKey("company_map.id"), primary_key=True),
    db.Column("user_id", db.Integer, db.ForeignKey("user.id"), primary_key=True),
)
class CompanyMap(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    company = db.Column(db.String(120), nullable=False, index=True)
    name = db.Column(db.String(200), nullable=False)
    project_id = db.Column(db.Integer, db.ForeignKey('project.id'), nullable=True, index=True)
    project = db.relationship('Project', backref=db.backref('maps', lazy=True))

    # Opcional (por mapa): habilita seleção MEIO/PONTA no lançamento.
    # Só aparece no Entry se mid_end_enabled=1.
    mid_end_enabled = db.Column(db.Boolean, nullable=False, default=False)
    included_splices_meio = db.Column(db.Integer, nullable=True)   # inclusas quando MEIO
    included_splices_ponta = db.Column(db.Integer, nullable=True)  # inclusas quando PONTA

    # Cores personalizadas por seção (JSON: { "SEC A": "#ff0000", ... })
    section_colors_json = db.Column(db.Text, nullable=True)

    # Splicers com permissão explícita para acessar esse mapa no módulo interativo.
    allowed_splicers = db.relationship(
        "User",
        secondary=map_splicer_access,
        back_populates="maps_with_access",
        lazy="select",
    )


class Record(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    map = db.Column(db.String(200))
    type = db.Column(db.String(120))
    splices = db.Column(db.Integer)
    device = db.Column(db.String(120))
    splicer = db.Column(db.String(120))
    created_date = db.Column(db.DateTime, nullable=True)
    company = db.Column(db.String(120), nullable=True)
    project_id = db.Column(db.Integer, db.ForeignKey('project.id'), nullable=True, index=True)
    project = db.relationship('Project', backref=db.backref('records', lazy=True))

    # Coordenadas (opcional): para mapa interativo
    latitude = db.Column(db.Float, nullable=True)
    longitude = db.Column(db.Float, nullable=True)

    # Informações adicionais do dispositivo (fibra usada, porta, etc.)
    device_info = db.Column(db.String(255), nullable=True)
    ft_in = db.Column(db.String(120), nullable=True)
    ft_out = db.Column(db.String(120), nullable=True)
    can_cable_count = db.Column(db.Integer, nullable=True)
    can_cables_json = db.Column(db.Text, nullable=True)

    # Snapshot (opcional): se o mapa estiver configurado com MEIO/PONTA,
    # salvamos o tipo escolhido e quantas fusões inclusas foram aplicadas.
    map_role = db.Column(db.String(10), nullable=True)  # 'MEIO' / 'PONTA'
    included_splices_applied = db.Column(db.Integer, nullable=True)

    # Seção lógica do mapa (grupo alimentado por um splitter/ramal específico)
    section = db.Column(db.String(120), nullable=True)
    geo_address = db.Column(db.String(255), nullable=True)
    pon_name = db.Column(db.String(120), nullable=True)
    splitter_name = db.Column(db.String(120), nullable=True)
    source_from = db.Column(db.String(120), nullable=True)
    source_out = db.Column(db.String(120), nullable=True)
    ote_label = db.Column(db.String(120), nullable=True)
    port_label = db.Column(db.String(120), nullable=True)

    # Subida (placed): marcado quando o dispositivo físico foi instalado no campo
    is_placed = db.Column(db.Boolean, default=False)
    placed_by = db.Column(db.String(120), nullable=True)   # nome de quem subiu
    placed_at = db.Column(db.DateTime, nullable=True)       # quando foi subido

    price_splices_usd = db.Column(db.Float, default=0.0)
    price_device_usd = db.Column(db.Float, default=0.0)
    total_usd = db.Column(db.Float, default=0.0)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    # Ribbon: quando o dispositivo e do tipo ribbon, armazena a qtd de fitas
    ribbon_count = db.Column(db.Integer, nullable=True)
    # Ativo/Inativo: False = dispositivo desativado (pin vermelho no mapa)
    # NULL ou True = ativo (comportamento padrao)
    is_active = db.Column(db.Boolean, nullable=True, default=None)
    # Codigos de cobranca calculados automaticamente no lancamento
    # Armazenados como JSON: ["FS01", "FS15"] etc.
    billing_codes_json = db.Column(db.Text, nullable=True)

    # Testes (níveis por fusão, armazenados como CSV) e flag de concluído
    test_levels = db.Column(db.Text, nullable=True)
    test_done = db.Column(db.Boolean, default=False)
    test_date = db.Column(db.DateTime, nullable=True)


class RecordPhoto(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    record_id = db.Column(db.Integer, db.ForeignKey('record.id'), nullable=False)
    filename = db.Column(db.String(255), nullable=False)
    content_type = db.Column(db.String(100))
    data = deferred(db.Column(db.LargeBinary, nullable=False))  # may be empty bytes when stored in R2
    r2_key = db.Column(db.String(512))
    r2_thumb_key = db.Column(db.String(512))
    size_bytes = db.Column(db.Integer)
    thumb_data = deferred(db.Column(db.LargeBinary))
    thumb_content_type = db.Column(db.String(100))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    is_test = db.Column(db.Boolean, default=False)

    record = db.relationship('Record', backref=db.backref('photos', lazy=True))


class Expense(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    user = db.relationship("User", foreign_keys=[user_id], backref=db.backref("expenses", lazy=True))
    description = db.Column(db.String(255), nullable=False)
    category = db.Column(db.String(120), nullable=True)
    amount = db.Column(db.Float, nullable=False, default=0.0)
    date = db.Column(db.Date, nullable=False, default=date.today)

    paid = db.Column(db.Boolean, nullable=False, default=False)
    paid_at = db.Column(db.DateTime, nullable=True)
    paid_by = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)


    paid_by_user = db.relationship("User", foreign_keys=[paid_by], backref=db.backref("expenses_paid", lazy=True))
    created_at = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)




class Payroll(db.Model):
    """Folha de pagamento gerada pelo admin para um splicer em um período."""
    id = db.Column(db.Integer, primary_key=True)
    # Splicer que vai receber
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    user = db.relationship("User", foreign_keys=[user_id], backref=db.backref("payrolls", lazy=True))
    # Período coberto
    start_date = db.Column(db.Date, nullable=False)
    end_date = db.Column(db.Date, nullable=False)
    # Projeto (opcional — pode ser global ou por projeto)
    project_id = db.Column(db.Integer, db.ForeignKey("project.id"), nullable=True)
    project = db.relationship("Project", backref=db.backref("payrolls", lazy=True))
    company = db.Column(db.String(120), nullable=True)
    # Totais calculados no momento da geração
    total_records = db.Column(db.Integer, default=0)
    total_splices = db.Column(db.Integer, default=0)
    total_amount_usd = db.Column(db.Float, default=0.0)
    # Prazo de pagamento em dias (copiado do projeto no momento da criação)
    payment_days = db.Column(db.Integer, default=30)
    # Data limite para pagamento
    due_date = db.Column(db.Date, nullable=True)
    # Status
    status = db.Column(db.String(20), default="pending")  # pending | paid | cancelled
    paid_at = db.Column(db.DateTime, nullable=True)
    paid_by = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)
    paid_by_user = db.relationship("User", foreign_keys=[paid_by], backref=db.backref("payrolls_paid", lazy=True))
    # Quem criou
    created_by = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)
    created_by_user = db.relationship("User", foreign_keys=[created_by], backref=db.backref("payrolls_created", lazy=True))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    notes = db.Column(db.Text, nullable=True)
    # Custo real pago ao splicer (calculado com a modalidade dele)
    splicer_cost_usd = db.Column(db.Float, nullable=True)
    # Modalidade de pagamento usada no fechamento
    pricing_id = db.Column(db.Integer, db.ForeignKey("splicer_pricing.id"), nullable=True)
    pricing = db.relationship("SplicerPricing", foreign_keys=[pricing_id])
    plan_name = db.Column(db.String(120), nullable=True)  # snapshot do label da tabela

class SplicerPricing(db.Model):
    """Tabela de preços compartilhada por um ou mais splicers dentro de um projeto.

    Vários splicers podem ser atribuídos à mesma tabela via SplicerPricingAssignment.
    Cada splicer só pode ter uma tabela por projeto (enforced no assignment).
    """
    __tablename__ = "splicer_pricing"
    id = db.Column(db.Integer, primary_key=True)
    project_id = db.Column(db.Integer, db.ForeignKey("project.id"), nullable=False, index=True)
    project = db.relationship("Project", backref=db.backref("splicer_pricings", lazy=True))
    # Nome/rótulo da tabela (ex: "Padrão", "Sênior", "João Silva")
    label = db.Column(db.String(120), nullable=False)
    # Fusões inclusas por lançamento (sem cobrança)
    included_splices = db.Column(db.Integer, nullable=False, default=0)
    # Preços por dispositivo: {"OTE": 5.00, "CAN": 8.00, ...}
    device_prices_json = db.Column(db.Text, nullable=True)
    # Faixas de fusão: [{"min": 1, "max": null, "price": 2.50}, ...]
    tiers_json = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    # Splicers atribuídos a esta tabela
    assignments = db.relationship("SplicerPricingAssignment", backref="pricing", lazy=True,
                                  cascade="all, delete-orphan")

    def get_tiers(self):
        try:
            return json.loads(self.tiers_json or "[]") or []
        except Exception:
            return []

    def get_device_prices(self) -> dict:
        try:
            return json.loads(self.device_prices_json or "{}") or {}
        except Exception:
            return {}

    def device_price_for(self, device_name: str) -> float:
        if not device_name:
            return 0.0
        prices = self.get_device_prices()
        key = device_name.strip().upper()
        for k, v in prices.items():
            if k.strip().upper() == key:
                return float(v or 0)
        return 0.0

    def price_for_splices(self, total_splices: int) -> float:
        included = int(self.included_splices or 0)
        charge = max(int(total_splices or 0) - included, 0)
        if charge == 0:
            return 0.0
        tiers = self.get_tiers()
        price_per = 0.0
        for t in sorted(tiers, key=lambda x: x.get("min", 0)):
            t_min = int(t.get("min", 0))
            t_max = t.get("max")
            if total_splices >= t_min and (t_max is None or total_splices <= int(t_max)):
                price_per = float(t.get("price", 0))
                break
        return charge * price_per

    def total_for_record(self, splices: int, device_name: str = "") -> float:
        return self.price_for_splices(splices) + self.device_price_for(device_name or "")


class SplicerPricingAssignment(db.Model):
    """Associa um splicer a uma tabela de preços dentro de um projeto.
    Um splicer só pode estar em UMA tabela por projeto.
    """
    __tablename__ = "splicer_pricing_assignment"
    id = db.Column(db.Integer, primary_key=True)
    pricing_id = db.Column(db.Integer, db.ForeignKey("splicer_pricing.id"), nullable=False, index=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False, index=True)
    user = db.relationship("User", backref=db.backref("pricing_assignments", lazy=True))
    project_id = db.Column(db.Integer, db.ForeignKey("project.id"), nullable=False, index=True)

    __table_args__ = (
        db.UniqueConstraint("user_id", "project_id", name="uq_spa_user_project"),
    )

# --------- User loader ---------
@login_manager.user_loader
def load_user(user_id: str):
    try:
        return User.query.get(int(user_id))
    except Exception:
        return None

# Garante migração antes da primeira requisição (necessário no Render/Gunicorn)
_migration_done = False

@app.before_request
def run_migration_once():
    global _migration_done
    if _migration_done:
        return
    _migration_done = True
    try:
        from sqlalchemy import inspect as sa_inspect, text as sa_text
        insp = sa_inspect(db.engine)

        def _ensure(table, col, typ):
            try:
                existing = [c["name"] for c in insp.get_columns(table)]
                if col not in existing:
                    db.session.execute(sa_text(f'ALTER TABLE "{table}" ADD COLUMN {col} {typ}'))
                    db.session.commit()
            except Exception:
                db.session.rollback()

        _ensure("user", "is_active",           "BOOLEAN DEFAULT TRUE")
        _ensure("user", "can_access_expenses",  "BOOLEAN DEFAULT FALSE")
        _ensure("user", "can_view_values",      "BOOLEAN DEFAULT TRUE")
        _ensure("project", "payment_days",      "INTEGER DEFAULT 30")

        # Preenche valores padrão para registros existentes (Postgres)
        try:
            db.session.execute(sa_text('UPDATE "user" SET is_active = TRUE WHERE is_active IS NULL'))
            db.session.execute(sa_text('UPDATE "user" SET can_access_expenses = FALSE WHERE can_access_expenses IS NULL'))
            db.session.execute(sa_text('UPDATE "user" SET can_view_values = TRUE WHERE can_view_values IS NULL'))
            db.session.commit()
        except Exception:
            db.session.rollback()
    except Exception:
        pass

def ensure_db_schema():
    """Best-effort schema updates (no destructive changes)."""
    try:
        insp = inspect(db.engine)
        cols = {c["name"] for c in insp.get_columns("record_photo")}
    except Exception:
        cols = set()

    needed = {
        "r2_key": "VARCHAR(512)",
        "r2_thumb_key": "VARCHAR(512)",
        "size_bytes": "INTEGER",
    }

    dialect = getattr(db.engine.dialect, "name", "").lower()
    for col, coltype in needed.items():
        if col in cols:
            continue
        try:
            if dialect == "postgresql":
                db.session.execute(text(f'ALTER TABLE record_photo ADD COLUMN IF NOT EXISTS {col} {coltype}'))
            else:
                # SQLite: no IF NOT EXISTS on older versions; we already checked.
                db.session.execute(text(f'ALTER TABLE record_photo ADD COLUMN {col} {coltype}'))
            db.session.commit()
        except Exception:
            db.session.rollback()

# --------- DB init & migrations simples ---------
with app.app_context():
    db.create_all()
    ensure_db_schema()

    # ── Migração crítica das colunas novas do user ──
    # DEVE rodar ANTES de qualquer User.query para evitar
    # "column does not exist" no PostgreSQL (Render/Gunicorn)
    try:
        _insp = inspect(db.engine)
        _existing_user_cols = [c["name"] for c in _insp.get_columns("user")]

        def _ensure_user_col(col, typ, default_sql):
            if col not in _existing_user_cols:
                try:
                    db.session.execute(text(f'ALTER TABLE "user" ADD COLUMN {col} {typ}'))
                    db.session.execute(text(f'UPDATE "user" SET {col} = {default_sql} WHERE {col} IS NULL'))
                    db.session.commit()
                except Exception:
                    db.session.rollback()

        _ensure_user_col("is_active",          "BOOLEAN", "TRUE")
        _ensure_user_col("can_access_expenses", "BOOLEAN", "FALSE")
        _ensure_user_col("can_view_values",     "BOOLEAN", "TRUE")
    except Exception:
        pass

    # garante usuário padrão
    if not User.query.filter_by(username="admin").first():
        db.session.add(User(username="admin", password="admin", is_admin=True, splicer_name="ADMIN"))
        db.session.commit()

    # migração simples de colunas (funciona tanto em SQLite quanto em Postgres)
    def ensure(table, col, typ):
        """Garante que uma coluna exista na tabela informada."""
        inspector = inspect(db.engine)
        existing = [c["name"] for c in inspector.get_columns(table)]
        if col not in existing:
            # Em Postgres, usar aspas duplas no nome da tabela evita problemas de case
            db.session.execute(text(f'ALTER TABLE "{table}" ADD COLUMN {col} {typ}'))
            db.session.commit()

    ensure("record", "company", "VARCHAR(120)")
    ensure("device_type", "company", "VARCHAR(120)")
    ensure("splice_tier", "company", "VARCHAR(120)")
    ensure("record", "project_id", "INTEGER")
    ensure("record", "map_role", "VARCHAR(10)")
    ensure("record", "included_splices_applied", "INTEGER")
    ensure("record", "test_levels", "TEXT")
    ensure("record", "test_done", "BOOLEAN")
    ensure("record", "test_date", "TIMESTAMP")
    ensure("record", "latitude", "DOUBLE PRECISION")
    ensure("record", "longitude", "DOUBLE PRECISION")
    ensure("record", "device_info", "VARCHAR(255)")
    ensure("record", "ft_in", "VARCHAR(120)")
    ensure("record", "ft_out", "VARCHAR(120)")
    ensure("record", "can_cable_count", "INTEGER")
    ensure("record", "can_cables_json", "TEXT")
    ensure("record", "section", "VARCHAR(120)")
    ensure("record", "geo_address", "VARCHAR(255)")
    ensure("record", "pon_name", "VARCHAR(120)")
    ensure("record", "splitter_name", "VARCHAR(120)")
    ensure("record", "source_from", "VARCHAR(120)")
    ensure("record", "source_out", "VARCHAR(120)")
    ensure("record", "ote_label", "VARCHAR(120)")
    ensure("record", "port_label", "VARCHAR(120)")
    ensure("record", "is_placed", "BOOLEAN")
    ensure("record", "placed_by", "VARCHAR(120)")
    ensure("record", "placed_at", "TIMESTAMP")
    ensure("device_type", "project_id", "INTEGER")
    ensure("device_type", "value_meio_usd", "DOUBLE PRECISION")
    ensure("device_type", "value_ponta_usd", "DOUBLE PRECISION")
    ensure("device_type", "is_ribbon", "BOOLEAN")
    ensure("device_type", "ribbon_price_usd", "DOUBLE PRECISION")
    ensure("record", "ribbon_count", "INTEGER")
    ensure("record", "is_active", "BOOLEAN")
    ensure("record", "billing_codes_json", "TEXT")
    ensure("device_type", "billing_code", "VARCHAR(30)")
    ensure("device_type", "billing_code_meio", "VARCHAR(30)")
    ensure("device_type", "billing_code_ponta", "VARCHAR(30)")
    ensure("splice_tier", "code_splice", "VARCHAR(30)")
    # HourlyRate e HourRecord sao criados pelo db.create_all() automaticamente
    # mas caso as tabelas ja existam precisamos garantir as colunas
    try:
        db.create_all()
    except Exception:
        pass
    ensure("hour_record", "map_name", "VARCHAR(200)")
    # Garante FALSE como padrão para is_ribbon em registros existentes (NULL -> FALSE)
    try:
        db.session.execute(text('UPDATE device_type SET is_ribbon = FALSE WHERE is_ribbon IS NULL'))
        db.session.commit()
    except Exception:
        db.session.rollback()

    # Recalcula precos de lancamentos ribbon com total zerado
    try:
        ribbon_types = DeviceType.query.filter(DeviceType.is_ribbon == True).all()
        ribbon_names = {dt.name.lower(): dt for dt in ribbon_types}
        if ribbon_names:
            broken = Record.query.filter(Record.total_usd == 0.0).all()
            changed = False
            for rec in broken:
                rtype = (rec.type or "").lower().strip()
                dt = ribbon_names.get(rtype)
                if not dt:
                    continue
                rcount = rec.ribbon_count or (int(rec.splices or 0) if (rec.splices or 0) > 0 else None)
                if rcount and dt.ribbon_price_usd:
                    price_fitas = float(rcount) * float(dt.ribbon_price_usd)
                    price_enc = float(dt.value_usd or 0.0)
                    rec.price_splices_usd = price_fitas
                    rec.price_device_usd = price_enc
                    rec.total_usd = price_fitas + price_enc
                    rec.ribbon_count = rcount
                    rec.splices = 0
                    changed = True
            if changed:
                db.session.commit()
    except Exception:
        db.session.rollback()
    ensure("splice_tier", "project_id", "INTEGER")
    ensure("company_map", "project_id", "INTEGER")
    ensure("company_map", "mid_end_enabled", "BOOLEAN")
    ensure("company_map", "included_splices_meio", "INTEGER")
    ensure("company_map", "included_splices_ponta", "INTEGER")
    ensure("company_map", "section_colors_json", "TEXT")
    ensure("record_photo", "thumb_data", "BYTEA")
    ensure("record_photo", "thumb_content_type", "VARCHAR(100)")
    ensure("record_photo", "is_test", "BOOLEAN")
    ensure("project", "company", "VARCHAR(120)")
    ensure("project", "name", "VARCHAR(200)")
    ensure("project", "included_splices", "INTEGER")
    ensure("company_config", "invoice_address", "TEXT")
    ensure("system_config", "geoapify_api_key", "VARCHAR(255)")
    ensure("system_config", "board_header", "VARCHAR(120)")
    ensure("user", "is_admin", "BOOLEAN")
    ensure("user", "splicer_name", "VARCHAR(120)")
    ensure("user", "is_company_owner", "BOOLEAN")
    ensure("user", "is_active", "BOOLEAN")
    ensure("user", "can_access_expenses", "BOOLEAN")
    ensure("user", "can_view_values", "BOOLEAN")
    ensure("project", "payment_days", "INTEGER")
    ensure("expense", "paid", "BOOLEAN")
    ensure("expense", "paid_at", "TIMESTAMP")
    ensure("expense", "paid_by", "INTEGER")
    ensure("invoice", "created_by", "INTEGER")
    ensure("invoice", "pdf_filename", "VARCHAR(255)")
    ensure("invoice", "pdf_content_type", "VARCHAR(100)")
    ensure("invoice", "pdf_data", "BYTEA" if 'postgres' in db.engine.name else "BLOB")

    # ── Tabelas de preços de splicers ──
    try:
        _insp2 = inspect(db.engine)
        _tables_now = _insp2.get_table_names()
        # Garante coluna pricing_id no payroll (troca de plan_id antigo)
        if "payroll" in _tables_now:
            _pr_cols = {c["name"] for c in _insp2.get_columns("payroll")}
            for _col, _typ in [
                ("splicer_cost_usd", "DOUBLE PRECISION"),
                ("pricing_id", "INTEGER"),
                ("plan_name", "VARCHAR(120)"),
            ]:
                if _col not in _pr_cols:
                    try:
                        db.session.execute(text(f'ALTER TABLE "payroll" ADD COLUMN {_col} {_typ}'))
                        db.session.commit()
                    except Exception:
                        db.session.rollback()
    except Exception:
        pass

    # garante valor padrão para despesas antigas
    try:
        db.session.execute(text('UPDATE "expense" SET paid = 0 WHERE paid IS NULL'))
        db.session.commit()
    except Exception:
        pass

    # garante valores padrão para novas colunas de permissão de usuários
    try:
        db.session.execute(text('UPDATE "user" SET is_active = 1 WHERE is_active IS NULL'))
        db.session.execute(text('UPDATE "user" SET can_access_expenses = 0 WHERE can_access_expenses IS NULL'))
        db.session.execute(text('UPDATE "user" SET can_view_values = 1 WHERE can_view_values IS NULL'))
        db.session.commit()
    except Exception:
        pass

# --------- Login ---------
@app.route("/login", methods=["GET", "POST"])
def login():
    if current_user.is_authenticated:
        return redirect(url_for("index"))

    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        password = (request.form.get("password") or "").strip()

        user = User.query.filter_by(username=username).first()
        if user and user.password == password:
            if not getattr(user, 'is_active', True):
                flash("Usuário desativado. Contate o administrador.", "danger")
                return render_template("login.html")
            login_user(user)
            flash("Login realizado com sucesso.", "success")
            next_page = request.args.get("next")
            return redirect(next_page or url_for("index"))

        flash("Usuário ou senha inválidos.", "danger")

    return render_template("login.html")

# --------- Helpers de preço ---------
def included_splices_for(company: str | None, project_id: int | None = None) -> int:
    """Quantas fusões são inclusas para essa empresa/projeto."""
    # Prioridade: projeto -> empresa -> padrão antigo
    if project_id:
        pr = Project.query.get(int(project_id))
        if pr and pr.included_splices is not None:
            return int(pr.included_splices or 0)

    if company:
        cfg = CompanyConfig.query.filter_by(name=company).first()
        if cfg:
            return int(cfg.included_splices or 0)

    return 1

def _get_device_type(name: str, company: str | None, project_id: int | None = None):
    """Busca o DeviceType com prioridade: projeto > empresa > global."""
    if not name:
        return None
    q = DeviceType.query.filter(DeviceType.name.ilike(name))
    if project_id:
        q = q.filter(or_(DeviceType.project_id == project_id, DeviceType.project_id.is_(None)))
    else:
        q = q.filter(DeviceType.project_id.is_(None))
    if company:
        q = q.filter(or_(DeviceType.company == company, DeviceType.company.is_(None)))
        order_clauses = []
        if project_id:
            order_clauses.append(case((DeviceType.project_id == project_id, 0), else_=1))
        order_clauses.append(case((DeviceType.company == company, 0), else_=1))
        return q.order_by(*order_clauses).first()
    return q.first()


def device_is_ribbon(name: str, company: str | None, project_id: int | None = None) -> tuple:
    """Retorna (is_ribbon, ribbon_price_usd) para o dispositivo."""
    dt = _get_device_type(name, company, project_id)
    if dt and bool(getattr(dt, "is_ribbon", False)):
        return True, float(getattr(dt, "ribbon_price_usd", None) or 0.0)
    return False, 0.0


def device_value_for(name: str, company: str | None, project_id: int | None = None, map_role: str | None = None) -> float:
    """Retorna o valor do dispositivo considerando a regra MEIO/PONTA.

    Quando map_role for 'MEIO' ou 'PONTA' e o DeviceType tiver os campos
    value_meio_usd / value_ponta_usd preenchidos, usa esses valores.
    Caso contrario, cai no value_usd padrao.
    Dispositivos ribbon nao tem valor de dispositivo fixo (o preco e por fita).
    """
    dt = _get_device_type(name, company, project_id)
    if not dt:
        return 0.0
    # Ribbon: cobra value_usd normalmente (enclosure/caixa)
    # o valor por fita e calculado separadamente em compute_prices
    # Aplica regra MEIO/PONTA se disponivel
    role = (map_role or "").strip().upper()
    if role == "MEIO" and dt.value_meio_usd is not None:
        return float(dt.value_meio_usd)
    if role == "PONTA" and dt.value_ponta_usd is not None:
        return float(dt.value_ponta_usd)
    return float(dt.value_usd)

def tier_price_for(count: int, company: str | None, project_id: int | None = None) -> float:
    """Retorna o $/fusão da faixa respeitando prioridade:

    1. Faixa específica do PROJETO (independente da empresa).
    2. Se não houver faixa de projeto, usa faixa da EMPRESA (sem projeto).
    3. Se ainda assim não tiver, usa faixa GLOBAL (sem empresa e sem projeto).

    Em todos os casos, escolhe a faixa cujo intervalo contenha `count`:
        min_splices <= count <= max_splices (ou max_splices is NULL)
    e, dentre elas, pega a de maior `min_splices` (faixa mais específica).
    """
    from sqlalchemy import or_ as _or

    def _best_for(base_query):
        q = (
            base_query.filter(SpliceTier.min_splices <= count)
            .filter(_or(SpliceTier.max_splices == None, SpliceTier.max_splices >= count))
            .order_by(SpliceTier.min_splices.desc())
        )
        return q.first()

    # 1) Faixas específicas do PROJETO (se houver project_id)
    if project_id:
        tier_proj = _best_for(SpliceTier.query.filter(SpliceTier.project_id == project_id))
        if tier_proj:
            return float(tier_proj.price_per_splice_usd or 0.0)

    # 2) Faixas por EMPRESA (sem projeto)
    if company:
        tier_company = _best_for(
            SpliceTier.query.filter(
                SpliceTier.company == company,
                SpliceTier.project_id.is_(None),
            )
        )
        if tier_company:
            return float(tier_company.price_per_splice_usd or 0.0)

    # 3) Faixas GLOBAIS (sem empresa e sem projeto)
    tier_global = _best_for(
        SpliceTier.query.filter(
            SpliceTier.company.is_(None),
            SpliceTier.project_id.is_(None),
        )
    )
    if tier_global:
        return float(tier_global.price_per_splice_usd or 0.0)

    # Se não houver nenhuma faixa cadastrada
    return 0.0


def resolve_included_override(company: str | None, project_id: int | None, map_obj, map_val: str | None, map_role: str | None):
    """
    Resolve quantas fusões inclusas devem ser aplicadas para este lançamento,
    levando em conta configuração MEIO/PONTA do mapa.

    Retorna (included_override, included_applied, map_cfg_obj).
    """
    map_cfg = map_obj
    # tenta buscar pelo nome, garantindo que sempre encontramos o CompanyMap correto
    if not map_cfg and map_val and company:
        q = CompanyMap.query.filter_by(company=company, name=map_val)
        if project_id is not None:
            q = q.filter_by(project_id=project_id)
        map_cfg = q.first()
        if not map_cfg:
            map_cfg = CompanyMap.query.filter_by(company=company, name=map_val, project_id=None).first()

    included_override = None
    included_applied = None

    if map_cfg and bool(getattr(map_cfg, "mid_end_enabled", False)) and map_role in ("MEIO", "PONTA"):
        if map_role == "MEIO":
            included_override = int(getattr(map_cfg, "included_splices_meio", 0) or 0)
        else:
            included_override = int(getattr(map_cfg, "included_splices_ponta", 0) or 0)
        included_applied = included_override

    return included_override, included_applied, map_cfg




def parse_timestamp_block_from_text(text: str):
    """Interpreta o bloco de 4 linhas usado no Timestamp.

    Espera algo como:
        MAP: RN51E
        DEVICE: FT10 - P1
        FUSION: 12
        T: P

    Retorna dict com: map_name, device_name, splices (int), map_role ('MEIO' ou 'PONTA'),
    ou None se não conseguir interpretar.
    """
    if not text:
        return None

    # Normaliza quebras de linha
    lines = [ln.strip() for ln in str(text).replace("\r", "\n").split("\n") if ln.strip()]
    map_name = None
    device_name = None
    splices_val = None
    map_role = None

    for ln in lines:
        upper = ln.upper()
        if upper.startswith("MAP"):
            parts = ln.split(":", 1)
            if len(parts) == 2:
                map_name = parts[1].strip()
        elif upper.startswith("DEVICE") or upper.startswith("DISPOSITIVO"):
            parts = ln.split(":", 1)
            if len(parts) == 2:
                device_name = parts[1].strip()
        elif "FUS" in upper:  # FUSION / FUSÕES / FUSAO
            import re as _re
            m = _re.search(r"(\d+)", ln)
            if m:
                try:
                    splices_val = int(m.group(1))
                except Exception:
                    splices_val = None
        elif upper.startswith("T"):
            parts = ln.split(":", 1)
            if len(parts) == 2:
                val = parts[1].strip().upper()
                if val in ("P", "M"):
                    map_role = "PONTA" if val == "P" else "MEIO"

    if not (map_name and device_name and splices_val is not None and map_role):
        return None

    return {
        "map_name": map_name,
        "device_name": device_name,
        "splices": splices_val,
        "map_role": map_role,
    }


def extract_timestamp_fields_with_ai(image_bytes: bytes):
    """Usa Claude (Anthropic) para ler dados de uma foto Timemark de splice box.

    Retorna (dict, error_message). Se der certo, error_message = None e o dict contém:
    map_name, device_name, splices (int), map_role ('MEIO' ou 'PONTA'),
    ft_in (str|None), ft_out (str|None), gps (str|None), photo_datetime (str|None).

    MEIO  = tem IN e OUT (enclosure do meio da rota)
    PONTA = tem só IN (enclosure terminal / ponta da rota)
    """
    if not image_bytes:
        return None, "Imagem vazia"

    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        # Fallback: tenta OpenAI se configurado
        api_key_oai = os.environ.get("OPENAI_API_KEY")
        if not api_key_oai:
            return None, "ANTHROPIC_API_KEY não configurada no servidor."
        # --- Fallback OpenAI (legado) ---
        model = os.environ.get("OPENAI_VISION_MODEL", "gpt-4o-mini")
        try:
            img_b64 = base64.b64encode(image_bytes).decode("ascii")
            prompt = (
                "Você é um assistente que lê fotos de splice box (fibra óptica) com carimbo Timemark.\n"
                "Retorne APENAS JSON válido, sem explicações:\n"
                "{\"map_name\":\"...\",\"device_name\":\"...\",\"splices\":0,\"map_role\":\"PONTA\","
                "\"ft_in\":null,\"ft_out\":null,\"gps\":null,\"photo_datetime\":null}\n"
                "map_role: 'PONTA' se só tiver IN, 'MEIO' se tiver IN e OUT.\n"
                "ft_in/ft_out: números de feet visíveis na splice tray (ex: '48224'). null se não visível.\n"
                "gps: 'lat,lon' se visível no carimbo. photo_datetime: data/hora do carimbo."
            )
            payload = {
                "model": model,
                "messages": [{"role": "user", "content": [
                    {"type": "text", "text": prompt},
                    {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{img_b64}"}},
                ]}],
                "max_tokens": 400, "temperature": 0.0,
            }
            headers = {"Authorization": f"Bearer {api_key_oai}", "Content-Type": "application/json"}
            resp = requests.post("https://api.openai.com/v1/chat/completions", headers=headers,
                                 data=json.dumps(payload), timeout=30)
            if resp.status_code != 200:
                return None, f"Erro HTTP {resp.status_code} ao chamar OpenAI."
            content = resp.json().get("choices", [{}])[0].get("message", {}).get("content", "").strip()
            if content.startswith("```"):
                content = content.strip("`\n ")
                if content.lower().startswith("json"):
                    content = content[4:].lstrip()
            parsed = json.loads(content)
        except Exception as e:
            return None, f"Erro fallback OpenAI: {e}"
    else:
        # --- Claude (Anthropic) ---
        try:
            img_b64 = base64.b64encode(image_bytes).decode("ascii")

            prompt = (
                "Você é um especialista em leitura de fotos de splice box de fibra óptica (field terminals FTTX).\n"
                "As fotos são tiradas com o app Timemark que adiciona carimbo com nome, data, endereço e GPS.\n\n"
                "=== REGRAS DE LEITURA ===\n\n"
                "DEVICE (nome do dispositivo):\n"
                "  - Leia da ETIQUETA BRANCA dentro da splice box (ex: FT33, FT102, FT62)\n"
                "  - Confirmação: o mesmo nome aparece no carimbo Timemark embaixo da foto\n"
                "  - Use SEMPRE a etiqueta dentro da box como fonte primária\n\n"
                "SPLICES (número de fusões):\n"
                "  - NÃO conte fibras físicas nem conectores\n"
                "  - Leia a ETIQUETA NUMERADA na splice tray (ex: #19-20, #1-12, #1-24)\n"
                "  - Calcule: fim - início + 1 (ex: #19-20 = 20-19+1 = 2 fusões)\n"
                "  - Se houver múltiplas etiquetas, some todas (ex: #1-12 e #13-24 = 24 fusões)\n"
                "  - Se não encontrar etiqueta numerada, conte os splices visíveis na tray\n\n"
                "MAP_ROLE (tipo do dispositivo):\n"
                "  - PONTA = só tem cabo(s) de ENTRADA. Terminal de rede, cliente final.\n"
                "  - MEIO = tem cabo de ENTRADA e cabo de SAÍDA. Passagem/repetidor de rota.\n"
                "  - Observe os cabos externos entrando na caixa para determinar\n\n"
                "FT_IN / FT_OUT:\n"
                "  - Números escritos na splice box ou nos cabos (ex: 'IN 48224', 'OUT 48224')\n"
                "  - null se não visível. ft_out sempre null se PONTA.\n\n"
                "GPS:\n"
                "  - Leia do carimbo Timemark: latitude e longitude (ex: '38.823849,-82.229083')\n"
                "  - Converta graus com símbolo N/W para decimal: W é negativo\n\n"
                "DATETIME:\n"
                "  - Data e hora do carimbo Timemark (ex: '2026-06-06 13:13')\n\n"
                "=== EXEMPLO desta foto ===\n"
                "Etiqueta dentro da box: FT33, etiqueta tray: #19-20 → splices=2\n"
                "Carimbo: FT33, 38.823849°N 82.229083°W → gps='38.823849,-82.229083'\n\n"
                "Retorne APENAS JSON válido, sem explicações, sem markdown:\n"
                "{\"device_name\":\"FT33\",\"splices\":2,\"map_role\":\"PONTA\","
                "\"ft_in\":null,\"ft_out\":null,\"gps\":\"38.823849,-82.229083\",\"photo_datetime\":\"2026-06-06 13:13\"}"
            )

            payload = {
                "model": "claude-sonnet-4-20250514",
                "max_tokens": 500,
                "messages": [{
                    "role": "user",
                    "content": [
                        {
                            "type": "image",
                            "source": {
                                "type": "base64",
                                "media_type": "image/jpeg",
                                "data": img_b64,
                            },
                        },
                        {"type": "text", "text": prompt},
                    ],
                }],
            }

            headers = {
                "x-api-key": api_key,
                "anthropic-version": "2023-06-01",
                "Content-Type": "application/json",
            }

            resp = requests.post(
                "https://api.anthropic.com/v1/messages",
                headers=headers,
                data=json.dumps(payload),
                timeout=40,
            )
            if resp.status_code != 200:
                return None, f"Erro HTTP {resp.status_code} ao chamar Claude."

            data = resp.json()
            content = ""
            for block in data.get("content", []):
                if block.get("type") == "text":
                    content = block.get("text", "").strip()
                    break

            if not content:
                return None, "Resposta vazia do Claude."

            # Remove possíveis marcas de bloco de código
            if content.startswith("```"):
                content = content.strip("`\n ")
                if content.lower().startswith("json"):
                    content = content[4:].lstrip()

            try:
                parsed = json.loads(content)
            except Exception as e:
                return None, f"Não consegui interpretar o JSON do Claude: {e}"

        except Exception as e:
            return None, f"Erro ao chamar Claude: {e}"

    # --- Normaliza resultado (comum OpenAI e Claude) ---
    map_name = (parsed.get("map_name") or "").strip() or None
    device_name = (parsed.get("device_name") or "").strip() or None

    splices_val = parsed.get("splices")
    try:
        splices_val = int(splices_val) if splices_val is not None else None
    except Exception:
        splices_val = None

    role_raw = (parsed.get("map_role") or "").strip().upper()
    if role_raw in ("PONTA", "P"):
        map_role = "PONTA"
    elif role_raw in ("MEIO", "M"):
        map_role = "MEIO"
    else:
        # Auto-detecta: se ft_out existe → MEIO, senão → PONTA
        ft_out_raw = (str(parsed.get("ft_out") or "")).strip()
        map_role = "MEIO" if ft_out_raw and ft_out_raw.lower() != "null" else "PONTA"

    ft_in = (str(parsed.get("ft_in") or "")).strip() or None
    if ft_in and ft_in.lower() == "null":
        ft_in = None

    ft_out = (str(parsed.get("ft_out") or "")).strip() or None
    if ft_out and ft_out.lower() == "null":
        ft_out = None
    if map_role == "PONTA":
        ft_out = None

    gps = (str(parsed.get("gps") or "")).strip() or None
    if gps and gps.lower() == "null":
        gps = None

    photo_datetime = (str(parsed.get("photo_datetime") or "")).strip() or None
    if photo_datetime and photo_datetime.lower() == "null":
        photo_datetime = None

    result = {
        "map_name": map_name,
        "device_name": device_name,
        "splices": splices_val,
        "map_role": map_role,
        "ft_in": ft_in,
        "ft_out": ft_out,
        "gps": gps,
        "photo_datetime": photo_datetime,
    }

    # map_name não é obrigatório — o GPS localiza o mapa automaticamente
    if not (result["device_name"] and result["splices"] is not None and result["map_role"]):
        return None, "Claude não retornou os campos obrigatórios (device_name, splices, map_role)."

    return result, None


def _best_tier_for(count: int, company: str | None, project_id: int | None = None):
    """Retorna o objeto SpliceTier que corresponde ao count, com prioridade projeto > empresa > global."""
    def _best(q):
        tiers = q.all()
        for t in sorted(tiers, key=lambda x: x.min_splices or 0, reverse=True):
            if (t.min_splices or 0) <= count:
                return t
        return None
    if project_id:
        t = _best(SpliceTier.query.filter(SpliceTier.project_id == project_id))
        if t:
            return t
    if company:
        t = _best(SpliceTier.query.filter(SpliceTier.company == company, SpliceTier.project_id.is_(None)))
        if t:
            return t
    return _best(SpliceTier.query.filter(SpliceTier.company.is_(None), SpliceTier.project_id.is_(None)))


def compute_billing_codes(
    splices: int,
    device_name: str,
    company: str | None,
    project_id: int | None = None,
    map_role: str | None = None,
    ribbon_count: int | None = None,
) -> list:
    """Calcula os codigos de cobranca para um lancamento.

    Retorna lista de codigos unicos (ex: ["FS01", "FS15"]).
    Lista vazia se o projeto nao usa codigos.
    """
    codes = []

    # Codigo do dispositivo (enclosure/caixa) — varia por MEIO/PONTA
    dt = _get_device_type(device_name or "", company, project_id)
    if dt:
        role = (map_role or "").strip().upper()
        dev_code = None
        if role == "MEIO" and getattr(dt, "billing_code_meio", None):
            dev_code = dt.billing_code_meio.strip()
        elif role == "PONTA" and getattr(dt, "billing_code_ponta", None):
            dev_code = dt.billing_code_ponta.strip()
        elif getattr(dt, "billing_code", None):
            dev_code = dt.billing_code.strip()
        if dev_code:
            codes.append(dev_code)

    # Codigo de fusoes (por faixa — unico, sem distincao MEIO/PONTA)
    is_rib = dt and bool(getattr(dt, "is_ribbon", False))
    total = int(ribbon_count or 0) if is_rib else int(splices or 0)
    if total > 0:
        tier = _best_tier_for(total, company, project_id)
        if tier and getattr(tier, "code_splice", None):
            code = tier.code_splice.strip()
            if code and code not in codes:
                codes.append(code)

    return codes


def compute_prices(
    splices: int,
    device_name: str,
    company: str | None,
    project_id: int | None = None,
    included_override: int | None = None,
    map_role: str | None = None,
    ribbon_count: int | None = None,
):
    """Calcula preco de fusoes e dispositivo para um lancamento manual.

    - Ribbon: se o dispositivo for ribbon, calcula ribbon_count x ribbon_price_usd.
      Fusoes e valor de dispositivo sao ignorados nesse caso.
    - MEIO/PONTA: included_override tem prioridade sobre a configuracao padrao.
    """
    # --- Ribbon ---
    # price_splices = fitas x valor_por_fita
    # price_device  = valor fixo do enclosure/caixa (value_usd do DeviceType)
    is_rib, ribbon_price = device_is_ribbon(device_name or "", company, project_id)
    if is_rib:
        count = int(ribbon_count or 0)
        price_ribbon = count * ribbon_price
        price_enclosure = device_value_for(device_name or "", company, project_id, map_role=map_role)
        return price_ribbon, price_enclosure, price_ribbon + price_enclosure

    total_splices = int(splices or 0)

    # Fusões inclusas: override do mapa (MEIO/PONTA) > projeto > empresa > padrão
    if included_override is not None:
        included = int(included_override or 0)
    else:
        included = int(included_splices_for(company, project_id) or 0)

    included = max(included, 0)

    # Só cobra o que ultrapassar as inclusas
    charge = max(total_splices - included, 0)

    # A faixa de preço é escolhida pelo TOTAL de fusões, não apenas as cobradas
    price_per_splice = tier_price_for(total_splices, company, project_id) if charge > 0 else 0.0
    price_splices = charge * price_per_splice

    # Valor do dispositivo
    price_device = device_value_for(device_name or "", company, project_id, map_role=map_role)

    return price_splices, price_device, price_splices + price_device





# --------- Decorators ---------

def admin_required(f):
    @wraps(f)
    @login_required
    def wrapper(*args, **kwargs):
        if not getattr(current_user, "is_admin", False):
            flash("Apenas o administrador pode acessar essa área.", "danger")
            return redirect(url_for("index"))
        return f(*args, **kwargs)
    return wrapper

# --------- Rotas ---------
@app.route("/", methods=["GET", "POST"])
@login_required
def index():
    # Importar planilha foi removido do sistema; qualquer POST apenas mostra aviso.
    if request.method == "POST":
        flash("A importação de planilha foi desativada neste sistema.", "warning")
        return redirect(url_for("index"))

    # filtros vindos da URL
    company_filter = request.args.get("company") or None
    splicer_filter = request.args.get("splicer") or None
    map_filter = request.args.get("map") or None
    device_filter = request.args.get("device") or None
    start_raw = request.args.get("start") or None
    end_raw = request.args.get("end") or None

    # base da consulta
    query = Record.query

    # regras de visibilidade por tipo de usuário
    is_admin = getattr(current_user, "is_admin", False)
    is_owner = getattr(current_user, "is_company_owner", False)
    enforced_splicer = None

    if is_admin:
        # Admin enxerga todos os registros, sem restrição adicional.
        pass
    elif is_owner:
        # Dono de empresa enxerga todos os registros da PRÓPRIA empresa.
        owner_company = getattr(current_user, "company_name", None)
        if owner_company:
            company_filter = owner_company
            # ignoramos qualquer empresa passada manualmente na URL
    else:
        # Splicer normal vê apenas os lançamentos feitos por ele mesmo.
        enforced_splicer = getattr(current_user, "splicer_name", None) or current_user.username
        query = query.filter(Record.splicer == enforced_splicer)
        # evita que tente filtrar manualmente outro splicer pela URL
        splicer_filter = None

    # filtros principais
    if company_filter:
        query = query.filter(Record.company == company_filter)
    if splicer_filter and (is_admin or is_owner):
        # apenas admin / dono de empresa podem filtrar por outro splicer
        query = query.filter(Record.splicer == splicer_filter)
    if map_filter:
        query = query.filter(Record.map.ilike(f"%{map_filter}%"))
    if device_filter:
        query = query.filter(Record.device.ilike(f"%{device_filter}%"))

    # filtros por data
    if start_raw:
        try:
            start_dt = datetime.fromisoformat(start_raw)
            query = query.filter(Record.created_date >= start_dt)
        except ValueError:
            pass
    if end_raw:
        try:
            end_dt = datetime.fromisoformat(end_raw)
            query = query.filter(Record.created_date <= end_dt)
        except ValueError:
            pass

    # Oculta dispositivos importados do KMZ que ainda não tiveram lançamento (splicer/splices)
    query = query.filter((Record.splices > 0) | (Record.total_usd > 0) | (Record.price_device_usd > 0) | (Record.price_splices_usd > 0) | (Record.ribbon_count.isnot(None)))

    records = query.order_by(Record.created_date.desc().nullslast(), Record.id.desc()).all()

    # Busca lançamentos de horas com os mesmos filtros e adiciona à lista
    hq = HourRecord.query
    if is_admin:
        pass
    elif is_owner:
        owner_company = getattr(current_user, "company_name", None)
        if owner_company:
            hq = hq.filter(HourRecord.company == owner_company)
    else:
        hq = hq.filter(HourRecord.splicer == enforced_splicer)
    if company_filter:
        hq = hq.filter(HourRecord.company == company_filter)
    if splicer_filter and (is_admin or is_owner):
        hq = hq.filter(HourRecord.splicer == splicer_filter)
    if map_filter:
        hq = hq.filter(HourRecord.map_name.ilike(f"%{map_filter}%"))
    if start_raw:
        try:
            hq = hq.filter(HourRecord.created_date >= datetime.fromisoformat(start_raw))
        except ValueError:
            pass
    if end_raw:
        try:
            hq = hq.filter(HourRecord.created_date <= datetime.fromisoformat(end_raw))
        except ValueError:
            pass
    hour_records = hq.order_by(HourRecord.created_date.desc().nullslast()).all()

    # Cria objetos compatíveis com o template para os lançamentos de horas
    class HourRow:
        """Wrapper para HourRecord compatível com o template de Record."""
        is_hour_record = True
        def __init__(self, hr):
            self.id = hr.id
            self.map = hr.map_name or "-"
            self.type = "HORAS"
            self.device = f"{hr.hours}h — {hr.description or ''}"
            self.splicer = hr.splicer
            self.company = hr.company
            self.project_id = hr.project_id
            self.created_date = hr.created_date
            self.splices = None
            self.ribbon_count = None
            self.map_role = None
            self.billing_codes_json = f'["{hr.billing_code}"]' if hr.billing_code else None
            self.price_splices_usd = hr.total_usd  # exibe no campo fusões
            self.price_device_usd = 0.0
            self.total_usd = hr.total_usd
            self.photos = []
            self.can_view_values = True
            self.is_placed = False
            self.test_done = False
            self._hr = hr  # referência original

    hour_rows = [HourRow(hr) for hr in hour_records]

    # Mescla e ordena por data
    all_rows = records + hour_rows
    all_rows.sort(key=lambda r: (r.created_date or datetime.min), reverse=True)

    total_rows = len(all_rows)
    total_amount = sum((r.total_usd or 0) for r in all_rows)

    # Substitui records pela lista mesclada para o template
    records = all_rows

    companies = [c.name for c in CompanyConfig.query.order_by(CompanyConfig.name).all()]
    companies_from_records = {
        c for (c,) in db.session.query(Record.company).distinct().all() if c
    }
    all_companies = sorted(set(companies) | companies_from_records)

    # lista de splicers / usuários cadastrados (para o filtro)
    splicers_from_records = {
        s for (s,) in db.session.query(Record.splicer).distinct().all() if s
    }
    splicers_from_users = {
        (u.splicer_name or u.username)
        for u in User.query.all()
        if (u.splicer_name or u.username)
    }
    all_splicers = sorted(splicers_from_records | splicers_from_users)

    # Para usuários normais, o dropdown DEVE mostrar só ele mesmo.
    # Para dono de empresa, mantemos a lista completa (pode filtrar por qualquer splicer da empresa).
    if not is_admin:
        if enforced_splicer:
            all_splicers = [enforced_splicer]
            splicer_filter = enforced_splicer


    # Mapas que o usuário tem permissão para EDITAR (apenas para splicer comum).
    editable_maps = []
    if is_admin:
        # admin edita tudo; o template ignora esta lista nesse caso
        editable_maps = []
    elif not is_owner:
        try:
            editable_maps = [m.name for m in getattr(current_user, "maps_with_access", [])]
        except Exception:
            editable_maps = []
    return render_template(
        "index.html",
        records=records,
        total_rows=total_rows,
        total_amount=total_amount,
        companies=all_companies,
        splicers=all_splicers,
        company_filter=company_filter or "",
        splicer_filter=splicer_filter or "",
        map_filter=map_filter or "",
        device_filter=device_filter or "",
        start=start_raw or "",
        end=end_raw or "",
        editable_maps=editable_maps,
        can_view_values=is_admin or getattr(current_user, 'can_view_values', True),
    )

def build_filtered_record_query_from_request():
    """Reaproveita os filtros da tela principal para buscar os registros."""
    company_filter = request.args.get("company") or None
    map_filter = request.args.get("map") or None
    device_filter = request.args.get("device") or None
    splicer_filter = request.args.get("splicer") or None
    start_raw = request.args.get("start") or None
    end_raw = request.args.get("end") or None

    query = Record.query

    # Restrições de visibilidade por tipo de usuário
    is_admin = getattr(current_user, "is_admin", False)
    is_owner = getattr(current_user, "is_company_owner", False)

    if is_admin:
        # Admin enxerga todos os registros; os filtros são aplicados mais abaixo.
        pass
    elif is_owner:
        # Dono de empresa enxerga todos os registros da própria empresa.
        # Ignoramos o filtro de empresa vindo da URL, se houver.
        owner_company = _current_user_company_name()
        if owner_company:
            company_filter = owner_company
    else:
        # Splicer comum: sempre vê apenas os próprios registros.
        enforced_splicer = getattr(current_user, "splicer_name", None) or current_user.username
        query = query.filter(Record.splicer == enforced_splicer)
        # Evita que o splicer tente ver registros de outra pessoa via URL.
        splicer_filter = None

    # Filtros por campos principais
    if company_filter:
        query = query.filter(Record.company == company_filter)
    if map_filter:
        query = query.filter(Record.map == map_filter)
    if device_filter:
        query = query.filter(Record.device == device_filter)

    # Filtro de splicer (somente admin ou dono de empresa)
    if splicer_filter and (is_admin or is_owner):
        query = query.filter(Record.splicer == splicer_filter)

    # Filtro por datas
    # OBS: o modelo Record não possui coluna "date". A data usada no sistema é "created_date" (ou "created_at" como fallback).
    record_day = func.date(func.coalesce(Record.created_date, Record.created_at))

    if start_raw:
        try:
            start_date = datetime.strptime(start_raw, "%Y-%m-%d").date()
            query = query.filter(record_day >= start_date)
        except ValueError:
            pass

    if end_raw:
        try:
            end_date = datetime.strptime(end_raw, "%Y-%m-%d").date()
            query = query.filter(record_day <= end_date)
        except ValueError:
            pass

    # Oculta dispositivos importados do KMZ que ainda não viraram produção
    query = query.filter((Record.splices > 0) | (Record.total_usd > 0) | (Record.price_device_usd > 0) | (Record.price_splices_usd > 0) | (Record.ribbon_count.isnot(None)))

    return query


@app.route("/photo/<int:photo_id>")
@login_required
def photo_file(photo_id: int):
    """Retorna o binário de uma foto salva em RecordPhoto.

    Performance:
    - suporte a thumbnail via ?size=thumb
    - cache agressivo (as fotos são imutáveis por ID)
    - suporta fotos antigas no banco (BLOB) e novas no R2 (r2_key)
    """
    photo = RecordPhoto.query.get_or_404(photo_id)

    want_thumb = (request.args.get("size") == "thumb")

    # ETag leve (não faz md5 do arquivo inteiro)
    etag = f'W/"{photo.id}-{"t" if want_thumb else "o"}-{photo.r2_thumb_key or ""}-{photo.r2_key or ""}-{photo.size_bytes or 0}"'
    if request.headers.get("If-None-Match") == etag:
        resp = make_response("", 304)
        resp.headers["ETag"] = etag
        resp.headers["Cache-Control"] = "public, max-age=31536000, immutable"
        return resp

    def get_original_bytes() -> tuple[bytes, str]:
        if getattr(photo, "data", None) and len(photo.data) > 0:
            return photo.data, (photo.content_type or "image/jpeg")
        if getattr(photo, "r2_key", None) and r2_enabled():
            try:
                b = r2_get_bytes(photo.r2_key)
                return b, (photo.content_type or "application/octet-stream")
            except Exception:
                pass
        return b"", (photo.content_type or "application/octet-stream")

    blob = b""
    mimetype = "application/octet-stream"

    if want_thumb:
        # 1) thumb no banco
        if getattr(photo, "thumb_data", None):
            blob = photo.thumb_data
            mimetype = photo.thumb_content_type or "image/jpeg"
        # 2) thumb no R2
        elif getattr(photo, "r2_thumb_key", None) and r2_enabled():
            try:
                blob = r2_get_bytes(photo.r2_thumb_key)
                mimetype = "image/jpeg"
            except Exception:
                blob = b""
        # 3) gerar thumb sob demanda (uma vez) e salvar
        if not blob:
            orig, _ct = get_original_bytes()
            if orig:
                try:
                    # gera thumb pequeno (rápido)
                    img = Image.open(BytesIO(orig))
                    img = ImageOps.exif_transpose(img)
                    img.thumbnail((420, 420))
                    out = BytesIO()
                    img.save(out, format="JPEG", quality=70, optimize=True, progressive=True)
                    thumb_bytes = out.getvalue()

                    if getattr(photo, "r2_key", None) and r2_enabled():
                        tkey = photo.r2_thumb_key or (photo.r2_key + ".thumb.jpg")
                        r2_put_bytes(tkey, thumb_bytes, content_type="image/jpeg")
                        photo.r2_thumb_key = tkey
                    else:
                        photo.thumb_data = thumb_bytes
                        photo.thumb_content_type = "image/jpeg"

                    db.session.commit()

                    blob = thumb_bytes
                    mimetype = "image/jpeg"
                except Exception:
                    db.session.rollback()

        if not blob:
            # fallback: retorna original
            blob, mimetype = get_original_bytes()
    else:
        blob, mimetype = get_original_bytes()

    resp = send_file(
        BytesIO(blob),
        mimetype=mimetype,
        as_attachment=False,
        download_name=photo.filename,
        conditional=True,
    )
    resp.headers["ETag"] = etag
    resp.headers["Cache-Control"] = "public, max-age=31536000, immutable"
    return resp


    resp = send_file(
        BytesIO(blob),
        mimetype=mimetype,
        as_attachment=False,
        download_name=photo.filename,
        conditional=True,
    )
    if etag:
        resp.headers["ETag"] = etag
    resp.headers["Cache-Control"] = "public, max-age=31536000, immutable"
    return resp



@app.route("/photo_thumb/<int:pid>")
@login_required
def photo_thumb(pid: int):
    """
    Rota legada para thumbnails de foto.

    Alguns templates antigos ainda chamam url_for('photo_thumb', pid=...).
    Para manter compatibilidade e evitar BuildError, essa rota apenas
    reutiliza a lógica de `photo_file` com size=thumb.
    """
    # Repassa para a rota principal de foto, forçando ?size=thumb
    # (não faz redirect para simplificar)
    with current_app.test_request_context(query_string={"size": "thumb"}):
        return photo_file(pid)

@app.route("/record/<int:rid>/photos", methods=["POST"])
@login_required
def record_add_photos(rid: int):
    """Anexa fotos a um lançamento já criado.
    Usado pelo modo rápido: salva o lançamento primeiro e sobe as fotos depois.
    """
    rec = Record.query.get_or_404(rid)

    # Permissões básicas: splicer só mexe nos próprios registros (admin pode tudo)
    is_admin = bool(getattr(current_user, "is_admin", False))
    is_owner = bool(getattr(current_user, "is_owner", False))
    if not is_admin:
        # se não for admin/dono, deve ser o mesmo splicer
        current_splicer = (getattr(current_user, "splicer_name", None) or current_user.username)
        if (rec.splicer or "") != current_splicer:
            abort(403)

    files = request.files.getlist("photos")
    created = []
    if files:
        for f in files[:5]:

            if not f or not f.filename:
                continue

            # Performance + NOLOSS:
            # 1) Always store *optimized* bytes in Postgres (fast + guarantees we never lose the photo).
            # 2) If R2 is configured, upload to R2 in a background thread and then set r2_key.
            filename, content_type, data, thumb_data, thumb_ct = process_uploaded_photo(f)
            if not data:
                continue

            photo = RecordPhoto(
                record_id=rec.id,
                filename=filename,
                content_type=content_type,
                data=data if data else b"",  # NOT NULL constraint
                thumb_data=thumb_data,
                thumb_content_type=thumb_ct,
                r2_key=None,
                r2_thumb_key=None,
                size_bytes=int(len(data) if data else 0),
            )

            db.session.add(photo)
            db.session.flush()

            if r2_enabled() and data:
                try:
                    key = r2_key_for_record_photo(rec.id, filename)
                    thumb_key = key + ".thumb.jpg" if thumb_data else None
                    enqueue_r2_upload(int(photo.id), key, data, content_type, thumb_key=thumb_key, thumb_bytes=thumb_data, thumb_content_type=thumb_ct)
                except Exception as e:
                    print(f"[R2] Failed to enqueue upload for photo id={getattr(photo,'id',None)}: {e}")
                    traceback.print_exc()

            created.append(int(photo.id))
        db.session.commit()

    return jsonify({"ok": True, "photo_ids": created, "record_id": int(rec.id)})

@app.route("/photos_filtered_zip")
@login_required
def photos_filtered_zip():
    """Gera um .zip com TODAS as fotos dos registros filtrados na tela principal."""

    query = build_filtered_record_query_from_request()
    records = query.all()

    if not records:
        flash("Nenhum registro encontrado para gerar o .zip de fotos.", "warning")
        return redirect(request.referrer or url_for("index"))

    record_ids = [r.id for r in records]
    photos = RecordPhoto.query.filter(RecordPhoto.record_id.in_(record_ids)).all()

    if not photos:
        flash("Nenhuma foto encontrada para os filtros selecionados.", "warning")
        return redirect(request.referrer or url_for("index"))

    mem = BytesIO()
    with zipfile.ZipFile(mem, "w", zipfile.ZIP_DEFLATED) as zf:
        rec_by_id = {r.id: r for r in records}
        for photo in photos:
            rec = rec_by_id.get(photo.record_id)
            company_part = (rec.company or "SEM_EMPRESA").replace("/", "-") if rec else "SEM_EMPRESA"
            map_part = (rec.map or "SEM_MAP").replace("/", "-") if rec else "SEM_MAP"
            device_part = (rec.device or f"ID-{photo.record_id}").replace("/", "-") if rec else f"ID-{photo.record_id}"
            safe_filename = photo.filename or f"foto_{photo.id}.jpg"
            zip_path = f"{company_part}/{map_part}/{device_part}/ID-{photo.record_id}_PH-{photo.id}_{safe_filename}"
            try:
                payload = None
                # Prefer DB data when present; if cleared after R2 upload, fetch from R2.
                if photo.data:
                    payload = bytes(photo.data)
                elif getattr(photo, "r2_key", None):
                    payload = r2_get_bytes(photo.r2_key)

                # Guarantee bytes
                if hasattr(payload, "read"):
                    payload = payload.read()
                if isinstance(payload, str):
                    payload = payload.encode("utf-8", errors="ignore")
                if payload is None:
                    current_app.logger.warning(
                        f"[ZIP] Missing payload for photo id={photo.id} record_id={photo.record_id} r2_key={getattr(photo,'r2_key',None)}"
                    )
                    continue
                if not isinstance(payload, (bytes, bytearray)):
                    try:
                        payload = bytes(payload)
                    except Exception:
                        payload = str(payload).encode("utf-8", errors="ignore")

                if len(payload) == 0:
                    current_app.logger.warning(
                        f"[ZIP] Empty payload for photo id={photo.id} record_id={photo.record_id} r2_key={getattr(photo,'r2_key',None)}"
                    )
                    continue

                zf.writestr(zip_path, payload)
            except Exception as e:
                current_app.logger.exception(f"[ZIP] Failed to add photo id={photo.id} to zip: {e}")
                continue
            except Exception as e:
                current_app.logger.exception(f"[ZIP] Failed to add photo id={photo.id} to zip: {e}")
                continue

    mem.seek(0)
    return send_file(
        mem,
        mimetype="application/zip",
        as_attachment=True,
        download_name="fotos_filtradas.zip",
    )




@app.route("/record/<int:rid>/photos_zip")
@login_required
def record_photos_zip(rid):
    """Gera um .zip com TODAS as fotos de um único lançamento (record).

    IMPORTANT: when CLEAR_DB_AFTER_R2=1 the binary data may have been cleared from Postgres.
    In that case, we fetch the original bytes from Cloudflare R2 using photo.r2_key.
    """
    record = Record.query.get_or_404(rid)

    photos = (
        db.session.query(RecordPhoto)
        .filter(RecordPhoto.record_id == rid)
        .order_by(RecordPhoto.created_at.asc())
        .all()
    )

    if not photos:
        flash("Este dispositivo não possui fotos para download.", "warning")
        return redirect(request.referrer or url_for("index"))

    mem = BytesIO()
    skipped = 0
    with zipfile.ZipFile(mem, "w", zipfile.ZIP_DEFLATED) as zf:
        for photo in photos:
            safe_filename = photo.filename or f"foto_{photo.id}.jpg"
            device_part = (record.device or f"ID-{record.id}").replace("/", "-")
            zip_path = f"{device_part}/ID-{record.id}_PH-{photo.id}_{safe_filename}"

            payload = b""
            try:
                if photo.data and isinstance(photo.data, (bytes, bytearray)) and len(photo.data) > 10:
                    payload = bytes(photo.data)
                elif getattr(photo, "r2_key", None):
                    payload = r2_get_bytes(photo.r2_key)
            except Exception as e:
                print(f"[ZIP] Failed to fetch photo id={photo.id} r2_key={getattr(photo,'r2_key',None)}: {e}")
                traceback.print_exc()
                payload = b""

            # Avoid writing empty/invalid files into the zip
            if not payload or len(payload) < 10:
                skipped += 1
                continue

            zf.writestr(zip_path, payload)

    mem.seek(0)
    ts = datetime.utcnow().strftime("%Y%m%d-%H%M%S")
    zip_name = f"fotos_dispositivo_{(record.device or record.id)}_{ts}.zip"

    if skipped and len(photos) == skipped:
        flash("Não foi possível montar o ZIP (fotos vazias). Verifique R2 e r2_key.", "danger")

    return send_file(
        mem,
        mimetype="application/zip",
        as_attachment=True,
        download_name=zip_name,
    )


def _parse_can_cables_from_form(form):
    raw_count = (form.get("can_cable_count") or "").strip()
    try:
        count = max(0, int(raw_count or 0))
    except Exception:
        count = 0

    cables = []
    for i in range(1, count + 1):
        name = (form.get(f"can_cable_{i}_name") or f"Cabo {i}").strip() or f"Cabo {i}"
        ft_in = (form.get(f"can_cable_{i}_ft_in") or "").strip()
        ft_out = (form.get(f"can_cable_{i}_ft_out") or "").strip()
        cables.append({"name": name, "ft_in": ft_in, "ft_out": ft_out})
    return count, cables


def _record_ft_lines(rec):
    lines = []
    if (rec.type or "").strip().upper().startswith("CAN"):
        try:
            cables = json.loads(rec.can_cables_json or "[]") if rec.can_cables_json else []
        except Exception:
            cables = []
        for idx, c in enumerate(cables, start=1):
            cname = (c.get("name") or f"Cabo {idx}").strip() or f"Cabo {idx}"
            cin = (c.get("ft_in") or "").strip()
            cout = (c.get("ft_out") or "").strip()
            part = f"{cname}: IN {cin or '-'} | OUT {cout or '-'}"
            lines.append(part)
    else:
        if (rec.ft_in or "").strip() or (rec.ft_out or "").strip():
            lines.append(f"FT IN: {(rec.ft_in or '-').strip() or '-'}")
            lines.append(f"FT OUT: {(rec.ft_out or '-').strip() or '-'}")
    return lines


@app.route("/entry", methods=["GET", "POST"])
@login_required
def entry():
    """Lançamento manual de produção (uma linha por vez).

    Para splicer: modo rápido (não escolhe projeto). O sistema deriva o projeto pelo mapa.
    """
    # Dono de empresa (company_owner) tem acesso apenas para visualização.
    # Ele não pode lançar produção manualmente. Se tentar acessar /entry,
    # redirecionamos para a tela principal.
    is_owner = bool(getattr(current_user, "is_company_owner", False))
    is_admin = bool(getattr(current_user, "is_admin", False))
    if is_owner and not is_admin:
        flash("Dono de empresa só pode visualizar. Utilize um usuário splicer para lançar produção.", "danger")
        return redirect(url_for("index"))

    companies = [c.name for c in CompanyConfig.query.order_by(CompanyConfig.name).all()]

    # Modo rápido (padrão): não escolhe projeto. O sistema deriva o projeto pelo mapa.
    # Para ADMIN, se quiser o modo completo (com seleção de projeto), use /entry?full=1
    full_mode = bool(getattr(current_user, "is_admin", False)) and (request.args.get("full") == "1")
    is_splicer = not full_mode
    entry_url = url_for('entry', full='1') if full_mode else url_for('entry')

    # Permite reset do "modo rápido" (opcional)
    if request.method == "GET" and request.args.get("reset") == "1":
        for k in ("entry_company", "entry_map_id", "entry_map_role", "entry_type"):
            session.pop(k, None)

    projects_by_company = {}
    for pr in Project.query.order_by(Project.company, Project.name).all():
        projects_by_company.setdefault(pr.company, []).append({"id": pr.id, "name": pr.name})

    maps_by_project = {}
    maps_by_company = {}          # mapas SEM projeto (legado) por empresa
    maps_by_company_all = {}      # todos mapas por empresa (inclui mapas vinculados a projeto) – usado no splicer

    for m in CompanyMap.query.order_by(CompanyMap.company, CompanyMap.name).all():
        mobj = {
            "id": int(m.id),
            "name": m.name,
            "project_id": int(m.project_id) if m.project_id else None,
            "mid_end_enabled": bool(getattr(m, "mid_end_enabled", False)),
            "included_splices_meio": int(getattr(m, "included_splices_meio", 0) or 0),
            "included_splices_ponta": int(getattr(m, "included_splices_ponta", 0) or 0),
        }
        maps_by_company_all.setdefault(m.company, []).append(mobj)

        if m.project_id:
            maps_by_project.setdefault(str(m.project_id), []).append(mobj)
        else:
            maps_by_company.setdefault(m.company, []).append(mobj)

    devices_by_project = {}
    devices_by_company = {}
    for dt in DeviceType.query.order_by(DeviceType.company, DeviceType.name).all():
        dobj = {"name": dt.name, "is_ribbon": bool(getattr(dt, "is_ribbon", False))}
        if dt.project_id:
            devices_by_project.setdefault(str(dt.project_id), []).append(dobj)
        else:
            key = dt.company or "__global__"
            devices_by_company.setdefault(key, []).append(dobj)

    splicer_options = []
    if getattr(current_user, "is_admin", False):
        splicer_options = [
            (u.splicer_name or u.username)
            for u in User.query.order_by(User.username).all()
            if (u.splicer_name or u.username)
        ]

    default_splicer = getattr(current_user, "splicer_name", None) or current_user.username

    # Prefill "modo rápido" (splicer) – mantém enquanto o usuário ficar nessa tela
    pre_company = session.get("entry_company") if is_splicer else None
    pre_map_id = session.get("entry_map_id") if is_splicer else None
    pre_map_role = session.get("entry_map_role") if is_splicer else None
    pre_type = session.get("entry_type") if is_splicer else None

    if request.method == "POST":
        company = (request.form.get("company") or "").strip() or None

        # Para splicer, projeto vem do mapa; para admin pode vir do form.
        project_id_raw = (request.form.get("project_id") or "").strip()
        project_id = int(project_id_raw) if project_id_raw.isdigit() else None

        map_id_raw = (request.form.get("map_id") or "").strip()
        map_val = (request.form.get("map") or "").strip()

        map_obj = None
        if map_id_raw.isdigit():
            map_obj = CompanyMap.query.get(int(map_id_raw))
        else:
            # compatibilidade: se ainda vier pelo nome
            if company and map_val:
                q = CompanyMap.query.filter_by(company=company, name=map_val)
                map_obj = q.first()

        # Se for splicer, sempre deriva project_id pelo mapa escolhido
        if is_splicer and map_obj:
            project_id = int(map_obj.project_id) if map_obj.project_id else None
            map_val = map_obj.name

        # Garantia extra: se ainda não houver project_id mas o mapa pertence a um projeto,
        # usamos o projeto do mapa (também para admin em modo completo).
        if map_obj is not None and map_obj.project_id is not None and project_id is None:
            project_id = int(map_obj.project_id)

        type_val = (request.form.get("type") or "").strip()
        device_name = (request.form.get("device_name") or "").strip()

        device_for_price = type_val or device_name

        splices_raw = request.form.get("splices") or "0"
        ribbon_count_raw = (request.form.get("ribbon_count") or "").strip()
        created_raw = request.form.get("created") or ""
        splicer = (request.form.get("splicer") or "").strip() or default_splicer
        confirm_duplicate = (request.form.get("confirm_duplicate") == "yes")
        ft_in = (request.form.get("ft_in") or "").strip()
        ft_out = (request.form.get("ft_out") or "").strip()
        can_cable_count, can_cables = _parse_can_cables_from_form(request.form)

        existing = None
        if map_val and device_name:
            dup_query = Record.query.filter(Record.map == map_val, Record.device == device_name)
            if company:
                dup_query = dup_query.filter(Record.company == company)
            if project_id is not None:
                dup_query = dup_query.filter(Record.project_id == project_id)
            existing = dup_query.order_by(Record.created_date.desc().nullslast(), Record.id.desc()).first()

        is_admin = bool(getattr(current_user, "is_admin", False))
        if existing:
            # Regra de negócio: um dispositivo só pode ser lançado uma vez neste mapa/projeto.
            # Para lançar novamente é preciso o admin decidir; usuários comuns são bloqueados.
            if not is_admin:
                flash(
                    "Este dispositivo já foi lançado neste mapa e não pode ser lançado novamente. "
                    "Se houver erro, peça para o admin excluir ou ajustar o lançamento anterior.",
                    "danger",
                )
                return redirect(entry_url)

            # Admin: ainda permite duplicar, mas exige confirmação explícita.
            if not confirm_duplicate:
                flash(
                    "Este dispositivo já foi lançado neste mapa. Data: "
                    + (existing.created_date.date().isoformat() if existing.created_date else "-")
                    + f", Splicer: {existing.splicer}. Se desejar lançar novamente, confirme o lançamento.",
                    "warning",
                )
                return render_template(
                    "entry.html",
                    splicer_options=splicer_options,
                    companies=companies,
                    projects_by_company=projects_by_company,
                    maps_by_project=maps_by_project,
                    maps_by_company=maps_by_company,
                    maps_by_company_all=maps_by_company_all,
                    devices_by_project=devices_by_project,
                    devices_by_company=devices_by_company,
                    default_splicer=default_splicer,
                    today=date.today().isoformat(),
                    duplicate_record=existing,
                    form_company=company,
                    form_project_id=project_id_raw,
                    form_map=map_val,
                    form_type=type_val,
                    form_device_name=device_name,
                    form_splices=splices_raw,
                    form_created=created_raw or date.today().isoformat(),
                    form_ft_in=ft_in,
                    form_ft_out=ft_out,
                    form_can_cable_count=str(can_cable_count or ""),
                    form_can_cables_json=json.dumps(can_cables, ensure_ascii=False),
                    confirm_duplicate=True,
                    is_splicer=is_splicer,
                )
        try:
            splices = int(splices_raw or 0)
        except ValueError:
            splices = 0

        try:
            ribbon_count = int(ribbon_count_raw) if ribbon_count_raw else None
        except ValueError:
            ribbon_count = None

        if created_raw:
            try:
                created_date = datetime.strptime(created_raw, "%Y-%m-%d")
            except ValueError:
                created_date = datetime.utcnow()
        else:
            today = date.today()
            created_date = datetime(today.year, today.month, today.day)

        # Regra opcional por MAPA: se o mapa estiver com MEIO/PONTA habilitado,
        # o usuário precisa escolher e o sistema aplica as fusões inclusas do mapa.
        map_role = (request.form.get("map_role") or "").strip().upper() or None

        # Regras de MEIO/PONTA por mapa:
        # 1) descobre o CompanyMap correto
        map_cfg = map_obj
        if not map_cfg and map_val and company:
            q = CompanyMap.query.filter_by(company=company, name=map_val)
            if project_id is not None:
                q = q.filter_by(project_id=project_id)
            map_cfg = q.first()
            if not map_cfg:
                map_cfg = CompanyMap.query.filter_by(company=company, name=map_val, project_id=None).first()

        # Se o mapa estiver configurado para MEIO/PONTA, o campo passa a ser obrigatório
        if map_cfg and bool(getattr(map_cfg, "mid_end_enabled", False)) and map_role not in ("MEIO", "PONTA"):
            flash("Este mapa exige selecionar MEIO ou PONTA.", "danger")
            return redirect(entry_url)

        included_override, included_applied, map_cfg = resolve_included_override(
            company,
            project_id,
            map_cfg,
            map_val,
            map_role,
        )

        price_splices, price_device, total = compute_prices(
            splices,
            device_for_price,
            company,
            project_id,
            included_override=included_override,
            map_role=map_role,
            ribbon_count=ribbon_count,
        )

        is_rib, _ = device_is_ribbon(device_for_price or "", company, project_id)
        _bcodes = compute_billing_codes(splices, device_for_price or "", company, project_id, map_role=map_role, ribbon_count=ribbon_count)
        rec = Record(
            map=map_val,
            type=type_val,
            splices=(0 if is_rib else splices),
            device=device_name,
            splicer=splicer,
            created_date=created_date,
            company=company,
            project_id=project_id,
            map_role=map_role,
            included_splices_applied=included_applied,
            price_splices_usd=price_splices,
            price_device_usd=price_device,
            total_usd=total,
            ribbon_count=(ribbon_count if is_rib else None),
            billing_codes_json=(json.dumps(_bcodes, ensure_ascii=False) if _bcodes else None),
            ft_in=(None if type_val.strip().upper().startswith("CAN") else (ft_in or None)),
            ft_out=(None if type_val.strip().upper().startswith("CAN") else (ft_out or None)),
            can_cable_count=(can_cable_count if type_val.strip().upper().startswith("CAN") else None),
            can_cables_json=(json.dumps(can_cables, ensure_ascii=False) if type_val.strip().upper().startswith("CAN") and can_cables else None),
        )
        db.session.add(rec)
        db.session.commit()

        # Se for modo AJAX (usado quando existem fotos), responde rápido com o ID
        ajax_mode = (request.form.get("_ajax") == "1") or (request.args.get("ajax") == "1")
        if ajax_mode:
            # mantém seleções do modo rápido
            if is_splicer:
                session["entry_company"] = company
                if map_obj:
                    session["entry_map_id"] = int(map_obj.id)
                session["entry_map_role"] = map_role or ""
                session["entry_type"] = type_val or ""
            return jsonify({"ok": True, "record_id": int(rec.id)})

        files = request.files.getlist("photos")
        if files:
            for f in files[:5]:
                if not f or not f.filename:
                    continue
                filename, content_type, data, thumb_data, thumb_ct = process_uploaded_photo(f)
                if not data:
                    continue

                photo = RecordPhoto(
                    record_id=rec.id,
                    filename=filename[:255],
                    content_type=content_type,
                    data=data if data else b"",
                    thumb_data=thumb_data,
                    thumb_content_type=thumb_ct,
                    r2_key=None,
                    r2_thumb_key=None,
                    size_bytes=int(len(data) if data else 0),
                )
                db.session.add(photo)
                db.session.flush()

                if r2_enabled() and data:
                    try:
                        key = r2_key_for_record_photo(rec.id, filename)
                        thumb_key = key + ".thumb.jpg" if thumb_data else None
                        enqueue_r2_upload(int(photo.id), key, data, content_type, thumb_key=thumb_key, thumb_bytes=thumb_data, thumb_content_type=thumb_ct)
                    except Exception as e:
                        print(f"[R2] Failed to enqueue upload for photo id={getattr(photo,'id',None)}: {e}")
                        traceback.print_exc()

            db.session.commit()

        # Modo rápido: memoriza última seleção (somente para splicer)
        focus_map_id = None
        if is_splicer and company and (map_cfg or map_obj):
            last_map = map_cfg or map_obj
            session["entry_company"] = company
            session["entry_map_id"] = int(last_map.id)
            session["entry_map_role"] = map_role or ""
            session["entry_type"] = type_val or ""
            focus_map_id = int(last_map.id)

        flash("Lançamento salvo.", "success")

        # Se o usuário é splicer e temos um mapa associado, volta para o mapa
        # já focando no dispositivo recém-lançado.
        if focus_map_id is not None:
            return redirect(url_for("map_view", map_id=focus_map_id, focus_record=rec.id))

        return redirect(entry_url)

    # GET
    return render_template(
        "entry.html",
        splicer_options=splicer_options,
        companies=companies,
        projects_by_company=projects_by_company,
        maps_by_project=maps_by_project,
        maps_by_company=maps_by_company,
        maps_by_company_all=maps_by_company_all,
        devices_by_project=devices_by_project,
        devices_by_company=devices_by_company,
        default_splicer=default_splicer,
        today=date.today().isoformat(),
        form_company=pre_company or "",
        form_map_id=str(pre_map_id or ""),
        form_map_role=pre_map_role or "",
        form_type=pre_type or "",
        form_ft_in="",
        form_ft_out="",
        form_can_cable_count="",
        form_can_cables_json="[]",
        is_splicer=is_splicer,
    )





@app.route("/photo-entry", methods=["GET", "POST"])
@login_required
def photo_entry():
    """Lançamento de produção a partir de uma foto com carimbo (Timestamp).

    Versão beta:
      - Você pode deixar o campo de texto preenchido (4 linhas MAP/DEVICE/FUSION/T),
        ou marcar a opção de ler automaticamente com IA.
      - Se a IA estiver configurada (OPENAI_API_KEY), o sistema tenta ler os dados direto da foto.
      - Se algo falhar, cai no modo texto manual.
    """
    is_owner = bool(getattr(current_user, "is_company_owner", False))
    is_admin = bool(getattr(current_user, "is_admin", False))
    if is_owner and not is_admin:
        flash("Dono de empresa só pode visualizar. Utilize um usuário splicer para lançar produção.", "danger")
        return redirect(url_for("index"))


    if request.method == "POST":
        stamp_text = (request.form.get("stamp_text") or "").strip()
        use_ai = bool(request.form.get("use_ai"))

        # Aceita várias fotos; usa a primeira como base para IA/EXIF,
        # mas anexa todas ao lançamento.
        photos = [p for p in request.files.getlist("photos") if getattr(p, "filename", None)]
        if not photos:
            flash("Envie pelo menos uma foto.", "danger")
            return redirect(url_for("photo_entry"))

        main_photo = photos[0]

        raw_bytes = None
        try:
            raw_bytes = main_photo.read()
            try:
                main_photo.seek(0)
            except Exception:
                pass
        except Exception:
            raw_bytes = None

        parsed = None
        ai_error = None

        # Primeiro tenta IA, se marcado e se houver bytes da imagem.
        if use_ai and raw_bytes:
            parsed, ai_error = extract_timestamp_fields_with_ai(raw_bytes)
            if parsed is None and ai_error:
                flash(f"Não consegui ler automaticamente os dados da foto (IA): {ai_error}", "warning")

        # Se IA não foi usada ou falhou, tenta interpretar o texto manual.
        if parsed is None:
            if not stamp_text:
                flash("Cole o texto em 4 linhas (MAP / DEVICE / FUSION / T) ou ative a leitura automática por IA.", "danger")
                return redirect(url_for("photo_entry"))
            parsed = parse_timestamp_block_from_text(stamp_text)
            if not parsed:
                flash("Não consegui entender o texto. Confirme que está no formato MAP/DEVICE/FUSION/T.", "danger")
                return redirect(url_for("photo_entry"))
        map_name = parsed["map_name"]
        device_name = parsed["device_name"]
        splices_val = parsed["splices"]
        map_role = parsed["map_role"]  # 'MEIO' ou 'PONTA'
        # ft_in/ft_out: prioriza leitura do Claude, aceita override manual do form
        ft_in_ai = (parsed.get("ft_in") or "").strip()
        ft_out_ai = (parsed.get("ft_out") or "").strip()
        ft_in_form = (request.form.get("ft_in") or "").strip()
        ft_out_form = (request.form.get("ft_out") or "").strip()
        ft_in = ft_in_form or ft_in_ai or ""
        ft_out = ft_out_form or ft_out_ai or ""
        can_cable_count, can_cables = _parse_can_cables_from_form(request.form)
        photo_is_can = bool(request.form.get("is_can"))

        # Determina empresa/projeto com base em um Record existente desse dispositivo.
        base_query = Record.query.filter(Record.map == map_name, Record.device == device_name)
        user_company = getattr(current_user, "company_name", None) or getattr(current_user, "default_company", None)
        if user_company:
            base_query = base_query.filter(Record.company == user_company)
        base_rec = base_query.order_by(Record.id.asc()).first()
        if not base_rec:
            flash(f"Não encontrei nenhum dispositivo '{device_name}' no mapa '{map_name}'.", "danger")
            return redirect(url_for("photo_entry"))

        company = base_rec.company
        project_id = base_rec.project_id
        # IMPORTANT: muitos dispositivos importados (ou criados) podem estar com type NULL.
        # Para manter consistência e garantir cobrança do dispositivo no lançamento por foto,
        # o padrão do sistema é OTE (o usuário pode trocar manualmente para CAN depois).
        type_val = (base_rec.type or "OTE").strip() or "OTE"
        if photo_is_can and not str(type_val).upper().startswith("CAN"):
            type_val = "CAN"
        map_val = map_name

        # Busca o objeto CompanyMap para aplicar regras de MEIO/PONTA.
        map_obj = CompanyMap.query.filter(
            CompanyMap.company == company,
            CompanyMap.name == map_name,
            CompanyMap.project_id == project_id,
        ).order_by(CompanyMap.id.asc()).first()

        included_override, included_applied, map_cfg = resolve_included_override(
            company=company,
            project_id=project_id,
            map_obj=map_obj,
            map_val=map_val,
            map_role=map_role,
        )

        device_for_price = type_val or device_name

        # Ribbon: le ribbon_count do form (lancamento por foto pode ter o campo manual)
        ribbon_count_photo_raw = (request.form.get("ribbon_count") or "").strip()
        try:
            ribbon_count_photo = int(ribbon_count_photo_raw) if ribbon_count_photo_raw else None
        except ValueError:
            ribbon_count_photo = None

        price_splices, price_device, total = compute_prices(
            splices=splices_val,
            device_name=device_for_price,
            company=company,
            project_id=project_id,
            included_override=included_override,
            map_role=map_role,
            ribbon_count=ribbon_count_photo,
        )

        is_rib_photo, _ = device_is_ribbon(device_for_price or "", company, project_id)
        _bcodes_photo = compute_billing_codes(splices_val, device_for_price or "", company, project_id, map_role=map_role, ribbon_count=ribbon_count_photo)

        rec = base_rec
        # Garante que o lançamento fique com type preenchido (default OTE).
        rec.type = type_val
        rec.splicer = (getattr(current_user, "splicer_name", None) or current_user.username)
        rec.map_role = map_role
        rec.splices = (0 if is_rib_photo else splices_val)
        rec.ribbon_count = (ribbon_count_photo if is_rib_photo else None)
        rec.billing_codes_json = json.dumps(_bcodes_photo, ensure_ascii=False) if _bcodes_photo else None
        rec.price_splices_usd = price_splices
        rec.price_device_usd = price_device
        rec.total_usd = total
        rec.included_splices_applied = included_applied
        if str(type_val).upper().startswith("CAN"):
            rec.ft_in = None
            rec.ft_out = None
            rec.can_cable_count = can_cable_count or None
            rec.can_cables_json = json.dumps(can_cables, ensure_ascii=False) if can_cables else None
        else:
            rec.ft_in = ft_in or None
            rec.ft_out = ft_out or None
            rec.can_cable_count = None
            rec.can_cables_json = None

        # Data de criação: tenta usar a data EXIF da foto.
        photo_date = extract_photo_date_from_exif(raw_bytes) if raw_bytes else None
        if photo_date:
            rec.created_date = datetime.combine(photo_date, datetime.min.time())
        elif not rec.created_date:
            rec.created_date = datetime.utcnow()


        # Processa e anexa as fotos ao lançamento.
        for p in photos:
            try:
                filename, content_type, data_bytes, thumb_bytes, thumb_ct = process_uploaded_photo(p)
            except Exception:
                continue
            if data_bytes:
                photo_obj = RecordPhoto(
                    record_id=rec.id,
                    filename=(filename or "foto.jpg")[:255],
                    content_type=content_type,
                    data=data_bytes,
                    thumb_data=thumb_bytes,
                    thumb_content_type=thumb_ct,
                )
                db.session.add(photo_obj)

        db.session.commit()
        flash(f"Lançamento atualizado para {map_name} / {device_name} com {splices_val} fusões e {len(photos)} foto(s).", "success")
        # Depois de lançar por foto, volta para o mapa correspondente,
        # focando no dispositivo que acabou de ser atualizado.
        if map_obj and getattr(map_obj, "id", None):
            return redirect(url_for("map_view", map_id=map_obj.id, focus_record=rec.id))
        return redirect(url_for("index"))

    # GET: monta o texto padrão (MAP/DEVICE/FUSION/T) se vierem dados via querystring.
    if request.method == "GET":
        map_name_q = (request.args.get("map_name") or "").strip()
        device_name_q = (request.args.get("device_name") or "").strip()

        stamp_text = ""
        if map_name_q or device_name_q:
            # Monta o bloco em 4 linhas, deixando FUSION e T para você preencher.
            lines = []
            lines.append(f"MAP: {map_name_q}" if map_name_q else "MAP:")
            lines.append(f"DEVICE: {device_name_q}" if device_name_q else "DEVICE:")
            lines.append("FUSION: ")
            lines.append("T: ")
            stamp_text = "\n".join(lines)

        return render_template(
            "photo_entry.html",
            stamp_text=stamp_text,
            form_ft_in="",
            form_ft_out="",
            form_can_cable_count="",
            form_can_cables_json="[]",
            device_type="",
        )

    return render_template(
        "photo_entry.html",
        stamp_text="",
        form_ft_in="",
        form_ft_out="",
        form_can_cable_count="",
        form_can_cables_json="[]",
        device_type="",
    )


@app.route("/entry-focus/<int:rid>")
@login_required
def entry_focus_redirect(rid: int):
    """
    Redireciona para a tela correta conforme o tipo de usuário:
    - Admin: tela de edição completa (/record/<id>/edit)
    - Splicer / Dono de empresa: tela de visualização (/record/<id>/view)
    """
    if bool(getattr(current_user, "is_admin", False)):
        return redirect(url_for("record_edit", rid=rid))
    return redirect(url_for("record_view", rid=rid))


@app.route("/record/<int:rid>/edit", methods=["GET", "POST"])
@login_required
def record_edit(rid):
    """Editar um lançamento existente."""
    rec = Record.query.get_or_404(rid)

    # Permissões de edição:
    # - Apenas Admin pode editar lançamentos de dispositivos já criados.
    # - Dono de empresa e splicer comum não podem editar.
    is_admin = bool(getattr(current_user, "is_admin", False))
    is_owner = bool(getattr(current_user, "is_company_owner", False))

    if not is_admin:
        flash("Apenas administradores podem editar lançamentos de dispositivos.", "danger")
        abort(403)

    companies = [c.name for c in CompanyConfig.query.order_by(CompanyConfig.name).all()]

    projects_by_company = {}
    for pr in Project.query.order_by(Project.company, Project.name).all():
        projects_by_company.setdefault(pr.company, []).append({"id": pr.id, "name": pr.name})

    maps_by_project = {}
    maps_by_company = {}
    maps_by_company_all = {}
    for m in CompanyMap.query.order_by(CompanyMap.company, CompanyMap.name).all():
        # Mantém o mesmo formato usado em /entry (lista de objetos),
        # para o JS funcionar também na tela de edição.
        mobj = {
            "id": int(m.id),
            "name": m.name,
            "project_id": int(m.project_id) if m.project_id else None,
            "mid_end_enabled": bool(getattr(m, "mid_end_enabled", False)),
            "included_splices_meio": int(getattr(m, "included_splices_meio", 0) or 0),
            "included_splices_ponta": int(getattr(m, "included_splices_ponta", 0) or 0),
        }
        maps_by_company_all.setdefault(m.company, []).append(mobj)

        if m.project_id:
            maps_by_project.setdefault(str(m.project_id), []).append(mobj)
        else:
            maps_by_company.setdefault(m.company, []).append(mobj)

    devices_by_project = {}
    devices_by_company = {}
    for dt in DeviceType.query.order_by(DeviceType.company, DeviceType.name).all():
        dobj = {"name": dt.name, "is_ribbon": bool(getattr(dt, "is_ribbon", False))}
        if dt.project_id:
            devices_by_project.setdefault(str(dt.project_id), []).append(dobj)
        else:
            key = dt.company or "__global__"
            devices_by_company.setdefault(key, []).append(dobj)

    splicer_options = []
    if getattr(current_user, "is_admin", False):
        splicer_options = [
            (u.splicer_name or u.username)
            for u in User.query.order_by(User.username).all()
            if (u.splicer_name or u.username)
        ]

    default_splicer = getattr(current_user, "splicer_name", None) or current_user.username

    if request.method == "POST":
        company = (request.form.get("company") or "").strip() or None
        project_id_raw = (request.form.get("project_id") or "").strip()
        project_id = int(project_id_raw) if project_id_raw.isdigit() else None
        map_val = (request.form.get("map") or "").strip()
        type_val = (request.form.get("type") or "").strip()
        device_name = (request.form.get("device_name") or "").strip()

        device_for_price = type_val or device_name

        splices_raw = request.form.get("splices") or "0"
        ribbon_count_raw = (request.form.get("ribbon_count") or "").strip()
        created_raw = request.form.get("created") or ""
        splicer = (request.form.get("splicer") or "").strip() or default_splicer
        ft_in = (request.form.get("ft_in") or "").strip()
        ft_out = (request.form.get("ft_out") or "").strip()
        can_cable_count, can_cables = _parse_can_cables_from_form(request.form)

        try:
            splices = int(splices_raw or 0)
        except ValueError:
            splices = 0

        try:
            ribbon_count = int(ribbon_count_raw) if ribbon_count_raw else None
        except ValueError:
            ribbon_count = None

        if created_raw:
            try:
                created_date = datetime.strptime(created_raw, "%Y-%m-%d")
            except ValueError:
                created_date = datetime.utcnow()
        else:
            today = date.today()
            created_date = datetime(today.year, today.month, today.day)

        # --- Regras MEIO/PONTA também na edição ---
        map_role = (request.form.get("map_role") or rec.map_role or "").strip().upper() or None

        map_cfg = None
        if map_val and company:
            q = CompanyMap.query.filter_by(company=company, name=map_val)
            if project_id is not None:
                q = q.filter_by(project_id=project_id)
            map_cfg = q.first()
            if not map_cfg:
                map_cfg = CompanyMap.query.filter_by(company=company, name=map_val, project_id=None).first()

        if map_cfg and bool(getattr(map_cfg, "mid_end_enabled", False)) and map_role not in ("MEIO", "PONTA"):
            flash("Este mapa exige selecionar MEIO ou PONTA.", "danger")
            return redirect(url_for("record_edit", rid=rec.id))

        included_override, included_applied, map_cfg = resolve_included_override(
            company,
            project_id,
            map_cfg,
            map_val,
            map_role,
        )

        price_splices, price_device, total = compute_prices(
            splices,
            device_for_price,
            company,
            project_id,
            included_override=included_override,
            map_role=map_role,
            ribbon_count=ribbon_count,
        )

        is_rib_edit, _ = device_is_ribbon(device_for_price or "", company, project_id)
        _bcodes_edit = compute_billing_codes(splices, device_for_price or "", company, project_id, map_role=map_role, ribbon_count=ribbon_count)
        rec.company = company
        rec.project_id = project_id
        rec.map = map_val
        rec.type = type_val
        rec.device = device_name
        rec.splices = (0 if is_rib_edit else splices)
        rec.ribbon_count = (ribbon_count if is_rib_edit else None)
        rec.billing_codes_json = json.dumps(_bcodes_edit, ensure_ascii=False) if _bcodes_edit else None
        rec.splicer = splicer
        rec.created_date = created_date
        rec.map_role = map_role
        rec.included_splices_applied = included_applied
        rec.price_splices_usd = price_splices
        rec.price_device_usd = price_device
        rec.total_usd = total
        if type_val.strip().upper().startswith("CAN"):
            rec.ft_in = None
            rec.ft_out = None
            rec.can_cable_count = can_cable_count
            rec.can_cables_json = json.dumps(can_cables, ensure_ascii=False) if can_cables else None
        else:
            rec.ft_in = ft_in or None
            rec.ft_out = ft_out or None
            rec.can_cable_count = None
            rec.can_cables_json = None

        db.session.commit()
        flash("Lançamento atualizado.", "success")

        # Resposta especial para AJAX (formulário com fotos)
        if request.form.get("_ajax") == "1":
            return jsonify({"ok": True, "record_id": rec.id})

        # Após salvar edição, se for splicer (não-admin), volta para o mapa
        is_admin = bool(getattr(current_user, "is_admin", False))
        if not is_admin and rec.map:
            mp_target = CompanyMap.query.filter_by(name=rec.map, company=rec.company).first()
            if mp_target is not None:
                return redirect(url_for("map_view", map_id=mp_target.id, focus_record=rec.id))

        # Comportamento padrão: volta para a página de visualização do dispositivo
        return redirect(url_for("record_view", rid=rec.id))

    form_created = rec.created_date.date().isoformat() if rec.created_date else date.today().isoformat()

    # Modo de edição:
    # - Admin vê tela completa
    # - Splicer (não-admin) usa o mesmo modo rápido do /entry (projeto derivado pelo mapa)
    is_admin = bool(getattr(current_user, "is_admin", False))
    is_splicer = not is_admin

    # Tenta resolver o mapa associado a este record para pré-selecionar no combo
    pre_map_id = None
    if rec.map:
        mp_for_rec = CompanyMap.query.filter_by(name=rec.map, company=rec.company).first()
        if mp_for_rec is not None:
            pre_map_id = mp_for_rec.id

    return render_template(
        "entry.html",
        splicer_options=splicer_options,
        companies=companies,
        projects_by_company=projects_by_company,
        maps_by_project=maps_by_project,
        devices_by_project=devices_by_project,
        maps_by_company=maps_by_company,
        maps_by_company_all=maps_by_company_all,
        devices_by_company=devices_by_company,
        default_splicer=default_splicer,
        today=date.today().isoformat(),
        is_edit=True,
        record=rec,
        is_splicer=is_splicer,
        form_company=rec.company,
        form_project_id=str(rec.project_id or ""),
        form_map_id=str(pre_map_id or ""),
        form_map_role=getattr(rec, "map_role", None) or "",
        form_type=rec.type,
        form_device_name=rec.device,
        form_splices=str(rec.splices or 0),
        form_created=form_created,
        form_ft_in=rec.ft_in or "",
        form_ft_out=rec.ft_out or "",
        form_can_cable_count=str(rec.can_cable_count or ""),
        form_can_cables_json=rec.can_cables_json or "[]",
        form_ribbon_count=str(rec.ribbon_count or ""),
    )




@app.route("/record/photo/<int:photo_id>/remove", methods=["POST"])
@login_required
def photo_remove_v2(photo_id: int):
    """Exclui uma foto específica de um lançamento.

    Regras:
    - Admin pode sempre excluir;
    - Dono de empresa (company_owner) nunca pode excluir;
    - Splicer comum pode excluir se tiver acesso ao mapa do registro
      e for o splicer responsável pelo lançamento (quando preenchido).
    """

    photo = RecordPhoto.query.get_or_404(photo_id)
    rec = photo.record

    is_admin = bool(getattr(current_user, "is_admin", False))
    is_owner = bool(getattr(current_user, "is_company_owner", False))

    # Dono de empresa nunca pode excluir
    if is_owner and not is_admin:
        abort(403)

    if not is_admin:
        # Garante acesso ao mapa (mesma lógica usada em outras rotas)
        mp = None
        if rec.map:
            mp = CompanyMap.query.filter_by(name=rec.map, company=rec.company).first()
        if mp is not None:
            ensure_map_access(mp)

        # Se o lançamento tem splicer definido, precisa ser o mesmo usuário
        current_splicer = (getattr(current_user, "splicer_name", None) or current_user.username or "").strip()
        if rec.splicer and rec.splicer.strip() and rec.splicer.strip() != current_splicer:
            abort(403)

    db.session.delete(photo)
    db.session.commit()
    flash("Foto removida.", "success")
    return redirect(request.referrer or url_for("record_view", rid=rec.id))



@app.route("/record/<int:rid>/view")
@login_required
def record_view(rid):
    """Página de detalhes de um dispositivo (registro) para ver e lançar testes.

    Regra especial:
    - Admin: pode abrir normalmente a tela completa do dispositivo;
    - Dono de empresa (cliente): é redirecionado para o mapa com foco no dispositivo;
    - Splicer: pode abrir, desde que tenha acesso ao mapa / empresa.
    """
    rec = Record.query.get_or_404(rid)

    is_admin = bool(getattr(current_user, "is_admin", False))
    is_owner = bool(getattr(current_user, "is_company_owner", False))
    can_view_values = is_admin or bool(getattr(current_user, "can_view_values", True))

    # Admin => tela completa
    if is_admin:
        mp = None
        if rec.map:
            mp = CompanyMap.query.filter_by(name=rec.map, company=rec.company).first()
        map_id_for_button = mp.id if mp is not None else None
        return render_template("record_view.html", rec=rec, map_id_for_button=map_id_for_button,
                               can_view_values=can_view_values)

    # Dono de empresa (cliente) => manda para o mapa com foco no dispositivo
    if is_owner and rec.map:
        mp = CompanyMap.query.filter_by(name=rec.map, company=rec.company).first()
        if mp is not None:
            return redirect(url_for("map_view", map_id=mp.id, focus_record=rec.id))

    # 2) Tenta resolver o mapa desse record e usar a mesma lógica de acesso de mapas
    mp = None
    if rec.map:
        mp = CompanyMap.query.filter_by(name=rec.map, company=rec.company).first()

    if mp is not None:
        # Usa a regra centralizada de acesso a mapas (admin/owner/splicers/sem lista => liberado)
        ensure_map_access(mp)
    else:
        # Fallback: se não tiver mapa associado, garante pelo menos que seja da mesma empresa
        if _current_user_company_name() != rec.company:
            # Também permitimos o próprio splicer do registro
            current_splicer = (getattr(current_user, "splicer_name", None) or current_user.username)
            if (rec.splicer or "") != current_splicer:
                abort(403)

    map_id_for_button = mp.id if mp is not None else None
    return render_template("record_view.html", rec=rec, map_id_for_button=map_id_for_button,
                           can_view_values=can_view_values)
@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("login"))



@app.route("/record/<int:record_id>/photos", methods=["POST"])
@login_required
def record_upload_photos(record_id):
    """
    Upload de fotos adicionais para um lançamento já existente.

    Este endpoint é usado pelo formulário de lançamento (entry.html)
    quando há fotos na fila. Ele valida o acesso ao mapa / empresa,
    processa todas as imagens recebidas e retorna JSON com o total salvo.
    """
    rec = Record.query.get_or_404(record_id)

    # Reaproveita a mesma lógica de permissão do record_view
    is_admin = getattr(current_user, "is_admin", False)
    is_owner = getattr(current_user, "is_company_owner", False)

    if not is_admin:
        # Usuário normal: precisa ter acesso ao mapa ou ser o splicer do registro
        mp = None
        if rec.map:
            mp = CompanyMap.query.filter_by(name=rec.map, company=rec.company).first()

        if mp is not None:
            ensure_map_access(mp)
        else:
            # Se não há mapa, cai na mesma regra do splicer da company
            if _current_user_company_name() != rec.company:
                current_splicer = (getattr(current_user, "splicer_name", None) or current_user.username)
                if (rec.splicer or "") != current_splicer:
                    abort(403)

    files = request.files.getlist("photos")
    if not files:
        return jsonify({"ok": True, "count": 0})

    is_test = bool(rec.is_test_for_device)
    added = 0
    for f in files:
        if not f or not getattr(f, "filename", None):
            continue
        if not f.filename.strip():
            continue

        processed = process_uploaded_photo(f, is_test=is_test)
        if not processed:
            continue

        photo = RecordPhoto(record_id=rec.id, **processed)
        db.session.add(photo)
        added += 1

    if added:
        db.session.commit()

    return jsonify({"ok": True, "count": added})


@app.route("/my-maps", methods=["GET"])
@login_required


def my_maps():
    """
    Tela para o usuário escolher em qual mapa quer trabalhar.

    Regras de visibilidade:
    - Admin: vê todos os mapas;
    - Dono da empresa (is_company_owner): vê SOMENTE os mapas da própria empresa;
    - Splicer comum:
        * Se o mapa não tiver nenhum splicer configurado (allowed_splicers vazio),
          o mapa aparece e pode ser usado;
        * Se o mapa tiver allowed_splicers configurado, ele só aparece se o
          usuário estiver nessa lista.
    """
    query = CompanyMap.query
    user_company = None

    if not current_user.is_admin:
        if getattr(current_user, "is_company_owner", False):
            # Dono da empresa só vê mapas da própria empresa
            user_company = (getattr(current_user, "company_name", None) or "").strip()
            if user_company:
                query = query.filter(CompanyMap.company == user_company)
            else:
                # Se não houver empresa vinculada, não mostra nenhum mapa
                query = query.filter(db.text("1=0"))
        else:
            # Splicer comum: aplica regras de allowed_splicers
            query = query.filter(
                or_(
                    ~CompanyMap.allowed_splicers.any(),
                    CompanyMap.allowed_splicers.any(id=current_user.id),
                )
            )

    maps = query.order_by(CompanyMap.company, CompanyMap.name).all()

    return render_template(
        "my_maps.html",
        maps=maps,
        user_company=user_company,
    )


@app.route("/settings", methods=["GET"])
@admin_required
def settings():
    """Tela principal de cadastro de empresas."""
    companies = CompanyConfig.query.order_by(CompanyConfig.name).all()
    syscfg = SystemConfig.query.first()
    if not syscfg:
        syscfg = SystemConfig()
        db.session.add(syscfg)
        db.session.commit()
    return render_template("settings.html", companies=companies, syscfg=syscfg)


@app.route("/settings/backup-db", methods=["GET"])
@login_required
@admin_required
def settings_backup_db():
    """Gera um backup do arquivo de banco de dados SQLite e envia para download."""
    try:
        from auto_migrate_all_dbs import find_candidate_dbs
    except ImportError:
        find_candidate_dbs = None

    db_path = None

    # Primeiro, tenta localizar o mesmo caminho usado pelo SQLAlchemy.
    db_url = app.config.get("SQLALCHEMY_DATABASE_URI", "") or ""
    if db_url.startswith("sqlite:///"):
        rel_path = db_url[len("sqlite:///"):]
        base_dir = os.path.dirname(os.path.abspath(__file__))
        candidate = os.path.join(base_dir, rel_path)
        if os.path.exists(candidate):
            db_path = candidate

    # Se ainda não encontrou, tenta usar o utilitário de migração (procura *.db* dentro do projeto).
    if db_path is None and find_candidate_dbs is not None:
        try:
            dbs = find_candidate_dbs()
        except Exception:
            dbs = []
        if dbs:
            # Se existir mais de um, pega o primeiro por simplicidade.
            db_path = dbs[0]

    if not db_path or not os.path.exists(db_path):
        flash("Não foi possível localizar um arquivo de banco de dados (.db) para backup.", "error")
        return redirect(url_for("settings"))

    import tempfile, shutil, datetime

    tmp_dir = tempfile.mkdtemp(prefix="splice-backup-")
    timestamp = datetime.datetime.utcnow().strftime("%Y%m%d-%H%M%S")
    base_name = os.path.basename(db_path)
    backup_name = f"splice-backup-{timestamp}-{base_name}"
    backup_path = os.path.join(tmp_dir, backup_name)

    try:
        shutil.copy2(db_path, backup_path)
    except Exception as e:
        flash(f"Erro ao gerar backup: {e}", "error")
        return redirect(url_for("settings"))

    # Envia o arquivo para o navegador fazer o download.
    return send_file(backup_path, as_attachment=True, download_name=backup_name, mimetype="application/octet-stream")


@app.route("/settings/company/add", methods=["POST"])
@login_required
def settings_company_add():
    name = (request.form.get("name") or "").strip()
    included_raw = request.form.get("included_splices") or "0"
    included = int(included_raw or 0)
    invoice_address = (request.form.get("invoice_address") or "").strip() or None

    if not name:
        flash("Nome da empresa é obrigatório.", "danger")
        return redirect(url_for("settings"))

    cfg = CompanyConfig.query.filter_by(name=name).first()
    if cfg:
        cfg.included_splices = included
        cfg.invoice_address = invoice_address
    else:
        cfg = CompanyConfig(name=name, included_splices=included, invoice_address=invoice_address)
        db.session.add(cfg)
    db.session.commit()
    flash("Empresa / fusões inclusas salva.", "success")
    return redirect(url_for("settings"))


@app.route("/settings/company/<int:cid>/rename", methods=["POST"])
@admin_required
def settings_company_rename(cid: int):
    """Renomeia uma empresa e atualiza todas as tabelas que referenciam o nome."""
    company = CompanyConfig.query.get_or_404(cid)
    new_name = (request.form.get("new_name") or "").strip()

    if not new_name:
        flash("Nome não pode ser vazio.", "danger")
        return redirect(url_for("settings"))

    if new_name == company.name:
        return redirect(url_for("settings"))

    # Verifica conflito
    if CompanyConfig.query.filter_by(name=new_name).first():
        flash(f"Já existe uma empresa com o nome '{new_name}'.", "danger")
        return redirect(url_for("settings"))

    old_name = company.name

    # Atualiza em cascata todas as tabelas que usam company como string
    for model_cls, col_attr in [
        (Project,    Project.company),
        (Invoice,    Invoice.company),
        (DeviceType, DeviceType.company),
        (SpliceTier, SpliceTier.company),
        (CompanyMap, CompanyMap.company),
        (Record,     Record.company),
        (Payroll,    Payroll.company),
    ]:
        try:
            model_cls.query.filter(col_attr == old_name).update(
                {col_attr.key: new_name}, synchronize_session="fetch"
            )
        except Exception:
            pass

    # Atualiza também usuários company_owner vinculados a esta empresa
    try:
        User.query.filter(User.company_name == old_name).update(
            {"company_name": new_name}, synchronize_session="fetch"
        )
    except Exception:
        pass

    company.name = new_name
    db.session.commit()
    flash(f"Empresa renomeada de '{old_name}' para '{new_name}'.", "success")
    return redirect(url_for("settings"))


@app.route("/settings/company/<int:cid>/delete", methods=["POST"])
@admin_required
def settings_company_delete(cid: int):
    """Exclui uma empresa se não houver registros vinculados."""
    company = CompanyConfig.query.get_or_404(cid)
    name = company.name
    force = request.form.get("force") == "1"

    # Conta vínculos
    n_records  = Record.query.filter_by(company=name).count()
    n_projects = Project.query.filter_by(company=name).count()
    n_invoices = Invoice.query.filter_by(company=name).count()
    n_payrolls = Payroll.query.filter_by(company=name).count()
    total = n_records + n_projects + n_invoices + n_payrolls

    if total > 0 and not force:
        flash(
            f"A empresa '{name}' possui {n_records} lançamento(s), "
            f"{n_projects} projeto(s), {n_invoices} invoice(s) e "
            f"{n_payrolls} payroll(s) vinculados. "
            f"Confirme a exclusão marcando a opção abaixo.",
            "warning"
        )
        return redirect(url_for("settings", confirm_delete=cid))

    # Exclui registros dependentes em cascata
    if force:
        try:
            Payroll.query.filter_by(company=name).delete()
            Invoice.query.filter_by(company=name).delete()
            # Records e fotos
            rids = [r.id for r in Record.query.filter_by(company=name).all()]
            if rids:
                RecordPhoto.query.filter(RecordPhoto.record_id.in_(rids)).delete(synchronize_session=False)
                Record.query.filter_by(company=name).delete()
            # Mapa, devices, tiers, projetos
            CompanyMap.query.filter_by(company=name).delete()
            DeviceType.query.filter_by(company=name).delete()
            SpliceTier.query.filter_by(company=name).delete()
            # Projetos e suas tabelas de splicer
            for p in Project.query.filter_by(company=name).all():
                for sp in SplicerPricing.query.filter_by(project_id=p.id).all():
                    db.session.delete(sp)
                db.session.delete(p)
        except Exception as e:
            db.session.rollback()
            flash(f"Erro ao excluir dados vinculados: {e}", "danger")
            return redirect(url_for("settings"))

    db.session.delete(company)
    db.session.commit()
    flash(f"Empresa '{name}' excluída com sucesso.", "success")
    return redirect(url_for("settings"))


@app.route("/settings/company/<int:cid>", methods=["GET", "POST"])
@admin_required
def settings_company_detail(cid: int):
    company = CompanyConfig.query.get_or_404(cid)

    # exclusão de mapa via querystring
    del_map_id = request.args.get("del_map")
    if del_map_id:
        mp = CompanyMap.query.get(int(del_map_id))
        if mp and mp.company == company.name:
            db.session.delete(mp)
            db.session.commit()
            flash("Mapa removido.", "success")
        return redirect(url_for("settings_company_detail", cid=company.id))

    # inclusão de mapa via POST
    if request.method == "POST":
        action = (request.form.get("action") or "").strip()

        # ── Renomear empresa ──
        if action == "rename":
            new_name = (request.form.get("new_name") or "").strip()
            if not new_name:
                flash("Nome não pode ser vazio.", "danger")
                return redirect(url_for("settings_company_detail", cid=cid))
            if new_name == company.name:
                return redirect(url_for("settings_company_detail", cid=cid))
            if CompanyConfig.query.filter_by(name=new_name).first():
                flash(f"Já existe uma empresa com o nome '{new_name}'.", "danger")
                return redirect(url_for("settings_company_detail", cid=cid))
            old_name = company.name
            for model_cls, col_attr in [
                (Project,    Project.company),
                (Invoice,    Invoice.company),
                (DeviceType, DeviceType.company),
                (SpliceTier, SpliceTier.company),
                (CompanyMap, CompanyMap.company),
                (Record,     Record.company),
                (Payroll,    Payroll.company),
            ]:
                try:
                    model_cls.query.filter(col_attr == old_name).update(
                        {col_attr.key: new_name}, synchronize_session="fetch"
                    )
                except Exception:
                    pass
            try:
                User.query.filter(User.company_name == old_name).update(
                    {"company_name": new_name}, synchronize_session="fetch"
                )
            except Exception:
                pass
            company.name = new_name
            db.session.commit()
            flash(f"Empresa renomeada para '{new_name}'.", "success")
            return redirect(url_for("settings_company_detail", cid=cid))

        # ── Excluir empresa ──
        if action == "delete":
            name = company.name
            force = request.form.get("force") == "1"
            n_records  = Record.query.filter_by(company=name).count()
            n_projects = Project.query.filter_by(company=name).count()
            n_invoices = Invoice.query.filter_by(company=name).count()
            n_payrolls = Payroll.query.filter_by(company=name).count()
            total = n_records + n_projects + n_invoices + n_payrolls
            if total > 0 and not force:
                flash(
                    f"A empresa possui {n_records} lançamento(s), "
                    f"{n_projects} projeto(s), {n_invoices} invoice(s) e "
                    f"{n_payrolls} payroll(s). Confirme a exclusão abaixo.",
                    "warning"
                )
                return redirect(url_for("settings_company_detail", cid=cid, confirm_delete=1))
            if force or total == 0:
                try:
                    Payroll.query.filter_by(company=name).delete()
                    Invoice.query.filter_by(company=name).delete()
                    rids = [r.id for r in Record.query.filter_by(company=name).all()]
                    if rids:
                        RecordPhoto.query.filter(RecordPhoto.record_id.in_(rids)).delete(synchronize_session=False)
                        Record.query.filter_by(company=name).delete()
                    CompanyMap.query.filter_by(company=name).delete()
                    DeviceType.query.filter_by(company=name).delete()
                    SpliceTier.query.filter_by(company=name).delete()
                    for p in Project.query.filter_by(company=name).all():
                        for sp in SplicerPricing.query.filter_by(project_id=p.id).all():
                            db.session.delete(sp)
                        db.session.delete(p)
                    db.session.delete(company)
                    db.session.commit()
                    flash(f"Empresa '{name}' excluída.", "success")
                    return redirect(url_for("settings"))
                except Exception as e:
                    db.session.rollback()
                    flash(f"Erro ao excluir: {e}", "danger")
                    return redirect(url_for("settings_company_detail", cid=cid))

        # ── Adicionar mapa ──
        new_map = (request.form.get("new_map") or "").strip()
        if new_map:
            exists = CompanyMap.query.filter_by(company=company.name, name=new_map).first()
            if not exists:
                db.session.add(CompanyMap(company=company.name, name=new_map))
                db.session.commit()
                flash("Mapa adicionado.", "success")
        return redirect(url_for("settings_company_detail", cid=company.id))

    types = DeviceType.query.filter_by(company=company.name, project_id=None).order_by(DeviceType.name).all()
    tiers = SpliceTier.query.filter_by(company=company.name, project_id=None).order_by(SpliceTier.min_splices).all()
    maps = CompanyMap.query.filter_by(company=company.name, project_id=None).order_by(CompanyMap.name).all()
    projects = Project.query.filter_by(company=company.name).order_by(Project.name).all()
    return render_template(
        "settings_company.html",
        company=company,
        projects=projects,
        types=types,
        tiers=tiers,
        maps=maps,
    )



@app.route("/settings/system", methods=["POST"])
@admin_required
def settings_system_update():
    """Atualiza os dados da sua empresa (emitente da invoice)."""
    name = (request.form.get("my_company_name") or "").strip() or None
    addr = (request.form.get("my_company_address") or "").strip() or None
    taxid = (request.form.get("my_company_tax_id") or "").strip() or None
    email = (request.form.get("my_company_email") or "").strip() or None
    phone = (request.form.get("my_company_phone") or "").strip() or None
    geoapify_api_key = (request.form.get("geoapify_api_key") or "").strip() or None
    board_header = (request.form.get("board_header") or "").strip() or None

    cfg = SystemConfig.query.first()
    if not cfg:
        cfg = SystemConfig()
        db.session.add(cfg)

    cfg.my_company_name = name
    cfg.my_company_address = addr
    cfg.my_company_tax_id = taxid
    cfg.my_company_email = email
    cfg.my_company_phone = phone
    cfg.geoapify_api_key = geoapify_api_key
    cfg.board_header = board_header

    db.session.commit()
    flash("Dados da sua empresa atualizados.", "success")
    return redirect(url_for("settings"))


@app.route("/settings/project/add", methods=["POST"])
@admin_required
def settings_project_add():
    company = (request.form.get("company") or "").strip()
    project_name = (request.form.get("project_name") or "").strip()

    if not company or not project_name:
        flash("Empresa e nome do projeto são obrigatórios.", "danger")
        return redirect(url_for("settings"))

    existing = Project.query.filter_by(company=company, name=project_name).first()
    if existing:
        flash("Projeto já existe.", "warning")
        comp = CompanyConfig.query.filter_by(name=company).first()
        return redirect(url_for("settings_company_detail", cid=comp.id if comp else 0))

    pr = Project(company=company, name=project_name, included_splices=None)
    db.session.add(pr)
    db.session.commit()

    # Copia dispositivos e faixas da empresa (defaults) para servir como base do projeto
    base_types = DeviceType.query.filter_by(company=company, project_id=None).all()
    for dt in base_types:
        db.session.add(DeviceType(name=dt.name, value_usd=dt.value_usd, company=company, project_id=pr.id))

    base_tiers = SpliceTier.query.filter_by(company=company, project_id=None).all()
    for t in base_tiers:
        db.session.add(
            SpliceTier(
                company=company,
                project_id=pr.id,
                min_splices=t.min_splices,
                max_splices=t.max_splices,
                price_per_splice_usd=t.price_per_splice_usd,
            )
        )

    db.session.commit()
    flash("Projeto criado (valores copiados da empresa).", "success")
    return redirect(url_for("settings_project_detail", pid=pr.id))


@app.route("/settings/project/<int:pid>", methods=["GET", "POST"])
@admin_required
def settings_project_detail(pid: int):
    project = Project.query.get_or_404(pid)

    # usado no link "Voltar para empresa" na página do projeto
    comp_cfg = CompanyConfig.query.filter_by(name=project.company).first()
    company_id = comp_cfg.id if comp_cfg else 0

    del_map_id = request.args.get("del_map")
    if del_map_id:
        mp = CompanyMap.query.get(int(del_map_id))
        if mp and mp.project_id == project.id:
            db.session.delete(mp)
            db.session.commit()
            flash("Mapa removido.", "success")
        return redirect(url_for("settings_project_detail", pid=project.id))

    if request.method == "POST":
        if request.form.get("action") == "update_project":
            inc_raw = (request.form.get("included_splices") or "").strip()
            project.included_splices = int(inc_raw) if inc_raw != "" else None
            pay_raw = (request.form.get("payment_days") or "").strip()
            project.payment_days = int(pay_raw) if pay_raw.isdigit() else 30
            db.session.commit()
            flash("Projeto atualizado.", "success")
            return redirect(url_for("settings_project_detail", pid=project.id))

        if request.form.get("action") == "update_map_rules":
            mid_raw = (request.form.get("included_meio") or "").strip()
            ponta_raw = (request.form.get("included_ponta") or "").strip()
            enabled = bool(request.form.get("mid_end_enabled"))
            mid = int(mid_raw) if mid_raw != "" else 0
            ponta = int(ponta_raw) if ponta_raw != "" else 0
            mid = max(mid, 0)
            ponta = max(ponta, 0)
            map_id_raw = (request.form.get("map_id") or "").strip()
            if map_id_raw.isdigit():
                mp = CompanyMap.query.get(int(map_id_raw))
                if mp and mp.project_id == project.id:
                    mp.mid_end_enabled = enabled
                    mp.included_splices_meio = mid
                    mp.included_splices_ponta = ponta
                    db.session.commit()
                    flash("Mapa atualizado.", "success")
            return redirect(url_for("settings_project_detail", pid=project.id))

        new_map = (request.form.get("new_map") or "").strip()
        if new_map:
            exists = CompanyMap.query.filter_by(company=project.company, name=new_map, project_id=project.id).first()
            if not exists:
                db.session.add(CompanyMap(company=project.company, name=new_map, project_id=project.id))
                db.session.commit()
                flash("Mapa adicionado.", "success")
        return redirect(url_for("settings_project_detail", pid=project.id))

    types = DeviceType.query.filter_by(company=project.company, project_id=project.id).order_by(DeviceType.name).all()
    tiers = SpliceTier.query.filter_by(company=project.company, project_id=project.id).order_by(SpliceTier.min_splices).all()
    maps = CompanyMap.query.filter_by(company=project.company, project_id=project.id).order_by(CompanyMap.name).all()

    return render_template(
        "settings_project.html",
        project=project,
        types=types,
        tiers=tiers,
        maps=maps,
        company_id=company_id,
    )



@app.route("/settings/map-access/<int:map_id>", methods=["GET", "POST"])
@login_required

def settings_map_access(map_id):
    mp = CompanyMap.query.get_or_404(map_id)
    project = mp.project

    # Somente admin ou dono de empresa podem editar o acesso do mapa.
    if (not getattr(current_user, "is_admin", False)) and (
        not getattr(current_user, "is_company_owner", False)
    ):
        abort(403)

    # Lista de possíveis splicers: todos os usuários não-admin
    splicers = (
        User.query.filter_by(is_admin=False)
        .order_by(User.splicer_name, User.username)
        .all()
    )

    if request.method == "POST":
        ids = request.form.getlist("splicer_ids")
        new_users = []
        for uid in ids:
            uid = (uid or "").strip()
            if not uid.isdigit():
                continue
            u = User.query.get(int(uid))
            if u and not u.is_admin:
                new_users.append(u)
        mp.allowed_splicers = new_users
        db.session.commit()
        flash("Splicers com acesso ao mapa atualizados.", "success")
        return redirect(url_for("settings_map_access", map_id=mp.id))

    return render_template(
        "settings_map_access.html",
        project=project,
        mp=mp,
        splicers=splicers,
    )


@app.route("/settings/device/add", methods=["POST"])
@login_required
def settings_device_add():
    name = (request.form.get("name") or "").strip()
    company = (request.form.get("company") or "").strip() or None
    project_id_raw = (request.form.get("project_id") or "").strip()
    project_id = int(project_id_raw) if project_id_raw.isdigit() else None
    next_url = (request.form.get("next") or "").strip() or None
    value_raw = request.form.get("value_usd") or "0"
    try:
        value = float(value_raw or 0)
    except ValueError:
        value = 0.0

    # Valores MEIO/PONTA (opcionais)
    def _opt_float(field):
        raw = (request.form.get(field) or "").strip()
        if raw == "":
            return None
        try:
            return float(raw)
        except ValueError:
            return None

    value_meio = _opt_float("value_meio_usd")
    value_ponta = _opt_float("value_ponta_usd")
    is_ribbon = bool(request.form.get("is_ribbon"))
    ribbon_price = _opt_float("ribbon_price_usd")
    billing_code = (request.form.get("billing_code") or "").strip() or None
    billing_code_meio = (request.form.get("billing_code_meio") or "").strip() or None
    billing_code_ponta = (request.form.get("billing_code_ponta") or "").strip() or None

    if not name:
        flash("Nome do dispositivo é obrigatório.", "danger")
        return redirect(next_url or url_for("settings"))

    dt = DeviceType.query.filter_by(name=name, company=company, project_id=project_id).first()
    if dt:
        dt.value_usd = value
        dt.value_meio_usd = value_meio
        dt.value_ponta_usd = value_ponta
        dt.is_ribbon = is_ribbon
        dt.ribbon_price_usd = ribbon_price if is_ribbon else None
        dt.billing_code = billing_code
        dt.billing_code_meio = billing_code_meio
        dt.billing_code_ponta = billing_code_ponta
    else:
        dt = DeviceType(
            name=name, company=company, project_id=project_id,
            value_usd=value, value_meio_usd=value_meio, value_ponta_usd=value_ponta,
            is_ribbon=is_ribbon,
            ribbon_price_usd=(ribbon_price if is_ribbon else None),
            billing_code=billing_code,
            billing_code_meio=billing_code_meio,
            billing_code_ponta=billing_code_ponta,
        )
        db.session.add(dt)
    db.session.commit()
    flash("Dispositivo salvo.", "success")
    return redirect(next_url or url_for("settings"))


@app.route("/settings/device/<int:did>/delete")
@login_required
def settings_device_delete(did: int):
    next_url = request.args.get("next") or None
    dt = DeviceType.query.get_or_404(did)
    db.session.delete(dt)
    db.session.commit()
    flash("Dispositivo removido.", "success")
    return redirect(next_url or url_for("settings"))


@app.route("/settings/tier/add", methods=["POST"])
@login_required
def settings_tier_add():
    company = (request.form.get("company") or "").strip() or None
    project_id_raw = (request.form.get("project_id") or "").strip()
    project_id = int(project_id_raw) if project_id_raw.isdigit() else None
    next_url = (request.form.get("next") or "").strip() or None
    min_raw = request.form.get("min_splices") or "0"
    max_raw = request.form.get("max_splices") or ""
    price_raw = request.form.get("price") or "0"

    try:
        min_s = int(min_raw or 0)
    except ValueError:
        min_s = 0
    max_s = int(max_raw) if max_raw else None
    try:
        price = float(price_raw or 0)
    except ValueError:
        price = 0.0

    if min_s < 0:
        flash("Splices mín. não pode ser negativo.", "danger")
        return redirect(next_url or url_for("settings"))

    def _strip_code(field):
        v = (request.form.get(field) or "").strip()
        return v if v else None

    tier = SpliceTier(
        company=company,
        project_id=project_id,
        min_splices=min_s,
        max_splices=max_s,
        price_per_splice_usd=price,
        code_splice=_strip_code("code_splice"),
    )
    db.session.add(tier)
    db.session.commit()
    flash("Faixa de fusões salva.", "success")
    return redirect(next_url or url_for("settings"))


@app.route("/settings/tier/<int:tid>/delete")
@login_required
def settings_tier_delete(tid: int):
    next_url = request.args.get("next") or None
    tier = SpliceTier.query.get_or_404(tid)
    db.session.delete(tier)
    db.session.commit()
    flash("Faixa de fusões removida.", "success")
    return redirect(next_url or url_for("settings"))


@app.route("/settings/device/<int:did>/edit", methods=["POST"])
@admin_required
def settings_device_edit(did: int):
    """Edita um DeviceType existente (valores e codigos de cobranca)."""
    dt = DeviceType.query.get_or_404(did)
    next_url = (request.form.get("next") or "").strip() or None

    def _f(field):
        return (request.form.get(field) or "").strip() or None
    def _float(field):
        v = (request.form.get(field) or "").strip()
        try: return float(v) if v else None
        except ValueError: return None

    dt.value_usd        = float((request.form.get("value_usd") or "0")) or 0.0
    dt.value_meio_usd   = _float("value_meio_usd")
    dt.value_ponta_usd  = _float("value_ponta_usd")
    dt.billing_code     = _f("billing_code")
    dt.billing_code_meio  = _f("billing_code_meio")
    dt.billing_code_ponta = _f("billing_code_ponta")
    is_ribbon = bool(request.form.get("is_ribbon"))
    dt.is_ribbon = is_ribbon
    dt.ribbon_price_usd = _float("ribbon_price_usd") if is_ribbon else None
    db.session.commit()

    # Recalcula billing_codes_json de todos os lancamentos que usam este dispositivo
    _recalc_billing_codes_for_device(dt)
    flash(f"Dispositivo '{dt.name}' atualizado.", "success")
    return redirect(next_url or url_for("settings"))


@app.route("/settings/tier/<int:tid>/edit", methods=["POST"])
@admin_required
def settings_tier_edit(tid: int):
    """Edita uma SpliceTier existente (preco e codigo de cobranca)."""
    tier = SpliceTier.query.get_or_404(tid)
    next_url = (request.form.get("next") or "").strip() or None

    try:
        tier.min_splices = int(request.form.get("min_splices") or 0)
    except ValueError:
        pass
    max_raw = (request.form.get("max_splices") or "").strip()
    tier.max_splices = int(max_raw) if max_raw.isdigit() else None
    try:
        tier.price_per_splice_usd = float(request.form.get("price") or 0)
    except ValueError:
        pass
    tier.code_splice = (request.form.get("code_splice") or "").strip() or None
    db.session.commit()

    # Recalcula billing_codes_json de todos os lancamentos afetados por esta faixa
    _recalc_billing_codes_for_tier(tier)
    flash("Faixa atualizada.", "success")
    return redirect(next_url or url_for("settings"))


def _recalc_billing_codes_for_device(dt: DeviceType):
    """Recalcula billing_codes_json para lancamentos que usam este DeviceType."""
    try:
        records = Record.query.filter(
            Record.project_id == dt.project_id,
            Record.company == dt.company,
        ).all() if dt.project_id else Record.query.filter(
            Record.company == dt.company
        ).all()
        changed = 0
        for rec in records:
            rtype = (rec.type or "").lower().strip()
            dname = (dt.name or "").lower().strip()
            if rtype != dname:
                continue
            bcodes = compute_billing_codes(
                int(rec.splices or 0), rec.type or "",
                rec.company, rec.project_id,
                map_role=rec.map_role,
                ribbon_count=getattr(rec, "ribbon_count", None),
            )
            rec.billing_codes_json = json.dumps(bcodes, ensure_ascii=False) if bcodes else None
            changed += 1
        if changed:
            db.session.commit()
    except Exception:
        db.session.rollback()


def _recalc_billing_codes_for_tier(tier: SpliceTier):
    """Recalcula billing_codes_json para lancamentos afetados por esta faixa."""
    try:
        records = Record.query.filter(
            Record.project_id == tier.project_id,
        ).all() if tier.project_id else Record.query.filter(
            Record.company == tier.company
        ).all()
        changed = 0
        for rec in records:
            bcodes = compute_billing_codes(
                int(rec.splices or 0), rec.type or "",
                rec.company, rec.project_id,
                map_role=rec.map_role,
                ribbon_count=getattr(rec, "ribbon_count", None),
            )
            rec.billing_codes_json = json.dumps(bcodes, ensure_ascii=False) if bcodes else None
            changed += 1
        if changed:
            db.session.commit()
    except Exception:
        db.session.rollback()




@app.route("/users", methods=["GET", "POST"])
@admin_required
def manage_users():
    """Cadastro simples de usuários. Apenas admin acessa."""
    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        password = (request.form.get("password") or "").strip()
        splicer_name = (request.form.get("splicer_name") or "").strip() or None
        company_name = (request.form.get("company_name") or "").strip() or None
        is_company_owner = bool(request.form.get("is_company_owner"))
        is_admin = bool(request.form.get("is_admin"))
        is_active = bool(request.form.get("is_active"))
        can_access_expenses = bool(request.form.get("can_access_expenses"))
        can_view_values = bool(request.form.get("can_view_values"))

        if not username or not password:
            flash("Usuário e senha são obrigatórios.", "danger")
            return redirect(url_for("manage_users"))

        user = User.query.filter_by(username=username).first()
        if user:
            user.password = password
            user.splicer_name = splicer_name
            user.company_name = company_name
            user.is_company_owner = is_company_owner
            user.is_admin = is_admin
            user.is_active = is_active
            user.can_access_expenses = can_access_expenses
            user.can_view_values = can_view_values
        else:
            user = User(
                username=username,
                password=password,
                splicer_name=splicer_name,
                company_name=company_name,
                is_company_owner=is_company_owner,
                is_admin=is_admin,
                is_active=is_active,
                can_access_expenses=can_access_expenses,
                can_view_values=can_view_values,
            )
            db.session.add(user)
        db.session.commit()
        flash("Usuário salvo com sucesso.", "success")
        return redirect(url_for("manage_users"))

    companies = CompanyConfig.query.order_by(CompanyConfig.name).all()
    users = User.query.order_by(User.username).all()
    # Mapa user_id -> lista de SplicerPricing via assignments
    assignments = SplicerPricingAssignment.query.all()
    user_pricings = {}
    for a in assignments:
        user_pricings.setdefault(a.user_id, []).append(a.pricing)
    return render_template("users.html", users=users, companies=companies,
                           user_pricings=user_pricings)


@app.route("/users/<int:uid>/pricings")
@admin_required
def user_pricing_summary(uid: int):
    """Retorna as tabelas de preços de um splicer em todos os projetos."""
    user = User.query.get_or_404(uid)
    assignments = SplicerPricingAssignment.query.filter_by(user_id=uid).all()
    result = []
    for a in assignments:
        sp = a.pricing
        result.append({
            "pricing_id": sp.id,
            "project_id": sp.project_id,
            "project_name": sp.project.name if sp.project else "",
            "company": sp.project.company if sp.project else "",
            "label": sp.label,
            "included_splices": sp.included_splices,
            "device_prices": sp.get_device_prices(),
            "tiers": sp.get_tiers(),
        })
    return jsonify({"user": user.splicer_name or user.username, "pricings": result})




@app.route("/users/<int:uid>/delete")
@admin_required
def user_delete(uid: int):
    user = User.query.get_or_404(uid)
    if user.username == "admin":
        flash("Não é permitido remover o usuário admin.", "danger")
        return redirect(url_for("manage_users"))
    if current_user.id == user.id:
        flash("Você não pode remover o próprio usuário logado.", "danger")
        return redirect(url_for("manage_users"))
    db.session.delete(user)
    db.session.commit()
    flash("Usuário removido.", "success")
    return redirect(url_for("manage_users"))


@app.route("/users/<int:uid>/toggle_active", methods=["POST"])
@admin_required
def user_toggle_active(uid: int):
    user = User.query.get_or_404(uid)
    if user.username == "admin":
        flash("Não é possível desativar o usuário admin.", "danger")
        return redirect(url_for("manage_users"))
    user.is_active = not getattr(user, 'is_active', True)
    db.session.commit()
    status = "ativado" if user.is_active else "desativado"
    flash(f"Usuário {user.username} {status}.", "success")
    return redirect(url_for("manage_users"))



@app.route("/export/pdf")
@login_required
def export_pdf():
    """Gera um PDF simples com os registros filtrados (mesma lógica da tela principal)."""

    # Flag opcional: se "no_values=1", não mostra colunas de valores em dinheiro.
    no_values = request.args.get("no_values") == "1"

    # Usa exatamente a mesma lógica de filtros/permissões da tela principal
    query = build_filtered_record_query_from_request()
    records = query.order_by(
        Record.created_date.asc().nullslast(),
        Record.id.asc()
    ).all()

    # totais do período
    total_amount = sum((r.total_usd or 0) for r in records)
    total_splices = sum((r.splices or 0) for r in records)
    total_hubs = sum(1 for r in records if (r.type or "").upper() == "HUB")

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 10, "Relatorio de Producao - SPLICER", ln=1)
    pdf.set_font("Arial", "", 9)

    # linha de totais
    pdf.cell(0, 8, f"Total de splices: {total_splices}", ln=1)
    pdf.cell(0, 8, f"Total de hubs: {total_hubs}", ln=1)
    if not no_values:
        pdf.cell(0, 8, f"Total no período: $ {total_amount:.2f}", ln=1)
    pdf.ln(4)

    # cabeçalho (tabela sem quebrar colunas)
    pdf.set_font("Arial", "B", 8)

    page_w = pdf.w - pdf.l_margin - pdf.r_margin  # largura útil da página

    def _scale_widths(base_widths):
        s = float(sum(base_widths)) or 1.0
        scale = page_w / s
        return [w * scale for w in base_widths]

    def fit_text(text, width):
        """Trunca o texto para caber na célula, sem invadir a coluna seguinte."""
        if text is None:
            text = ""
        text = _pdf_safe(str(text))
        if width <= 0:
            return ""
        if pdf.get_string_width(text) <= width:
            return text
        ell = "..."
        # garante que o ellipsis cabe
        while pdf.get_string_width(ell) > width and len(ell) > 0:
            ell = ell[:-1]
        if not ell:
            return ""
        # vai cortando até caber com "..."
        cut = text
        while cut and pdf.get_string_width(cut + ell) > width:
            cut = cut[:-1]
        return cut + ell

    if no_values:
        headers = ["Data", "Empresa", "Map", "Type", "Dispositivo", "Splices"]
        # base (soma ~190mm) -> escala para caber na largura útil
        col_widths = _scale_widths([20, 35, 35, 15, 65, 20])
    else:
        headers = ["Data", "Empresa", "Map", "Type", "Dispositivo", "Splices", "Fusoes $", "Total $"]
        col_widths = _scale_widths([20, 30, 30, 15, 40, 15, 20, 20])

    # cabeçalho
    for w, h in zip(col_widths, headers):
        pdf.cell(w, 7, h, border=1)
    pdf.ln()

    pdf.set_font("Arial", "", 8)

    for r in records:
        if no_values:
            row = [
                r.created_date.strftime("%Y-%m-%d") if r.created_date else "",
                _pdf_safe(r.company or ""),
                _pdf_safe(r.map or ""),
                _pdf_safe(r.type or ""),
                _pdf_safe(r.device or ""),
                str(r.splices or 0),
            ]
        else:
            row = [
                r.created_date.strftime("%Y-%m-%d") if r.created_date else "",
                _pdf_safe(r.company or ""),
                _pdf_safe(r.map or ""),
                _pdf_safe(r.type or ""),
                _pdf_safe(r.device or ""),
                str(r.splices or 0),
                f"{(r.price_splices_usd or 0):.2f}",
                f"{(r.total_usd or 0):.2f}",
            ]

        for w, val in zip(col_widths, row):
            pdf.cell(w, 6, fit_text(val, w - 1), border=1)
        pdf.ln()

    buf = BytesIO()
    pdf.output(buf)
    buf.seek(0)
    filename = "relatorio_producao.pdf"
    return send_file(buf, as_attachment=True, download_name=filename, mimetype="application/pdf")




@app.route("/invoices")
@admin_required
def invoices_list():
    """Lista simples de todas as invoices para controle contábil."""
    status_filter = request.args.get("status") or None
    query = Invoice.query.order_by(Invoice.created_at.desc())
    if status_filter in ("pending", "paid"):
        query = query.filter(Invoice.status == status_filter)
    invoices = query.all()
    return render_template("invoices.html", invoices=invoices, status_filter=status_filter)

@app.route("/invoice/<int:iid>/pdf")
@login_required
def invoice_pdf(iid: int):
    """Exibe/baixa o PDF salvo de uma invoice já criada."""
    inv = Invoice.query.get_or_404(iid)
    # permissões: admin vê tudo; (futuro) dono da empresa pode ver da própria empresa
    if not getattr(current_user, 'is_admin', False):
        if not (getattr(current_user, 'is_company_owner', False) and (inv.company == (current_user.company_name or ''))):
            abort(403)
    if not inv.pdf_data:
        flash('Esta invoice não tem PDF salvo. Gere novamente.', 'warning')
        return redirect(url_for('invoices_list'))
    download = (request.args.get('download') or '').lower() in ('1','true','yes')
    return send_file(BytesIO(inv.pdf_data), as_attachment=download, download_name=inv.pdf_filename or f"{inv.number}.pdf", mimetype=(inv.pdf_content_type or 'application/pdf'))


@app.route("/invoice/<int:iid>/toggle", methods=["POST"])
@admin_required
def invoice_toggle_status(iid: int):
    inv = Invoice.query.get_or_404(iid)
    inv.status = "paid" if inv.status != "paid" else "pending"
    db.session.commit()
    flash("Invoice status updated.", "success")
    return redirect(url_for("invoices_list"))

@app.route("/invoice/<int:iid>/delete", methods=["POST"])
@admin_required
def invoice_delete(iid: int):
    inv = Invoice.query.get_or_404(iid)
    db.session.delete(inv)
    db.session.commit()
    flash("Invoice deleted.", "success")
    return redirect(url_for("invoices_list"))

def _pdf_safe(text) -> str:
    """Converte texto para ASCII removendo acentos — necessário para fontes Arial/Helvetica no fpdf2."""
    if not text:
        return ""
    import unicodedata as _ud
    nfd = _ud.normalize("NFD", str(text))
    return nfd.encode("ascii", "ignore").decode("ascii")

@app.route("/export/invoice")
@login_required
def export_invoice():
    """Gera uma invoice (nota de cobrança) em PDF para o intervalo de datas e filtros informados.

    A invoice contém: nome do mapa, número do dispositivo, número de fusões,
    valor do dispositivo e total, somados por mapa/dispositivo no período.
    """
    # mesmos filtros do index / export_pdf
    company_filter = request.args.get("company") or None
    splicer_filter = request.args.get("splicer") or None
    map_filter = request.args.get("map") or None
    device_filter = request.args.get("device") or None
    start_raw = request.args.get("start") or None
    end_raw = request.args.get("end") or None
    no_values = False  # sempre com valores na invoice

    # invoice só pode ser gerada para UMA empresa específica
    if not company_filter:
        flash("Para gerar invoice, selecione uma empresa específica (filtro de empresa).", "danger")
        return redirect(url_for("index"))

    query = Record.query

    if company_filter:
        query = query.filter(Record.company == company_filter)
    if splicer_filter and getattr(current_user, "is_admin", False):
        query = query.filter(Record.splicer == splicer_filter)
    if map_filter:
        query = query.filter(Record.map.ilike(f"%{map_filter}%"))
    if device_filter:
        query = query.filter(Record.device.ilike(f"%{device_filter}%"))

    if start_raw:
        try:
            start_dt = datetime.fromisoformat(start_raw)
            query = query.filter(Record.created_date >= start_dt)
        except ValueError:
            start_dt = None
    else:
        start_dt = None

    if end_raw:
        try:
            end_dt = datetime.fromisoformat(end_raw)
            query = query.filter(Record.created_date <= end_dt)
        except ValueError:
            end_dt = None
    else:
        end_dt = None

    from datetime import datetime as _dt
    inv_date = _dt.utcnow().date().isoformat()
    inv_number = _dt.utcnow().strftime("INV-%Y%m%d-%H%M%S")

    # se não for admin, força o filtro para o próprio splicer
    if not getattr(current_user, "is_admin", False):
        enforced_splicer = getattr(current_user, "splicer_name", None) or current_user.username
        query = query.filter(Record.splicer == enforced_splicer)

    records = query.order_by(Record.created_date.asc().nullslast(), Record.id.asc()).all()

    # Busca lançamentos de horas com os mesmos filtros
    hq = HourRecord.query.filter(HourRecord.company == company_filter)
    if splicer_filter and getattr(current_user, "is_admin", False):
        hq = hq.filter(HourRecord.splicer == splicer_filter)
    if start_raw:
        try:
            hq = hq.filter(HourRecord.created_date >= datetime.fromisoformat(start_raw))
        except ValueError:
            pass
    if end_raw:
        try:
            hq = hq.filter(HourRecord.created_date <= datetime.fromisoformat(end_raw))
        except ValueError:
            pass
    if not getattr(current_user, "is_admin", False):
        enforced_splicer = getattr(current_user, "splicer_name", None) or current_user.username
        hq = hq.filter(HourRecord.splicer == enforced_splicer)
    hour_records = hq.order_by(HourRecord.created_date.asc().nullslast()).all()

    # Sem agrupamento: cada lancamento = uma linha na invoice
    grouped = {}
    for r in records:
        key = r.id  # chave unica por lancamento
        # Cada record tem sua propria entrada (chave = r.id)
        dt_src = r.created_date or getattr(r, "created_at", None)
        d_iso = dt_src.date().isoformat() if dt_src else "-"
        grouped[key] = {
            "map": (r.map or "-").strip(),
            "device": (r.device or "-").strip(),
            "launch_date": d_iso,
            "map_role": (r.map_role or "").strip(),
            "role": (r.map_role or "").strip(),
            "included": int(r.included_splices_applied) if r.included_splices_applied is not None else None,
            "splices": int(r.splices or 0),
            "price_device_usd": float(getattr(r, "price_device_usd", 0.0) or 0.0),
            "total_usd": float(getattr(r, "total_usd", 0.0) or 0.0),
        }

        # Coleta codigos de cobranca do registro
        # Se nao estiver salvo, calcula na hora (registros antigos)
        bcodes = []
        if r.billing_codes_json:
            try:
                bcodes = json.loads(r.billing_codes_json)
            except Exception:
                pass
        if not bcodes:
            # Calcula na hora para registros antigos sem billing_codes_json
            device_for_code = (r.type or r.device or "")
            bcodes = compute_billing_codes(
                int(r.splices or 0),
                device_for_code,
                r.company,
                r.project_id,
                map_role=r.map_role,
                ribbon_count=getattr(r, "ribbon_count", None),
            )
        grouped[key]["billing_codes"] = [c for c in bcodes if c]

    # lista final de linhas da invoice (um item por grupo mapa/device/tipo)
    # Recalcula total de cada linha com o device_price corrigido
    for l in grouped.values():
        l["total_usd"] = float(l.get("total_usd") or 0.0)
        stored_device = float(l.get("price_device_usd") or 0.0)
        if stored_device > 0 and l["total_usd"] == 0.0:
            l["total_usd"] = stored_device

    lines = [
        l for l in grouped.values()
        if (float(l.get("total_usd") or 0.0) > 0.0) or (int(l.get("splices") or 0) > 0) or (float(l.get("price_device_usd") or 0.0) > 0.0)
    ]

    # Adiciona lançamentos de horas como linhas separadas
    for hr in hour_records:
        if (hr.total_usd or 0) > 0:
            lines.append({
                "launch_date": hr.created_date.date().isoformat() if hr.created_date else "-",
                "map": hr.map_name or "-",
                "device": hr.description or "Hora trabalhada",
                "role": "HORAS",
                "included": None,
                "splices": 0,
                "price_device_usd": 0.0,
                "total_usd": float(hr.total_usd or 0),
                "billing_codes": [hr.billing_code] if hr.billing_code else [],
                "_is_hour": True,
                "_hours": hr.hours,
                "_rate": hr.rate_usd,
            })

    # Ordena todas as linhas por data
    lines.sort(key=lambda l: l.get("launch_date") or "")

    # total geral da invoice (soma de todos os grupos)
    total_invoice = sum(l["total_usd"] for l in lines)

    # persist invoice for accounting (PDF será salvo no banco para visualizar depois)
    inv_start_date = start_dt.date() if start_dt else None
    inv_end_date = end_dt.date() if end_dt else None
    inv_rec = Invoice(
        number=inv_number,
        company=company_filter or "",
        start_date=inv_start_date,
        end_date=inv_end_date,
        total_usd=float(total_invoice or 0.0),
        created_by=getattr(current_user, 'id', None),
    )
    db.session.add(inv_rec)
    # montar PDF da invoice
    # buscar dados da sua empresa (emitente)
    syscfg = SystemConfig.query.first()

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # header - your company (FROM)
    pdf.set_font("Arial", "B", 12)
    if syscfg and syscfg.my_company_name:
        pdf.cell(0, 6, _pdf_safe(syscfg.my_company_name), ln=1)
    if syscfg and syscfg.my_company_address:
        for line in (syscfg.my_company_address or "").splitlines():
            if line.strip():
                pdf.set_font("Arial", "", 9)
                pdf.cell(0, 5, _pdf_safe(line.strip()), ln=1)
    if syscfg and (syscfg.my_company_email or syscfg.my_company_phone):
        contact_parts = []
        if syscfg.my_company_email:
            contact_parts.append(syscfg.my_company_email)
        if syscfg.my_company_phone:
            contact_parts.append(syscfg.my_company_phone)
        pdf.cell(0, 5, _pdf_safe(" | ".join(contact_parts)), ln=1)
    pdf.ln(4)

    # invoice title and metadata
    pdf.set_font("Arial", "B", 14)
    pdf.cell(0, 8, "INVOICE", ln=1)

    pdf.set_font("Arial", "", 10)
    pdf.cell(0, 6, f"Invoice date: {inv_date}", ln=1)
    pdf.cell(0, 6, f"Invoice #: {inv_number}", ln=1)
    pdf.ln(4)

    # BILL TO (client)
    cfg_cli = CompanyConfig.query.filter_by(name=company_filter).first()
    pdf.set_font("Arial", "B", 10)
    pdf.cell(0, 6, "BILL TO:", ln=1)
    pdf.set_font("Arial", "", 9)
    if cfg_cli:
        if cfg_cli.invoice_address:
            for line in (cfg_cli.invoice_address or "").splitlines():
                if line.strip():
                    pdf.cell(0, 5, _pdf_safe(line.strip()), ln=1)
        else:
            pdf.cell(0, 5, _pdf_safe(cfg_cli.name), ln=1)
    else:
        pdf.cell(0, 5, _pdf_safe(company_filter or ""), ln=1)

    pdf.ln(4)

    # table header
    # OBS: FPDF "cell" não faz quebra de linha. Para não "estourar" a tabela e
    # mostrar o nome COMPLETO do device dentro da coluna, vamos desenhar a linha
    # com multi_cell e quebra manual (principalmente em '_' e '-').
    
    # === Tabela das linhas (uma linha por registro) ===
    # Coluna de data primeiro, como solicitado
    # Verifica se algum registro tem codigos de cobranca
    # A4 usable width = 190mm. Colunas somam exatamente 190.
    # Sem codigos: Date(22) Map(30) Device(50) Tipo(14) Incl(10) Splices(14) Dev$(24) Total(26) = 190
    # Com codigos: Date(20) Map(28) Device(42) Tipo(13) Incl(10) Splices(13) Dev$(22) Total(22) Codes(20) = 190
    has_billing = any(l.get("billing_codes") for l in lines)
    if has_billing:
        # Date(20)+Codes(28)+Map(18)+Device(32)+Tipo(14)+Incl(9)+Splices(12)+Dev$(24)+Total(33)=190
        col_widths = [20, 28, 18, 32, 14, 9, 12, 24, 33]
        headers   = ["Date", "Codes", "Map", "Device", "Tipo", "Incl.", "Spl.", "Dev $", "Total"]
    else:
        col_widths = [22, 30, 50, 14, 10, 14, 24, 26]
        headers   = ["Date", "Map", "Device", "Tipo", "Incl.", "Splices", "Dev $", "Total"]

    pdf.set_font("Arial", "B", 9)

    def _draw_table_header():
        for w, h in zip(col_widths, headers):
            pdf.cell(w, 7, _pdf_safe(h), border=1, align="C")
        pdf.ln(7)

    _draw_table_header()

    line_h = 6
    pdf.set_font("Arial", "", 8)

    for line in lines:
        row = [(line.get("launch_date") or "-")]
        if has_billing:
            row.append(_pdf_safe(", ".join(line.get("billing_codes") or [])))
        if line.get("_is_hour"):
            row += [
                _pdf_safe(line["map"] or "-"),
                _pdf_safe(line["device"] or "-"),
                "HORAS",
                "-",
                f"{line['_hours']}h",
                f"${line['_rate']:.2f}/h",
                f"${line['total_usd']:.2f}",
            ]
        else:
            row += [
                _pdf_safe(line["map"] or "-"),
                _pdf_safe(line["device"] or "-"),
                _pdf_safe(line["role"] or "-"),
                str(line["included"] if line["included"] is not None else "-"),
                str(line["splices"]),
                f"${line['price_device_usd']:.2f}",
                f"${line['total_usd']:.2f}",
            ]

        # quebra de página manual antes de desenhar a linha
        if pdf.get_y() + line_h > pdf.page_break_trigger:
            pdf.add_page()
            _draw_table_header()
            pdf.set_font("Arial", "", 9)

        for w, val in zip(col_widths, row):
            pdf.cell(w, line_h, _pdf_safe(str(val)), border=1)
        pdf.ln(line_h)

    pdf.ln(4)
    pdf.set_font("Arial", "B", 11)
    pdf.cell(0, 8, f"Invoice total: $ {total_invoice:.2f}", ln=1)

    # gerar bytes do PDF e salvar na Invoice para poder visualizar depois
    try:
        pdf_bytes = pdf.output(dest='S').encode('latin-1')
    except Exception:
        buf = BytesIO()
        pdf.output(buf)
        pdf_bytes = buf.getvalue()

    inv_rec.pdf_filename = f"{inv_number}.pdf"
    inv_rec.pdf_content_type = "application/pdf"
    inv_rec.pdf_data = pdf_bytes
    db.session.commit()

    download = (request.args.get('download') or '').lower() in ('1', 'true', 'yes')
    return send_file(
        BytesIO(pdf_bytes),
        as_attachment=download,
        download_name=inv_rec.pdf_filename or "invoice.pdf",
        mimetype="application/pdf",
    )


# ═══════════════════════════════════════════════════════
# PAYROLL — Folhas de pagamento
# ═══════════════════════════════════════════════════════
# ── TABELAS DE PREÇOS DE SPLICERS POR PROJETO ───────────
# ═══════════════════════════════════════════════════════

def _get_splicer_pricing(user_id: int, project_id: int):
    """Retorna a tabela de preços de um splicer num projeto via assignment, ou None."""
    if not user_id or not project_id:
        return None
    assignment = SplicerPricingAssignment.query.filter_by(
        user_id=user_id, project_id=project_id
    ).first()
    return assignment.pricing if assignment else None


def _calc_splicer_cost(records, pricing):
    """Calcula custo total do splicer para uma lista de records usando sua tabela."""
    if not pricing:
        return None
    total = 0.0
    for r in records:
        total += pricing.total_for_record(int(r.splices or 0), r.type or r.device or "")
    return total


@app.route("/settings/project/<int:pid>/splicer-pricing", methods=["GET", "POST"])
@admin_required
def project_splicer_pricing(pid: int):
    """Gerencia tabelas de preços de splicers num projeto (many-to-many)."""
    project = Project.query.get_or_404(pid)
    comp_cfg = CompanyConfig.query.filter_by(name=project.company).first()
    company_id = comp_cfg.id if comp_cfg else 0

    if request.method == "POST":
        action = (request.form.get("action") or "").strip()

        # ── Criar ou editar tabela de valores ──
        if action in ("add", "edit"):
            pricing_id = request.form.get("pricing_id")
            label = (request.form.get("label") or "").strip()
            included = int(request.form.get("included_splices") or 0)

            # Preços por dispositivo
            device_prices = {}
            i = 0
            while True:
                dev_name = request.form.get(f"dev_name_{i}")
                if dev_name is None:
                    break
                dev_name = dev_name.strip()
                try:
                    dev_price = float(request.form.get(f"dev_price_{i}") or 0)
                    if dev_name:
                        device_prices[dev_name] = dev_price
                except (ValueError, TypeError):
                    pass
                i += 1

            # Tiers de fusão
            tiers = []
            i = 0
            while True:
                min_raw = request.form.get(f"tier_min_{i}")
                if min_raw is None:
                    break
                try:
                    t_min = int(min_raw)
                    t_max_raw = (request.form.get(f"tier_max_{i}") or "").strip()
                    t_max = int(t_max_raw) if t_max_raw else None
                    t_price = float(request.form.get(f"tier_price_{i}") or 0)
                    tiers.append({"min": t_min, "max": t_max, "price": t_price})
                except (ValueError, TypeError):
                    pass
                i += 1

            if action == "edit" and pricing_id:
                sp = SplicerPricing.query.get(int(pricing_id))
                if sp and sp.project_id == pid:
                    sp.label = label or sp.label
                    sp.included_splices = included
                    sp.device_prices_json = json.dumps(device_prices)
                    sp.tiers_json = json.dumps(tiers)
                    db.session.commit()
                    flash(f"Tabela '{sp.label}' atualizada.", "success")
            else:
                sp = SplicerPricing(
                    project_id=pid,
                    label=label or "Nova tabela",
                    included_splices=included,
                    device_prices_json=json.dumps(device_prices),
                    tiers_json=json.dumps(tiers),
                )
                db.session.add(sp)
                db.session.commit()
                flash(f"Tabela '{sp.label}' criada.", "success")
            return redirect(url_for("project_splicer_pricing", pid=pid))

        # ── Remover tabela (cascata remove assignments) ──
        if action == "delete":
            pricing_id = int(request.form.get("pricing_id") or 0)
            sp = SplicerPricing.query.get(pricing_id)
            if sp and sp.project_id == pid:
                db.session.delete(sp)
                db.session.commit()
                flash("Tabela removida.", "success")
            return redirect(url_for("project_splicer_pricing", pid=pid))

        # ── Atribuir splicer a uma tabela ──
        if action == "assign":
            pricing_id = int(request.form.get("pricing_id") or 0)
            user_ids = request.form.getlist("user_ids")  # múltiplos splicers
            sp = SplicerPricing.query.get(pricing_id)
            if not sp or sp.project_id != pid:
                flash("Tabela inválida.", "danger")
                return redirect(url_for("project_splicer_pricing", pid=pid))
            added = 0
            for uid_raw in user_ids:
                try:
                    uid = int(uid_raw)
                except (ValueError, TypeError):
                    continue
                # Remove de outra tabela no mesmo projeto se existir
                old = SplicerPricingAssignment.query.filter_by(user_id=uid, project_id=pid).first()
                if old:
                    if old.pricing_id == pricing_id:
                        continue  # já está nesta tabela
                    db.session.delete(old)
                db.session.add(SplicerPricingAssignment(
                    pricing_id=pricing_id, user_id=uid, project_id=pid
                ))
                added += 1
            db.session.commit()
            flash(f"{added} splicer(s) atribuído(s) à tabela '{sp.label}'.", "success")
            return redirect(url_for("project_splicer_pricing", pid=pid))

        # ── Remover splicer de uma tabela ──
        if action == "unassign":
            assignment_id = int(request.form.get("assignment_id") or 0)
            a = SplicerPricingAssignment.query.get(assignment_id)
            if a and a.project_id == pid:
                db.session.delete(a)
                db.session.commit()
                flash("Splicer removido da tabela.", "success")
            return redirect(url_for("project_splicer_pricing", pid=pid))

    pricings = SplicerPricing.query.filter_by(project_id=pid).order_by(SplicerPricing.label).all()
    # Splicers já atribuídos a alguma tabela neste projeto
    all_assignments = SplicerPricingAssignment.query.filter_by(project_id=pid).all()
    assigned_user_ids = {a.user_id for a in all_assignments}
    all_splicers = User.query.filter_by(is_admin=False, is_company_owner=False, is_active=True).order_by(User.username).all()
    # Dispositivos do projeto (mesmo que o cliente vê)
    project_devices = DeviceType.query.filter_by(project_id=pid).order_by(DeviceType.name).all()
    # Fallback: dispositivos da empresa se projeto não tiver nenhum
    if not project_devices:
        project_devices = DeviceType.query.filter_by(company=project.company, project_id=None).order_by(DeviceType.name).all()

    return render_template(
        "project_splicer_pricing.html",
        project=project,
        company_id=company_id,
        pricings=pricings,
        all_splicers=all_splicers,
        assigned_user_ids=assigned_user_ids,
        all_assignments=all_assignments,
        project_devices=project_devices,
    )


@app.route("/api/project/<int:pid>/splicer-margin")
@admin_required
def api_splicer_margin(pid: int):
    """Retorna relatório de margem por splicer para um projeto no período informado."""
    project = Project.query.get_or_404(pid)
    start_raw = request.args.get("start") or None
    end_raw = request.args.get("end") or None

    start_dt = datetime.fromisoformat(start_raw) if start_raw else None
    end_dt = datetime.fromisoformat(end_raw + "T23:59:59") if end_raw else None

    pricings = SplicerPricing.query.filter_by(project_id=pid).all()

    results = []
    seen_splicers = set()

    for pricing in pricings:
        for assignment in pricing.assignments:
            splicer_user = assignment.user
            if not splicer_user:
                continue
            splicer_name = splicer_user.splicer_name or splicer_user.username
            if splicer_name in seen_splicers:
                continue

        q = Record.query.filter(Record.project_id == pid, Record.splicer == splicer_name)
        if start_dt:
            q = q.filter(Record.created_date >= start_dt)
        if end_dt:
            q = q.filter(Record.created_date <= end_dt)

        records = q.all()
        total_splices = sum(r.splices or 0 for r in records)
        revenue = sum(r.total_usd or 0.0 for r in records)
        cost = _calc_splicer_cost(records, pricing) or 0.0
        margin = revenue - cost

        results.append({
            "splicer": splicer_name,
            "total_records": len(records),
            "total_splices": total_splices,
            "revenue_usd": round(revenue, 2),
            "cost_usd": round(cost, 2),
            "margin_usd": round(margin, 2),
            "margin_pct": round((margin / revenue * 100) if revenue > 0 else 0, 1),
            "included_splices": pricing.included_splices,
            "device_prices": pricing.get_device_prices(),
            "tiers": pricing.get_tiers(),
        })
        seen_splicers.add(splicer_name)

    # Splicers com produção mas sem tabela
    q_all = Record.query.filter(Record.project_id == pid)
    if start_dt:
        q_all = q_all.filter(Record.created_date >= start_dt)
    if end_dt:
        q_all = q_all.filter(Record.created_date <= end_dt)
    for r in q_all.all():
        name = r.splicer or ""
        if name and name not in seen_splicers:
            results.append({
                "splicer": name,
                "total_records": 1,
                "total_splices": int(r.splices or 0),
                "revenue_usd": round(float(r.total_usd or 0), 2),
                "cost_usd": None,
                "margin_usd": None,
                "margin_pct": None,
                "included_splices": None,
                "device_value": None,
                "tiers": [],
            })
            seen_splicers.add(name)

    results.sort(key=lambda x: x["splicer"])
    return jsonify({"project": project.name, "results": results})


@app.route("/project/<int:pid>/margin-report")
@admin_required
def project_margin_report(pid: int):
    """Relatório de margem por splicer — página completa."""
    project = Project.query.get_or_404(pid)
    comp_cfg = CompanyConfig.query.filter_by(name=project.company).first()
    company_id = comp_cfg.id if comp_cfg else 0
    return render_template("project_margin_report.html", project=project, company_id=company_id)
# ═══════════════════════════════════════════════════════

@app.route("/payroll")

@admin_required
def payroll_list():
    """Lista todas as folhas de pagamento."""
    status_filter = request.args.get("status") or None
    user_filter_raw = request.args.get("user_id") or None
    user_filter = int(user_filter_raw) if user_filter_raw and user_filter_raw.isdigit() else None

    q = Payroll.query
    if status_filter:
        q = q.filter(Payroll.status == status_filter)
    if user_filter:
        q = q.filter(Payroll.user_id == user_filter)

    payrolls = q.order_by(Payroll.created_at.desc()).all()
    splicers = User.query.filter_by(is_admin=False, is_company_owner=False).order_by(User.username).all()
    return render_template("payroll.html",
                           payrolls=payrolls,
                           splicers=splicers,
                           status_filter=status_filter or "",
                           user_filter=user_filter or "")


@app.route("/payroll/new", methods=["GET", "POST"])
@admin_required
def payroll_new():
    """Cria uma nova folha de pagamento para um splicer."""
    splicers = User.query.filter_by(is_admin=False, is_company_owner=False).order_by(User.username).all()
    projects = Project.query.order_by(Project.company, Project.name).all()

    if request.method == "POST":
        user_id_raw = (request.form.get("user_id") or "").strip()
        start_raw = (request.form.get("start_date") or "").strip()
        end_raw = (request.form.get("end_date") or "").strip()
        project_id_raw = (request.form.get("project_id") or "").strip()
        notes = (request.form.get("notes") or "").strip() or None

        if not user_id_raw or not start_raw or not end_raw:
            flash("Splicer, data início e data fim são obrigatórios.", "danger")
            return render_template("payroll_new.html", splicers=splicers, projects=projects)

        try:
            splicer_user = User.query.get(int(user_id_raw))
            start_date = date.fromisoformat(start_raw)
            end_date = date.fromisoformat(end_raw)
        except Exception:
            flash("Dados inválidos.", "danger")
            return render_template("payroll_new.html", splicers=splicers, projects=projects)

        if end_date < start_date:
            flash("Data fim deve ser maior que data início.", "danger")
            return render_template("payroll_new.html", splicers=splicers, projects=projects)

        project_id = int(project_id_raw) if project_id_raw.isdigit() else None
        project = Project.query.get(project_id) if project_id else None
        company = project.company if project else None

        # Calcular totais dos lançamentos do splicer no período
        splicer_name = splicer_user.splicer_name or splicer_user.username
        q = Record.query.filter(
            Record.splicer == splicer_name,
            Record.created_date >= datetime.combine(start_date, datetime.min.time()),
            Record.created_date <= datetime.combine(end_date, datetime.max.time()),
        )
        if project_id:
            q = q.filter(Record.project_id == project_id)
        elif company:
            q = q.filter(Record.company == company)

        records = q.all()
        total_records = len(records)
        total_splices = sum(r.splices or 0 for r in records)
        total_amount = sum(r.total_usd or 0.0 for r in records)

        # Custo do splicer (modalidade de pagamento)
        splicer_plan = _get_splicer_pricing(splicer_user.id, project_id) if project_id else None
        splicer_cost = _calc_splicer_cost(records, splicer_plan)

        # Prazo de pagamento
        payment_days = getattr(project, "payment_days", None) or 30
        due_date = end_date + __import__("datetime").timedelta(days=payment_days)

        payroll = Payroll(
            user_id=splicer_user.id,
            start_date=start_date,
            end_date=end_date,
            project_id=project_id,
            company=company,
            total_records=total_records,
            total_splices=total_splices,
            total_amount_usd=total_amount,
            payment_days=payment_days,
            due_date=due_date,
            status="pending",
            created_by=current_user.id,
            notes=notes,
            splicer_cost_usd=splicer_cost,
            pricing_id=splicer_plan.id if splicer_plan else None,
            plan_name=splicer_plan.label if splicer_plan else None,
        )
        db.session.add(payroll)
        db.session.commit()
        flash(f"Payroll criado para {splicer_name}: {total_records} lançamentos, $ {total_amount:.2f}.", "success")
        return redirect(url_for("payroll_list"))

    return render_template("payroll_new.html", splicers=splicers, projects=projects)


@app.route("/payroll/<int:pid>/pay", methods=["POST"])
@admin_required
def payroll_mark_paid(pid: int):
    """Marca uma folha como paga."""
    p = Payroll.query.get_or_404(pid)
    if p.status == "paid":
        flash("Este payroll já foi marcado como pago.", "warning")
    else:
        p.status = "paid"
        p.paid_at = datetime.utcnow()
        p.paid_by = current_user.id
        db.session.commit()
        flash("Payroll marcado como pago.", "success")
    return redirect(url_for("payroll_list"))


@app.route("/payroll/<int:pid>/cancel", methods=["POST"])
@admin_required
def payroll_cancel(pid: int):
    """Cancela uma folha de pagamento."""
    p = Payroll.query.get_or_404(pid)
    if p.status == "paid":
        flash("Não é possível cancelar um payroll já pago.", "danger")
    else:
        p.status = "cancelled"
        db.session.commit()
        flash("Payroll cancelado.", "warning")
    return redirect(url_for("payroll_list"))


@app.route("/payroll/<int:pid>/delete", methods=["POST"])
@admin_required
def payroll_delete(pid: int):
    """Remove uma folha de pagamento."""
    p = Payroll.query.get_or_404(pid)
    db.session.delete(p)
    db.session.commit()
    flash("Payroll removido.", "success")
    return redirect(url_for("payroll_list"))



@app.route("/payroll/<int:pid>/pdf")
@admin_required
def payroll_pdf(pid: int):
    """Gera PDF profissional do payroll para envio ao splicer."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    import io

    p = Payroll.query.get_or_404(pid)
    syscfg = SystemConfig.query.first()
    splicer_name = p.user.splicer_name or p.user.username

    q = Record.query.filter(
        Record.splicer == splicer_name,
        Record.created_date >= datetime.combine(p.start_date, datetime.min.time()),
        Record.created_date <= datetime.combine(p.end_date, datetime.max.time()),
    )
    if p.project_id:
        q = q.filter(Record.project_id == p.project_id)
    elif p.company:
        q = q.filter(Record.company == p.company)
    records = q.order_by(Record.created_date).all()

    C_DARK   = rl_colors.HexColor("#0f0f1a")
    C_PURPLE = rl_colors.HexColor("#7c3aed")
    C_GREY   = rl_colors.HexColor("#6b7280")
    C_LIGHT  = rl_colors.HexColor("#f8f8fc")
    C_WHITE  = rl_colors.white
    C_GREEN  = rl_colors.HexColor("#059669")
    C_ORANGE = rl_colors.HexColor("#d97706")
    C_ROW_ALT= rl_colors.HexColor("#f3f4f6")
    C_BORDER = rl_colors.HexColor("#e5e7eb")
    BASE = getSampleStyleSheet()["Normal"]

    def PS(name, **kw):
        d = dict(fontName="Helvetica",fontSize=10,textColor=C_DARK,leading=14,parent=BASE)
        d.update(kw)
        return ParagraphStyle(name, **d)

    S_TH  = PS("th2", fontName="Helvetica-Bold",fontSize=8, textColor=C_WHITE, leading=10,alignment=TA_CENTER)
    S_TD  = PS("td2", fontName="Helvetica",     fontSize=8, textColor=C_DARK,  leading=10)
    S_TDR = PS("tdr2",fontName="Helvetica",     fontSize=8, textColor=C_DARK,  leading=10,alignment=TA_RIGHT)
    S_TDC = PS("tdc2",fontName="Helvetica",     fontSize=8, textColor=C_DARK,  leading=10,alignment=TA_CENTER)
    S_TL  = PS("tl2", fontName="Helvetica-Bold",fontSize=9, textColor=C_WHITE, leading=12)
    S_TR  = PS("tr2", fontName="Helvetica-Bold",fontSize=9, textColor=C_WHITE, leading=12,alignment=TA_RIGHT)
    S_NOTE= PS("nt2", fontName="Helvetica",     fontSize=8, textColor=C_GREY,  leading=11)
    S_LBL = PS("lb2", fontName="Helvetica-Bold",fontSize=8, textColor=C_GREY,  leading=11,spaceAfter=1)
    S_VAL = PS("vl2", fontName="Helvetica",     fontSize=11,textColor=C_DARK,  leading=15)
    S_VLB = PS("vb2", fontName="Helvetica-Bold",fontSize=13,textColor=C_DARK,  leading=18)

    company_name  = (syscfg.my_company_name  if syscfg else None) or "SPLICER"
    company_addr  = (syscfg.my_company_address if syscfg else None) or ""
    company_email = (syscfg.my_company_email if syscfg else None) or ""
    company_phone = (syscfg.my_company_phone if syscfg else None) or ""
    sub_parts = [x for x in [company_addr, company_email, company_phone] if x]

    status_map   = {"pending":"PENDENTE","paid":"PAGO","cancelled":"CANCELADO"}
    status_text  = status_map.get(p.status, p.status.upper())
    status_color = C_GREEN if p.status=="paid" else (C_ORANGE if p.status=="pending" else C_GREY)

    period_str  = f"{p.start_date.strftime('%d/%m/%Y')} → {p.end_date.strftime('%d/%m/%Y')}"
    due_str     = p.due_date.strftime('%d/%m/%Y') if p.due_date else "—"
    project_str = p.project.name if p.project else (p.company or "Todos os projetos")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
        leftMargin=20*mm, rightMargin=20*mm, topMargin=20*mm, bottomMargin=20*mm,
        title=f"Payroll #{p.id:04d} — {splicer_name}")
    W = A4[0] - 40*mm
    story = []

    # Header
    hdr = Table([[
        Paragraph(company_name, PS("cnx",fontName="Helvetica-Bold",fontSize=16,textColor=C_DARK,leading=20)),
        Paragraph(f"PAYROLL  #{p.id:04d}", PS("pidx",fontName="Helvetica-Bold",fontSize=20,textColor=C_PURPLE,leading=24,alignment=TA_RIGHT)),
    ]],colWidths=[W*0.55,W*0.45])
    hdr.setStyle(TableStyle([("VALIGN",(0,0),(-1,-1),"MIDDLE"),("BOTTOMPADDING",(0,0),(-1,-1),4)]))
    story.append(hdr)
    if sub_parts:
        story.append(Paragraph(" · ".join(sub_parts), PS("addrx",fontName="Helvetica",fontSize=9,textColor=C_GREY,leading=13)))
    story.append(Spacer(1,4*mm))
    story.append(HRFlowable(width="100%",thickness=2,color=C_PURPLE,spaceAfter=5*mm))

    # Info cards
    cw = W/3
    def cell(lbl,val,vs=None):
        return Table([[Paragraph(lbl,S_LBL)],[Paragraph(val,vs or S_VAL)]],colWidths=[cw-4*mm])
    def cards(row):
        t = Table([row],colWidths=[cw]*3)
        t.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,-1),C_LIGHT),
            ("LEFTPADDING",(0,0),(-1,-1),5*mm),("RIGHTPADDING",(0,0),(-1,-1),3*mm),
            ("TOPPADDING",(0,0),(-1,-1),3*mm),("BOTTOMPADDING",(0,0),(-1,-1),3*mm),
            ("VALIGN",(0,0),(-1,-1),"TOP"),
            ("LINEAFTER",(0,0),(1,0),0.5,C_BORDER),("BOX",(0,0),(-1,-1),0.5,C_BORDER),
        ]))
        return t
    story.append(cards([cell("SPLICER",splicer_name,S_VLB),cell("PERÍODO",period_str),cell("PROJETO / EMPRESA",project_str)]))
    story.append(Spacer(1,2*mm))
    story.append(cards([
        cell("VENCIMENTO",due_str),
        cell("PRAZO DE PAGAMENTO",f"{p.payment_days} dias"),
        cell("STATUS",status_text,PS("stx2",fontName="Helvetica-Bold",fontSize=11,textColor=status_color,leading=15)),
    ]))
    story.append(Spacer(1,6*mm))

    # Totals
    tw = W/3
    tot_tbl = Table([
        [Paragraph(str(p.total_records),PS("v1x",fontName="Helvetica-Bold",fontSize=26,textColor=C_PURPLE,leading=30,alignment=TA_CENTER)),
         Paragraph(str(p.total_splices), PS("v2x",fontName="Helvetica-Bold",fontSize=26,textColor=C_PURPLE,leading=30,alignment=TA_CENTER)),
         Paragraph(f"$ {p.total_amount_usd:,.2f}",PS("v3x",fontName="Helvetica-Bold",fontSize=20,textColor=C_GREEN,leading=24,alignment=TA_CENTER))],
        [Paragraph("LANÇAMENTOS",PS("l1x",fontName="Helvetica",fontSize=8,textColor=rl_colors.HexColor("#9ca3af"),leading=10,alignment=TA_CENTER)),
         Paragraph("FUSÕES",     PS("l2x",fontName="Helvetica",fontSize=8,textColor=rl_colors.HexColor("#9ca3af"),leading=10,alignment=TA_CENTER)),
         Paragraph("VALOR TOTAL USD",PS("l3x",fontName="Helvetica",fontSize=8,textColor=rl_colors.HexColor("#9ca3af"),leading=10,alignment=TA_CENTER))],
    ],colWidths=[tw]*3)
    tot_tbl.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,-1),C_DARK),
        ("TOPPADDING",(0,0),(-1,0),5*mm),("BOTTOMPADDING",(0,0),(-1,0),2*mm),
        ("TOPPADDING",(0,1),(-1,1),1*mm),("BOTTOMPADDING",(0,1),(-1,1),5*mm),
        ("LEFTPADDING",(0,0),(-1,-1),4*mm),("RIGHTPADDING",(0,0),(-1,-1),4*mm),
        ("LINEAFTER",(0,0),(1,-1),0.5,rl_colors.HexColor("#2d2d4e")),
    ]))
    story += [tot_tbl, Spacer(1,6*mm)]
    story.append(Paragraph("DETALHAMENTO DOS LANÇAMENTOS",PS("secx",fontName="Helvetica-Bold",fontSize=9,textColor=C_PURPLE,leading=12,spaceAfter=3*mm)))

    CW = [22*mm,32*mm,34*mm,22*mm,17*mm,18*mm,18*mm,20*mm]
    rows = [[Paragraph(h,S_TH) for h in ["DATA","MAPA","DEVICE","TIPO","FUSÕES","$ FUSÕES","$ DEVICE","$ TOTAL"]]]
    for r in records:
        rows.append([
            Paragraph(r.created_date.strftime("%d/%m/%Y") if r.created_date else "—",S_TDC),
            Paragraph((r.map or "—")[:25],S_TD),
            Paragraph((r.device or "—")[:28],S_TD),
            Paragraph((r.type or "—")[:16],S_TDC),
            Paragraph(str(r.splices or 0),S_TDC),
            Paragraph(f"${r.price_splices_usd or 0:.2f}",S_TDR),
            Paragraph(f"${r.price_device_usd or 0:.2f}",S_TDR),
            Paragraph(f"${r.total_usd or 0:.2f}",S_TDR),
        ])
    nr = len(rows)
    rows.append([
        Paragraph("TOTAL GERAL",S_TL),Paragraph("",S_NOTE),Paragraph("",S_NOTE),Paragraph("",S_NOTE),
        Paragraph(str(sum(r.splices or 0 for r in records)),S_TR),
        Paragraph(f"${sum(r.price_splices_usd or 0 for r in records):.2f}",S_TR),
        Paragraph(f"${sum(r.price_device_usd or 0 for r in records):.2f}",S_TR),
        Paragraph(f"${sum(r.total_usd or 0 for r in records):.2f}",S_TR),
    ])
    if not records:
        rows.append([Paragraph("Nenhum lançamento encontrado.",S_NOTE)]+[Paragraph("",S_NOTE)]*7)
    rt = Table(rows,colWidths=CW,repeatRows=1)
    ts2 = TableStyle([
        ("BACKGROUND",(0,0),(-1,0),C_PURPLE),
        ("TOPPADDING",(0,0),(-1,0),3*mm),("BOTTOMPADDING",(0,0),(-1,0),3*mm),
        ("TOPPADDING",(0,1),(-1,-2),2*mm),("BOTTOMPADDING",(0,1),(-1,-2),2*mm),
        ("LEFTPADDING",(0,0),(-1,-1),2*mm),("RIGHTPADDING",(0,0),(-1,-1),2*mm),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("GRID",(0,0),(-1,-2),0.3,C_BORDER),
        ("BACKGROUND",(0,nr),(-1,nr),C_DARK),
        ("TOPPADDING",(0,nr),(-1,nr),3*mm),("BOTTOMPADDING",(0,nr),(-1,nr),3*mm),
        ("SPAN",(0,nr),(3,nr)),
    ])
    for i in range(1,nr):
        if i%2==0: ts2.add("BACKGROUND",(0,i),(-1,i),C_ROW_ALT)
    rt.setStyle(ts2)
    story += [rt, Spacer(1,8*mm)]

    if p.notes:
        story += [HRFlowable(width="100%",thickness=0.5,color=C_BORDER,spaceAfter=3*mm),
                  Paragraph("OBSERVAÇÕES",PS("nhx",fontName="Helvetica-Bold",fontSize=8,textColor=C_GREY,leading=11,spaceAfter=2*mm)),
                  Paragraph(p.notes,S_NOTE),Spacer(1,6*mm)]

    story.append(HRFlowable(width="100%",thickness=0.5,color=C_BORDER,spaceAfter=3*mm))
    ft = Table([[
        Paragraph(f"Gerado em {datetime.utcnow().strftime('%d/%m/%Y %H:%M')} UTC  —  {company_name}", S_NOTE),
        Paragraph(f"Payroll #{p.id:04d}",PS("frx",fontName="Helvetica",fontSize=8,textColor=C_GREY,leading=11,alignment=TA_RIGHT)),
    ]],colWidths=[W*0.7,W*0.3])
    ft.setStyle(TableStyle([("VALIGN",(0,0),(-1,-1),"MIDDLE")]))
    story.append(ft)

    doc.build(story)
    buf.seek(0)
    filename = f"payroll_{p.id:04d}_{splicer_name.replace(' ','_')}.pdf"
    return send_file(buf, mimetype="application/pdf", as_attachment=True, download_name=filename)



@app.route("/payroll/<int:pid>/detail")
@admin_required
def payroll_detail(pid: int):
    """Detalhe de um payroll: lista os lançamentos do período."""
    p = Payroll.query.get_or_404(pid)
    splicer_name = p.user.splicer_name or p.user.username
    q = Record.query.filter(
        Record.splicer == splicer_name,
        Record.created_date >= datetime.combine(p.start_date, datetime.min.time()),
        Record.created_date <= datetime.combine(p.end_date, datetime.max.time()),
    )
    if p.project_id:
        q = q.filter(Record.project_id == p.project_id)
    elif p.company:
        q = q.filter(Record.company == p.company)
    records = q.order_by(Record.created_date).all()
    return render_template("payroll_detail.html", payroll=p, records=records)



@app.route("/export/excel")
@login_required
def export_excel():
    """Exporta os dados de produção em formato Excel (CSV) por empresa e período.

    O arquivo contém: nome do mapa, nome do dispositivo e número de fusões,
    já somados por mapa/dispositivo dentro do filtro.
    """
    company_filter = request.args.get("company") or None
    splicer_filter = request.args.get("splicer") or None
    map_filter = request.args.get("map") or None
    device_filter = request.args.get("device") or None
    start_raw = request.args.get("start") or None
    end_raw = request.args.get("end") or None

    if not company_filter:
        flash("To export Excel, select a company in the filter.", "danger")
        return redirect(url_for("index"))

    query = Record.query

    if company_filter:
        query = query.filter(Record.company == company_filter)
    if splicer_filter:
        query = query.filter(Record.splicer == splicer_filter)
    if map_filter:
        query = query.filter(Record.map == map_filter)

    start_dt = None
    end_dt = None
    if start_raw:
        try:
            start_dt = datetime.fromisoformat(start_raw)
            query = query.filter(Record.created_date >= start_dt)
        except ValueError:
            start_dt = None
    if end_raw:
        try:
            end_dt = datetime.fromisoformat(end_raw)
            query = query.filter(Record.created_date <= end_dt)
        except ValueError:
            end_dt = None

    # se não for admin, força o filtro para o próprio splicer
    if not getattr(current_user, "is_admin", False):
        enforced_splicer = getattr(current_user, "splicer_name", None) or current_user.username
        query = query.filter(Record.splicer == enforced_splicer)

    records = query.order_by(Record.created_date.asc().nullslast(), Record.id.asc()).all()

    if not records:
        flash("No records found for this filter.", "warning")
        return redirect(url_for("index"))

    # agrupar por mapa + dispositivo + data
    grouped = {}
    devices_unique = set()
    for r in records:
        map_name = (r.map or "").strip()
        device_name = (r.device or "").strip()
        date_value = r.created_date.date().isoformat() if r.created_date else None

        key = (map_name, device_name, date_value)
        if key not in grouped:
            grouped[key] = {
                "map": map_name or "-",
                "device": device_name or "-",
                "date": date_value or "",
                "splices": 0,
            }
        grouped[key]["splices"] += int(r.splices or 0)

        if device_name:
            devices_unique.add(device_name)

    lines = list(grouped.values())
    lines.sort(key=lambda x: (x["date"], x["map"], x["device"]))

    # gerar planilha Excel (XLSX) em memória
    wb = Workbook()
    ws = wb.active
    ws.title = "Production"

    # cabeçalho (Date | Map | Device | Splices)
    headers = ["Date", "Map", "Device", "Splices"]
    ws.append(headers)

    # linhas de dados
    total_splices = 0
    for line in lines:
        ws.append([line["date"], line["map"], line["device"], line["splices"]])
        total_splices += int(line["splices"] or 0)

    # linhas de totais
    total_devices = len(devices_unique)
    ws.append([])
    total_row_devices = ws.max_row + 1
    ws.cell(row=total_row_devices, column=1, value="TOTAL DEVICES")
    ws.cell(row=total_row_devices, column=2, value=total_devices)

    total_row_splices = total_row_devices + 1
    ws.cell(row=total_row_splices, column=1, value="TOTAL SPLICES")
    ws.cell(row=total_row_splices, column=4, value=total_splices)

    # aplicar estilos (negrito cabeçalho e totais, bordas)
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # cabeçalho em negrito
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.border = border

    # dados + bordas
    last_row = total_row_splices
    for row in range(2, last_row + 1):
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            # destacar totais
            if row in (total_row_devices, total_row_splices):
                cell.font = Font(bold=True)
            cell.border = border

    # autoajustar largura das colunas
    for col in range(1, 5):
        max_len = 0
        col_letter = ws.cell(row=1, column=col).column_letter
        for row in range(1, last_row + 1):
            val = ws.cell(row=row, column=col).value
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = max_len + 2

    buf = BytesIO()
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"splicer_{company_filter or 'all'}_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
@app.route("/record/<int:rid>/delete")
@login_required
def record_delete(rid: int):
    rec = Record.query.get_or_404(rid)

    # Apenas admin pode apagar qualquer registro.
    # Usuário comum só pode apagar o próprio lançamento.
    if not getattr(current_user, "is_admin", False):
        enforced_splicer = getattr(current_user, "splicer_name", None) or current_user.username
        if rec.splicer != enforced_splicer:
            abort(403)

    # Apaga primeiro as fotos vinculadas a esse lançamento.
    # Isso evita erro de integridade no PostgreSQL, pois a FK de record_photo
    # não aceita valor NULL em record_id.
    RecordPhoto.query.filter_by(record_id=rec.id).delete(synchronize_session=False)

    db.session.delete(rec)
    db.session.commit()
    flash("Registro removido.", "success")
    return redirect(url_for("index"))



@app.route("/expenses", methods=["GET", "POST"])
@login_required
def expenses():
    is_admin = getattr(current_user, "is_admin", False)
    is_owner = getattr(current_user, "is_company_owner", False)

    # Apenas admin ou usuários com permissão explícita podem acessar despesas
    if not is_admin and not getattr(current_user, 'can_access_expenses', False):
        flash("Você não tem permissão para acessar o módulo de despesas.", "danger")
        return redirect(url_for("index"))

    if request.method == "POST":
        description = (request.form.get("description") or "").strip()
        category = (request.form.get("category") or "").strip()
        amount_raw = (request.form.get("amount") or "").replace(",", ".")
        date_raw = request.form.get("date") or ""

        if not description:
            flash("Descrição é obrigatória.", "danger")
            return redirect(url_for("expenses"))

        try:
            amount = float(amount_raw)
        except ValueError:
            flash("Valor inválido para despesa.", "danger")
            return redirect(url_for("expenses"))

        if amount <= 0:
            flash("O valor da despesa deve ser maior que zero.", "danger")
            return redirect(url_for("expenses"))

        try:
            if date_raw:
                y, m, d = map(int, date_raw.split("-"))
                d = date(y, m, d)
            else:
                d = date.today()
        except Exception:
            d = date.today()

        exp = Expense(
            user_id=current_user.id,
            description=description,
            category=category or None,
            amount=amount,
            date=d,
        )
        db.session.add(exp)
        db.session.commit()
        flash("Despesa lançada com sucesso.", "success")
        return redirect(url_for("expenses"))

    user_filter = request.args.get("user_id") or None
    start_raw = request.args.get("start") or None
    end_raw = request.args.get("end") or None

    q = Expense.query

    show_paid = (request.args.get("show_paid") == "1")

    # por padrão mostra só despesas em aberto; com show_paid=1 mostra pagas
    if show_paid:
        q = q.filter(Expense.paid == True)
    else:
        q = q.filter(Expense.paid == False)

    if is_admin:
        if user_filter:
            try:
                q = q.filter(Expense.user_id == int(user_filter))
            except ValueError:
                pass
    else:
        q = q.filter(Expense.user_id == current_user.id)

    if start_raw:
        try:
            y, m, d = map(int, start_raw.split("-"))
            q = q.filter(Expense.date >= date(y, m, d))
        except Exception:
            pass

    if end_raw:
        try:
            y, m, d = map(int, end_raw.split("-"))
            q = q.filter(Expense.date <= date(y, m, d))
        except Exception:
            pass

    q = q.order_by(Expense.date.desc(), Expense.id.desc())
    expenses_list = q.all()
    total_amount = sum(e.amount for e in expenses_list)

    per_user = []
    users_for_filter = []

    if is_admin:
        agg_q = (
            db.session.query(
                User.id,
                User.username,
                User.splicer_name,
                func.sum(Expense.amount).label("total")
            )
            .join(Expense, Expense.user_id == User.id)
        )

        if show_paid:
            agg_q = agg_q.filter(Expense.paid == True)
        else:
            agg_q = agg_q.filter(Expense.paid == False)


        if start_raw:
            try:
                y, m, d = map(int, start_raw.split("-"))
                agg_q = agg_q.filter(Expense.date >= date(y, m, d))
            except Exception:
                pass

        if end_raw:
            try:
                y, m, d = map(int, end_raw.split("-"))
                agg_q = agg_q.filter(Expense.date <= date(y, m, d))
            except Exception:
                pass

        agg_q = agg_q.group_by(User.id, User.username, User.splicer_name)
        per_user = agg_q.all()

        users_for_filter = User.query.order_by(User.username).all()

    return render_template(
        "expenses.html",
        expenses=expenses_list,
        total_amount=total_amount,
        per_user=per_user,
        users_for_filter=users_for_filter,
        user_filter=user_filter,
        start_raw=start_raw,
        end_raw=end_raw,
        show_paid=show_paid,
    )


@app.route("/expenses/mark_paid", methods=["POST"])
@login_required
def expenses_mark_paid():
    is_admin = getattr(current_user, "is_admin", False)
    if not is_admin:
        flash("Acesso negado.", "danger")
        return redirect(url_for("expenses"))

    user_filter = request.form.get("user_id") or None
    start_raw = request.form.get("start") or None
    end_raw = request.form.get("end") or None

    q = Expense.query.filter(Expense.paid == False)

    if user_filter:
        try:
            q = q.filter(Expense.user_id == int(user_filter))
        except ValueError:
            pass

    if start_raw:
        try:
            y, m, d = map(int, start_raw.split("-"))
            q = q.filter(Expense.date >= date(y, m, d))
        except Exception:
            pass

    if end_raw:
        try:
            y, m, d = map(int, end_raw.split("-"))
            q = q.filter(Expense.date <= date(y, m, d))
        except Exception:
            pass

    try:
        updated = q.update(
            {
                Expense.paid: True,
                Expense.paid_at: datetime.utcnow(),
                Expense.paid_by: current_user.id,
            },
            synchronize_session=False,
        )
        db.session.commit()
    except Exception:
        db.session.rollback()
        updated = 0

    flash(f"{updated} despesas marcadas como pagas.", "success")
    return redirect(url_for("expenses", user_id=user_filter or "", start=start_raw or "", end=end_raw or ""))



@app.route("/expenses/delete/<int:expense_id>", methods=["POST"])
@login_required
def expenses_delete(expense_id):
    is_admin = getattr(current_user, "is_admin", False)
    is_owner = getattr(current_user, "is_company_owner", False)

    # Apenas admin ou usuários com permissão explícita podem acessar despesas
    if not is_admin and not getattr(current_user, 'can_access_expenses', False):
        flash("Você não tem permissão para acessar o módulo de despesas.", "danger")
        return redirect(url_for("index"))

    exp = Expense.query.get_or_404(expense_id)

    # Splicer só pode excluir as próprias despesas
    if not is_admin and exp.user_id != current_user.id:
        flash("Você não pode excluir despesas de outro usuário.", "danger")
        return redirect(url_for("expenses"))

    try:
        db.session.delete(exp)
        db.session.commit()
        flash("Despesa excluída com sucesso.", "success")
    except Exception:
        db.session.rollback()
        flash("Erro ao excluir despesa.", "danger")

    return redirect(request.referrer or url_for("expenses"))




def _current_user_company_name():
    """Devolve o nome da empresa do usuário logado (compatível com versões antigas)."""
    return getattr(current_user, "company_name", None) or getattr(current_user, "company", None)



def ensure_map_access(mp: CompanyMap):
    """Garante que o usuário logado pode acessar o mapa informado.

    Regras:
    - Admin: acesso total;
    - Dono da empresa (is_company_owner): acesso total;
    - Caso exista pelo menos um splicer configurado em mp.allowed_splicers,
      somente usuários presentes nessa lista podem acessar;
    - Caso não exista nenhum splicer configurado, qualquer usuário autenticado
      pode acessar o mapa (a proteção fica só por login).
    """
    if not current_user.is_authenticated:
        abort(403)

    # Admin ou dono da empresa: acesso total
    if getattr(current_user, "is_admin", False) or getattr(current_user, "is_company_owner", False):
        return

    # Se houver lista de splicers configurada, exige que o usuário esteja nela
    if mp.allowed_splicers and current_user not in mp.allowed_splicers:
        abort(403)


@app.route("/maps/<int:map_id>/view")
@login_required
def map_view(map_id):
    mp = CompanyMap.query.get_or_404(map_id)

    ensure_map_access(mp)

    # ID opcional de um dispositivo para focar o mapa ao carregar
    focus_record_id = request.args.get("focus_record", type=int)

    return render_template("map_view.html", map_obj=mp, focus_record_id=focus_record_id)

@app.route("/maps/<int:map_id>/delete", methods=["POST"])
@login_required
def delete_map(map_id):
    """
    DESABILITADO: exclusão de mapas foi bloqueada para evitar perda acidental
    de dados de produção. Esta rota existe apenas para compatibilidade, mas
    não executa nenhuma remoção.
    """
    flash("A exclusão de mapas está desabilitada para proteger seus dados. Entre em contato com o administrador do sistema se realmente precisar remover um mapa.", "warning")
    return redirect(url_for("my_maps"))




@app.route("/maps/<int:map_id>/section_colors", methods=["GET", "POST"])
@login_required
def map_section_colors(map_id):
    mp = CompanyMap.query.get_or_404(map_id)
    ensure_map_access(mp)

    # Modo leitura: qualquer usuário com acesso ao mapa pode ler as cores
    if request.method == "GET":
        try:
            colors = json.loads(mp.section_colors_json or "{}")
            if not isinstance(colors, dict):
                colors = {}
        except Exception:
            colors = {}
        return jsonify({"ok": True, "colors": colors})

    # Modo escrita: apenas admin ou dono da empresa podem alterar as cores das seções
    is_admin = bool(getattr(current_user, "is_admin", False))
    is_owner = bool(getattr(current_user, "is_company_owner", False)) and _current_user_company_name() == mp.company
    if not is_admin:
        abort(403)

    try:
        payload = request.get_json(force=True, silent=False) or {}
        colors = payload.get("colors") or {}
        if not isinstance(colors, dict):
            raise ValueError("Formato inválido")
    except Exception:
        return jsonify({"ok": False, "error": "JSON inválido de cores."}), 400

    # Normaliza: apenas strings simples "#rrggbb"
    normalized = {}
    for key, value in colors.items():
        if not isinstance(key, str) or not isinstance(value, str):
            continue
        key = key.strip()
        val = value.strip()
        if not key:
            continue
        normalized[key] = val

    mp.section_colors_json = json.dumps(normalized, ensure_ascii=False)
    db.session.commit()
    return jsonify({"ok": True, "colors": normalized})





@app.route("/api/maps/<int:map_id>/import_kmz", methods=["POST"])
@login_required
def api_import_kmz(map_id):
    """
    Endpoint para importar dispositivos via arquivo KMZ para um mapa específico.
    Usado pela tela de mapa (Meus Mapas -> Abrir mapa).
    """
    mp = CompanyMap.query.get_or_404(map_id)
    ensure_map_access(mp)

    # Apenas admin ou dono da empresa do mapa podem importar KMZ
    is_admin = bool(getattr(current_user, "is_admin", False))
    is_owner = bool(
        getattr(current_user, "is_company_owner", False)
        and _current_user_company_name() == mp.company
    )
    if not is_admin:
        abort(403)

    file_storage = (
        request.files.get("kmz_file")
        or request.files.get("file")
        or None
    )
    if not file_storage or not getattr(file_storage, "filename", None):
        return jsonify({"ok": False, "error": "Arquivo KMZ obrigatório."}), 400

    try:
        imported = import_kmz_for_map(mp, file_storage)
        db.session.commit()
    except Exception:
        db.session.rollback()
        return jsonify({"ok": False, "error": "Falha ao importar KMZ."}), 500

    return jsonify({"ok": True, "imported": imported})


@app.route("/api/maps/<int:map_id>/import_haf", methods=["POST"])
@login_required
def api_import_haf(map_id):
    """
    Importa dispositivos a partir de planilha HAF FIBER (.xlsx).
    Colunas usadas:
      - LAT  (índice 26)
      - LONG (índice 27)
      - COMMENT (índice 22) → usado como nome do dispositivo
      - HOUSE NUMBER (índice 0), STREET NAME (índice 3), STREET TYPE (índice 4),
        PRE DIRECTION (índice 2), POST DIRECTION (índice 5), CITY NAME (índice 11),
        STATE CODE (índice 12), ZIP CODE (índice 13) → montam o endereço da residência
    """
    mp = CompanyMap.query.get_or_404(map_id)
    ensure_map_access(mp)

    is_admin = bool(getattr(current_user, "is_admin", False))
    is_owner = bool(
        getattr(current_user, "is_company_owner", False)
        and _current_user_company_name() == mp.company
    )
    if not is_admin:
        abort(403)

    file_storage = request.files.get("haf_file") or request.files.get("file") or None
    if not file_storage or not getattr(file_storage, "filename", None):
        return jsonify({"ok": False, "error": "Arquivo Excel (.xlsx) obrigatório."}), 400

    filename_lower = (file_storage.filename or "").lower()
    if not (filename_lower.endswith(".xlsx") or filename_lower.endswith(".xls")):
        return jsonify({"ok": False, "error": "Formato inválido. Envie um arquivo .xlsx."}), 400

    try:
        from openpyxl import load_workbook
        import io

        content = file_storage.read()
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active

        imported = 0
        skipped = 0

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            try:
                lat = row[26]
                lon = row[27]
                comment = row[22]

                # Coordenadas são obrigatórias
                if lat is None or lon is None:
                    skipped += 1
                    continue
                try:
                    lat = float(lat)
                    lon = float(lon)
                except (TypeError, ValueError):
                    skipped += 1
                    continue

                # Nome do dispositivo vem do campo COMMENT
                device_name = str(comment).strip() if comment else f"HAF_{row_idx}"

                # Montar endereço legível da residência
                house_num  = str(row[0]).strip() if row[0] is not None else ""
                pre_dir    = str(row[2]).strip() if row[2] else ""
                street     = str(row[3]).strip() if row[3] else ""
                st_type    = str(row[4]).strip() if row[4] else ""
                post_dir   = str(row[5]).strip() if row[5] else ""
                city       = str(row[11]).strip() if row[11] else ""
                state      = str(row[12]).strip() if row[12] else ""
                zipcode    = str(row[13]).strip() if row[13] else ""

                parts = [p for p in [house_num, pre_dir, street, st_type, post_dir] if p]
                street_line = " ".join(parts)
                city_line   = ", ".join(p for p in [city, state, zipcode] if p)
                geo_address = ", ".join(p for p in [street_line, city_line] if p) or None

                # Verifica se já existe registro para este dispositivo neste mapa
                existing = Record.query.filter_by(
                    company=mp.company,
                    map=mp.name,
                    device=device_name,
                ).first()

                if existing:
                    # Atualiza coordenadas e endereço se estiverem vazios
                    if existing.latitude is None:
                        existing.latitude = lat
                    if existing.longitude is None:
                        existing.longitude = lon
                    if not getattr(existing, "geo_address", None) and geo_address:
                        existing.geo_address = geo_address
                else:
                    rec = Record(
                        company=mp.company,
                        map=mp.name,
                        project_id=mp.project_id,
                        device=device_name,
                        latitude=lat,
                        longitude=lon,
                        geo_address=geo_address,
                        splices=0,
                    )
                    db.session.add(rec)
                    imported += 1

            except Exception:
                skipped += 1
                continue

        db.session.commit()
        wb.close()

    except Exception as e:
        db.session.rollback()
        return jsonify({"ok": False, "error": f"Falha ao processar arquivo: {str(e)}"}), 500

    return jsonify({"ok": True, "imported": imported, "skipped": skipped})


@app.route("/api/maps/<int:map_id>/import_splice_report", methods=["POST"])
@login_required
def api_import_splice_report(map_id):
    """
    Lê o arquivo SIGNAL (gerado pelo sistema, ex: SIGNAL_V4.xlsx) e atualiza
    os Records deste mapa com:
      - pon_name      → PON 1, PON 2, ... (pela ordem das abas SE no arquivo)
      - splitter_name → nome do SE (ex: 2911E_SE_001)
      - ote_label     → fibras fusionadas (ex: F1,F2,F3)

    O arquivo deve ter uma aba "RESUMO" com colunas:
      [0] SPLITTER | [1] PORTA | [2] TIPO | [3] DESTINO | [4] Nº FIBRA | [5] ...
    As abas individuais (ex: SE_001, SE_004...) definem a ordem dos PONs.
    O campo 'pon_order' (opcional) permite sobrescrever a ordem: nomes de abas
    separados por vírgula/quebra de linha.
    """
    mp = CompanyMap.query.get_or_404(map_id)
    ensure_map_access(mp)

    is_admin = bool(getattr(current_user, "is_admin", False))
    is_owner = bool(
        getattr(current_user, "is_company_owner", False)
        and _current_user_company_name() == mp.company
    )
    if not is_admin:
        abort(403)

    file_storage = request.files.get("splice_file") or request.files.get("file") or None
    if not file_storage or not getattr(file_storage, "filename", None):
        return jsonify({"ok": False, "error": "Arquivo Excel (.xlsx) obrigatório."}), 400

    filename_lower = (file_storage.filename or "").lower()
    if not (filename_lower.endswith(".xlsx") or filename_lower.endswith(".xls")):
        return jsonify({"ok": False, "error": "Formato inválido. Envie um arquivo .xlsx."}), 400

    try:
        from openpyxl import load_workbook
        import io

        content = file_storage.read()
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)

        # Determine PON order: non-RESUMO sheets in file order, or user override
        pon_order_raw = (request.form.get("pon_order") or "").strip()
        import re as _re
        pon_order_user = [s.strip() for s in _re.split(r'[,;\n]+', pon_order_raw) if s.strip()]

        if pon_order_user:
            pon_sheets = pon_order_user
        else:
            # Auto: all sheets except RESUMO, in file order
            pon_sheets = [s for s in wb.sheetnames if s.upper() != "RESUMO"]

        # Validate sheet names
        missing = [s for s in pon_sheets if s not in wb.sheetnames]
        if missing:
            return jsonify({
                "ok": False,
                "error": f"Aba(s) não encontrada(s): {chr(44).join(missing)}"
            }), 400

        # Build PON name map: sheet_name -> "PON N"
        # Also extract the full splitter name from the sheet data (row 0 col 0)
        sheet_to_pon = {}
        sheet_to_splitter = {}
        for idx, sheet_name in enumerate(pon_sheets, start=1):
            sheet_to_pon[sheet_name] = f"PON {idx}"
            ws_tmp = wb[sheet_name]
            # Try to get the SE name from the title row (first non-empty cell in col A rows 1-3)
            splitter_name = sheet_name  # fallback to sheet name
            for row in ws_tmp.iter_rows(max_row=3, values_only=True):
                val = str(row[0]) if row[0] else ""
                if "SE_" in val or "SPL" in val.upper():
                    splitter_name = val.replace("MAPEAMENTO DE PORTAS  ·  ", "").split("  ·")[0].strip()
                    break
            sheet_to_splitter[sheet_name] = splitter_name

        # Read RESUMO sheet — it has all FT assignments
        # Columns: [0]SPLITTER [1]PORTA [2]TIPO [3]DESTINO [4]Nº FIBRA [5]BUFFER/FIBRA
        ft_data = {}  # ft_name -> {'sheet': str, 'pon': str, 'splitter': str, 'fibers': set}

        if "RESUMO" in wb.sheetnames:
            ws_resumo = wb["RESUMO"]
            for row in ws_resumo.iter_rows(min_row=3, values_only=True):
                splitter_cell = str(row[0]).strip() if row[0] else ""
                dest_type     = str(row[2]).strip() if row[2] else ""
                dest          = str(row[3]).strip() if row[3] else ""
                fnum_raw      = row[4]

                # Only rows that go to a FT device
                if not dest or "_FT_" not in dest:
                    continue
                if dest_type not in ("FT (saída)", "FT (entrada)"):
                    continue
                if fnum_raw is None:
                    continue

                try:
                    fnum = int(fnum_raw)
                except (ValueError, TypeError):
                    continue

                # Find which PON sheet this splitter belongs to
                # splitter_cell in RESUMO is the full SE name (e.g. "2911E_SE_001")
                # sheet names are like "SE_001" — match by suffix
                matched_sheet = None
                for sh in pon_sheets:
                    # Direct match (sheet_name == splitter_cell) or suffix match
                    if sh == splitter_cell:
                        matched_sheet = sh
                        break
                    # e.g. sheet="SE_001", splitter_cell="2911E_SE_001"
                    if splitter_cell.endswith(sh) or sh.endswith(splitter_cell.split("_", 1)[-1] if "_" in splitter_cell else splitter_cell):
                        matched_sheet = sh
                        break

                if not matched_sheet:
                    continue

                pon_name   = sheet_to_pon[matched_sheet]
                splitter_n = splitter_cell  # keep full name from RESUMO

                if dest not in ft_data:
                    ft_data[dest] = {
                        "sheet":    matched_sheet,
                        "pon":      pon_name,
                        "splitter": splitter_n,
                        "fibers":   set(),
                    }
                ft_data[dest]["fibers"].add(fnum)

        wb.close()

        # Load all records for this map
        all_records = Record.query.filter_by(
            company=mp.company,
            map=mp.name,
        ).all()

        rec_by_device = {(r.device or "").strip(): r for r in all_records}

        updated   = 0
        not_found = []

        for ft_device, info in ft_data.items():
            rec = rec_by_device.get(ft_device)

            if not rec:
                # Fallback: suffix match  e.g. "2911E_FT_001" → ends with "FT_001"
                suffix = "_".join(ft_device.split("_")[-2:])
                for dev_name, r in rec_by_device.items():
                    if dev_name.endswith(suffix):
                        rec = r
                        break

            if not rec:
                not_found.append(ft_device)
                continue

            fibers_str = ",".join(f"F{n}" for n in sorted(info["fibers"]))
            rec.pon_name      = info["pon"]
            rec.splitter_name = info["splitter"]
            rec.ote_label     = fibers_str
            updated += 1

        db.session.commit()

    except Exception as e:
        db.session.rollback()
        return jsonify({"ok": False, "error": f"Erro ao processar arquivo: {str(e)}"}), 500

    pon_map = {sheet_to_splitter.get(sh, sh): sheet_to_pon[sh] for sh in pon_sheets}

    return jsonify({
        "ok": True,
        "updated": updated,
        "not_found": not_found,
        "pon_map": pon_map,
        "total_ft_in_file": len(ft_data),
    })


@app.route("/api/maps/<int:map_id>/records", methods=["GET"])
@login_required
def api_map_records(map_id):
    mp = CompanyMap.query.get_or_404(map_id)

    ensure_map_access(mp)

    # Busca registros associados a este mapa, com coordenadas válidas
    query = Record.query.filter(Record.map == mp.name)

    # Se houver empresa configurada no mapa, filtramos também por empresa para maior segurança
    if mp.company:
        query = query.filter(Record.company == mp.company)

    # Apenas registros com latitude/longitude preenchidos
    query = query.filter(Record.latitude.isnot(None), Record.longitude.isnot(None))

    data = []
    for r in query.order_by(Record.id.asc()).all():
        # Separamos miniaturas de lançamento e de teste para mostrar no popup do mapa
        device_photos = [p for p in r.photos if not getattr(p, "is_test", False) and not (p.filename or "").startswith("placed__")][:4]
        test_photos   = [p for p in r.photos if getattr(p, "is_test", False)][:4]
        placed_photos = [p for p in r.photos if not getattr(p, "is_test", False) and (p.filename or "").startswith("placed__")][:4]
        device_type = (r.type or "OTE").strip() or "OTE"

        data.append({
            "id": r.id,
            "lat": r.latitude,
            "device": r.device or "",
            "lng": r.longitude,
            "info": r.device_info or "",
            "ft_lines": _record_ft_lines(r),
            "has_photos": len(device_photos),
            "photo_ids": [p.id for p in device_photos],
            "has_test_photos": len(test_photos),
            "test_photo_ids": [p.id for p in test_photos],
            "placed_photo_ids": [p.id for p in placed_photos],
            "splicer": r.splicer or "",
            "splices": int(r.splices or 0),
            "type": device_type,
            "test_done": bool(r.test_done),
            "test_levels": r.test_levels or "",
            "test_date": r.test_date.isoformat() if r.test_date else None,
            "section": r.section or "",
            "created_date": r.created_date.isoformat() if r.created_date else None,
            "address": getattr(r, 'geo_address', None) or '',
            "pon_name": getattr(r, 'pon_name', None) or '',
            "splitter_name": getattr(r, 'splitter_name', None) or '',
            "source_from": getattr(r, 'source_from', None) or '',
            "source_out": getattr(r, 'source_out', None) or '',
            "ote_label": getattr(r, 'ote_label', None) or '',
            "port_label": getattr(r, 'port_label', None) or '',
            "is_placed": bool(getattr(r, 'is_placed', False)),
            "placed_by": getattr(r, 'placed_by', None) or '',
            "placed_at": r.placed_at.isoformat() if getattr(r, 'placed_at', None) else None,
            "is_active": getattr(r, 'is_active', None) is not False,
            "ribbon_count": getattr(r, 'ribbon_count', None),
        })
    return jsonify({"records": data})


@app.route("/api/records/<int:record_id>/refresh-address", methods=["POST"])
@login_required
def api_refresh_record_address(record_id):
    rec = Record.query.get_or_404(record_id)
    if not bool(getattr(current_user, "is_admin", False)):
        abort(403)
    if rec.latitude is None or rec.longitude is None:
        return jsonify({"ok": False, "error": "Dispositivo sem coordenadas."}), 400
    address = geoapify_reverse_geocode(rec.latitude, rec.longitude)
    if not address:
        return jsonify({"ok": False, "error": "Geoapify não retornou endereço."}), 400
    rec.geo_address = address
    db.session.commit()
    return jsonify({"ok": True, "address": address})


@app.route("/api/maps/<int:map_id>/auto-network", methods=["POST"])
@login_required
def api_auto_network(map_id):
    mp = CompanyMap.query.get_or_404(map_id)
    if not bool(getattr(current_user, "is_admin", False)):
        abort(403)
    ensure_map_access(mp)
    updated = build_network_for_map(mp)
    return jsonify({"ok": True, "updated": updated})


@app.route("/api/records/<int:record_id>/update-from-map", methods=["POST"])
@login_required
def api_update_record_from_map(record_id):
    rec = Record.query.get_or_404(record_id)

    # Permissão básica: apenas mesma empresa ou admin
    if (not current_user.is_admin) and _current_user_company_name() != rec.company:
        is_owner = getattr(current_user, "is_company_owner", False)
        if not is_owner or _current_user_company_name() != rec.company:
            abort(403)

    # Atualiza dados básicos vindos do formulário do mapa
    is_admin = bool(getattr(current_user, "is_admin", False))
    is_owner = bool(getattr(current_user, "is_company_owner", False) and _current_user_company_name() == rec.company)

    splices_raw = (request.form.get("splices") or "").strip()
    if is_admin:
        try:
            if splices_raw != "":
                rec.splices = int(splices_raw)
        except ValueError:
            pass

    info = (request.form.get("device_info") or "").strip()
    if is_admin:
        new_device_name = (request.form.get('device_name') or '').strip()
        if new_device_name:
            rec.device = new_device_name
        rec.device_info = info or None
        rec.type = (request.form.get('device_type') or rec.type or 'OTE').strip() or 'OTE'
        rec.geo_address = (request.form.get('geo_address') or '').strip() or None
        rec.pon_name = (request.form.get('pon_name') or '').strip() or None
        rec.splitter_name = (request.form.get('splitter_name') or '').strip() or None
        rec.source_from = (request.form.get('source_from') or '').strip() or None
        rec.source_out = (request.form.get('source_out') or '').strip() or None
        rec.ote_label = (request.form.get('ote_label') or '').strip() or None
        rec.port_label = (request.form.get('port_label') or '').strip() or None
        rec.section = (request.form.get('section') or rec.section or '').strip() or None

    # Salvar pelo editor do mapa NÃO deve transformar o dispositivo em lançamento
    # nem alterar o splicer para ADMIN. Só recalculamos preços quando o registro
    # já era um lançamento real (splices/valores existentes).
    already_launched = bool((rec.splices or 0) > 0 or (rec.total_usd or 0) > 0 or (rec.price_device_usd or 0) > 0 or (rec.price_splices_usd or 0) > 0)

    if already_launched:
        company = rec.company
        project_id = rec.project_id
        device_for_price = rec.type or rec.device

        included_override = None
        included_applied = None
        included_override, included_applied, map_cfg = resolve_included_override(
            company,
            project_id,
            None,
            rec.map,
            rec.map_role,
        )

        price_splices, price_device, total = compute_prices(
            int(rec.splices or 0),
            device_for_price or "",
            company,
            project_id,
            included_override=included_override,
            map_role=rec.map_role,
            ribbon_count=rec.ribbon_count,
        )
        rec.price_splices_usd = price_splices
        rec.price_device_usd = price_device
        rec.total_usd = total
        rec.included_splices_applied = included_applied
    else:
        # Limpa qualquer resíduo financeiro/splicer em dispositivos de mapa sem produção
        rec.splicer = None
        rec.created_date = None
        rec.price_splices_usd = 0.0
        rec.price_device_usd = 0.0
        rec.total_usd = 0.0
        rec.included_splices_applied = None

    # Fotos opcionais
    files = request.files.getlist("photos")
    if files:
        for f in files[:5]:
            if not f or not getattr(f, "filename", None):
                continue
            filename, content_type, data, thumb_data, thumb_ct = process_uploaded_photo(f)
            if not data:
                continue

            photo = RecordPhoto(
                record_id=rec.id,
                filename=(filename or "photo")[:255],
                content_type=content_type,
                data=data,
                thumb_data=thumb_data,
                thumb_content_type=thumb_ct,
                size_bytes=len(data),
            )
            db.session.add(photo)

    db.session.commit()
    return jsonify({"ok": True})

@app.route("/api/records/<int:record_id>/toggle-active", methods=["POST"])
@login_required
def api_toggle_record_active(record_id):
    """Ativa ou desativa um dispositivo no mapa (pin fica vermelho). Somente admin."""
    if not bool(getattr(current_user, "is_admin", False)):
        return jsonify({"ok": False, "error": "Somente admin pode desativar dispositivos."}), 403
    rec = Record.query.get_or_404(record_id)
    # Toggle: None/True -> False, False -> True
    currently_active = getattr(rec, "is_active", None) is not False
    rec.is_active = not currently_active
    db.session.commit()
    return jsonify({"ok": True, "is_active": rec.is_active})


@app.route("/api/records/<int:record_id>/delete-from-map", methods=["POST"])
@login_required
def api_delete_record_from_map(record_id):
    """Remove permanentemente um dispositivo do mapa. Somente admin."""
    if not bool(getattr(current_user, "is_admin", False)):
        return jsonify({"ok": False, "error": "Somente admin pode excluir dispositivos."}), 403
    rec = Record.query.get_or_404(record_id)
    # Remove fotos associadas
    for photo in list(rec.photos):
        db.session.delete(photo)
    db.session.delete(rec)
    db.session.commit()
    return jsonify({"ok": True})


@app.route("/api/records/<int:record_id>/update-coords", methods=["POST"])
@login_required
def api_update_record_coords(record_id):
    rec = Record.query.get_or_404(record_id)

    # Permissão: somente ADMIN pode mover coordenadas
    if not current_user.is_admin:
        abort(403)

    data = request.get_json(silent=True) or {}
    try:
        lat = float(data.get("lat"))
        lng = float(data.get("lng"))
    except Exception:
        return jsonify({"ok": False, "error": "Lat/Lng inválidos."}), 400

    if not (-90 <= lat <= 90 and -180 <= lng <= 180):
        return jsonify({"ok": False, "error": "Lat/Lng fora do intervalo."}), 400

    rec.latitude = lat
    rec.longitude = lng
    db.session.commit()

    return jsonify({"ok": True, "id": rec.id, "lat": rec.latitude, "lng": rec.longitude})



@app.route("/api/records/<int:record_id>/save-test", methods=["POST"])
@login_required
def api_save_record_test(record_id):
    rec = Record.query.get_or_404(record_id)
    # Dispositivos CAN não exigem teste: bloqueia salvamento de teste
    if (rec.type or "").strip().upper().startswith("CAN"):
        return jsonify({"ok": False, "error": "Dispositivo CAN não exige teste."}), 400

    # Regra de permissão igual à tela de visualização do lançamento:
    # 1) Admin ou dono da empresa: acesso total
    is_admin = bool(getattr(current_user, "is_admin", False))
    is_owner = bool(getattr(current_user, "is_company_owner", False))

    if not is_admin:
        # 2) Tenta resolver o mapa desse record e usar a mesma lógica de acesso de mapas
        mp = None
        if rec.map:
            mp = CompanyMap.query.filter_by(name=rec.map, company=rec.company).first()

        if mp is not None:
            # Usa a regra centralizada de acesso a mapas
            ensure_map_access(mp)
        else:
            # Fallback: mesma empresa ou mesmo splicer do registro
            if _current_user_company_name() != rec.company:
                current_splicer = (getattr(current_user, "splicer_name", None) or current_user.username)
                if (rec.splicer or "") != current_splicer:
                    abort(403)


    levels_raw = (request.form.get("levels") or "").strip()
    rec.test_levels = levels_raw or None

    saved_test_photos = 0

    # Fotos de teste (opcional)
    files = request.files.getlist("photos")
    if files:
        for f in files[:5]:
            if not f or not getattr(f, "filename", None):
                continue
            filename, content_type, data, thumb_data, thumb_ct = process_uploaded_photo(f)
            if not data:
                continue

            photo = RecordPhoto(
                record_id=rec.id,
                filename=(filename or "test")[:255],
                content_type=content_type,
                data=data if data else b"",
                thumb_data=thumb_data,
                thumb_content_type=thumb_ct,
                size_bytes=int(len(data) if data else 0),
                is_test=True,
            )
            db.session.add(photo)
            saved_test_photos += 1

    rec.test_done = bool(levels_raw or saved_test_photos)
    if rec.test_done:
        rec.test_date = datetime.utcnow()
    db.session.commit()

    return jsonify({
        "ok": True,
        "record_id": int(rec.id),
        "test_done": bool(rec.test_done),
        "saved_test_photos": int(saved_test_photos),
    })


@app.route("/api/records/<int:record_id>/set-placed", methods=["POST"])
@login_required
def api_set_record_placed(record_id):
    """Marca um dispositivo como 'subido' (placed) e opcionalmente salva uma foto de instalação.
    Grava quem fez a subida (placed_by/placed_at) mas NUNCA altera o splicer original."""
    rec = Record.query.get_or_404(record_id)

    # Permissão: admin, owner ou splicer do mesmo mapa
    is_admin = bool(getattr(current_user, "is_admin", False))
    is_owner = bool(getattr(current_user, "is_company_owner", False))

    if not is_admin:
        mp = None
        if rec.map:
            mp = CompanyMap.query.filter_by(name=rec.map, company=rec.company).first()
        if mp is not None:
            ensure_map_access(mp)
        else:
            if _current_user_company_name() != rec.company:
                current_splicer = (getattr(current_user, "splicer_name", None) or current_user.username)
                if (rec.splicer or "") != current_splicer:
                    abort(403)

    # Quem está realizando a subida
    actor_name = (
        getattr(current_user, "splicer_name", None)
        or current_user.username
        or "Desconhecido"
    )

    # Salvar foto de subida (marcada com tag especial via is_test=False mas com photo_tag)
    # Usamos is_test=False e um flag de 'placed' via filename prefix para distinguir.
    files = request.files.getlist("photos")
    saved_placed_photos = 0
    if files:
        for f in files[:5]:
            if not f or not getattr(f, "filename", None):
                continue
            filename, content_type, data, thumb_data, thumb_ct = process_uploaded_photo(f)
            if not data:
                continue
            # Prefixamos o filename para identificar como foto de subida
            safe_fn = ("placed__" + (filename or "foto.jpg"))[:255]
            photo = RecordPhoto(
                record_id=rec.id,
                filename=safe_fn,
                content_type=content_type,
                data=data,
                thumb_data=thumb_data,
                thumb_content_type=thumb_ct,
                size_bytes=int(len(data)),
                is_test=False,
            )
            db.session.add(photo)
            # Enfileira upload para R2 se habilitado
            if r2_enabled():
                db.session.flush()
                r2_key = r2_key_for_record_photo(photo.id, safe_fn)
                photo.r2_key = r2_key
                if thumb_data:
                    thumb_key = r2_key.replace("/full/", "/thumb/")
                    photo.r2_thumb_key = thumb_key
                else:
                    thumb_key = None
                enqueue_r2_upload(photo.id, r2_key, data, content_type or "image/jpeg",
                                  thumb_key=thumb_key, thumb_bytes=thumb_data,
                                  thumb_content_type=thumb_ct)
            saved_placed_photos += 1

    # Marca como subido — preserva placed_by original se já havia sido subido antes
    # (só sobrescreve se for a primeira vez ou se admin forçar)
    force = request.form.get("force") == "1"
    if not rec.is_placed or force:
        rec.is_placed = True
        rec.placed_by = actor_name
        rec.placed_at = datetime.utcnow()

    db.session.commit()

    return jsonify({
        "ok": True,
        "record_id": int(rec.id),
        "is_placed": bool(rec.is_placed),
        "placed_by": rec.placed_by or "",
        "placed_at": rec.placed_at.isoformat() if rec.placed_at else None,
        "saved_photos": saved_placed_photos,
    })



@app.route("/maps/<int:map_id>/tests-report")
@login_required
def export_map_tests_report(map_id):
    mp = CompanyMap.query.get_or_404(map_id)
    ensure_map_access(mp)
    if not bool(getattr(current_user, "is_admin", False)):
        abort(403)

    query = Record.query.filter(Record.map == mp.name)
    if mp.company:
        query = query.filter(Record.company == mp.company)

    query = query.filter(Record.test_done.is_(True))

    rows = []
    max_ports = 0
    for r in query.order_by(func.coalesce(Record.test_date, Record.created_date).asc(), Record.device.asc(), Record.id.asc()).all():
        if (r.type or "").strip().upper().startswith("CAN"):
            continue
        levels = [x.strip() for x in (r.test_levels or "").split(",") if x.strip()]
        if not levels:
            continue
        max_ports = max(max_ports, len(levels))
        test_dt = r.test_date or r.created_date
        rows.append({
            "Date": test_dt.strftime("%Y-%m-%d") if test_dt else "",
            "Map": r.map or mp.name or "",
            "Device": r.device or "",
            "ports": levels,
        })

    # Permite CSV opcional via ?format=csv
    out_format = (request.args.get("format") or "pdf").strip().lower()
    if out_format == "csv":
        output = io.StringIO()
        fieldnames = ["Date", "Map", "Device"] + [f"Porta {i}" for i in range(1, max_ports + 1)]
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            base = {"Date": row["Date"], "Map": row["Map"], "Device": row["Device"]}
            for i in range(max_ports):
                base[f"Porta {i+1}"] = row["ports"][i] if i < len(row["ports"]) else ""
            writer.writerow(base)
        csv_bytes = output.getvalue().encode("utf-8-sig")
        filename = f"test-report-{(mp.name or 'map').replace(' ', '_')}.csv"
        return send_file(io.BytesIO(csv_bytes), mimetype="text/csv", as_attachment=True, download_name=filename)

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=10)
    pdf.add_page()
    pdf.set_font("Arial", "B", 14)
    title = f"Relatorio de Testes - {mp.company or ''} - {mp.name or ''}".strip(" -")
    pdf.cell(0, 10, title, ln=1)
    pdf.set_font("Arial", size=9)
    pdf.cell(0, 6, f"Gerado em: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", ln=1)
    pdf.ln(2)

    page_w = pdf.w - pdf.l_margin - pdf.r_margin
    date_w = 28
    map_w = 50
    device_w = 52
    port_cols = max(max_ports, 1)
    port_w = max(18, min(28, (page_w - date_w - map_w - device_w) / port_cols))

    pdf.set_font("Arial", "B", 9)
    headers = ["Date", "Map", "Device"] + [f"Porta {i}" for i in range(1, max_ports + 1)]
    widths = [date_w, map_w, device_w] + [port_w] * max_ports
    for h, w in zip(headers, widths):
        pdf.cell(w, 8, h, border=1, align="C")
    pdf.ln()

    pdf.set_font("Arial", size=8)
    for row in rows:
        values = [row["Date"], row["Map"], row["Device"]] + row["ports"] + [""] * (max_ports - len(row["ports"]))
        for val, w in zip(values, widths):
            text = str(val)[:40]
            pdf.cell(w, 7, text, border=1)
        pdf.ln()

    pdf_bytes = bytes(pdf.output(dest="S"))
    filename = f"test-report-{(mp.name or 'map').replace(' ', '_')}.pdf"
    return send_file(io.BytesIO(pdf_bytes), mimetype="application/pdf", as_attachment=True, download_name=filename)


@app.route("/api/maps/<int:map_id>/add-record", methods=["POST"])
@login_required
def api_add_record_to_map(map_id):
    mp = CompanyMap.query.get_or_404(map_id)

    ensure_map_access(mp)

    device_name = (request.form.get("device") or "").strip()
    lat_raw = (request.form.get("lat") or "").strip()
    lng_raw = (request.form.get("lng") or "").strip()
    device_info = (request.form.get("device_info") or "").strip()

    if not device_name or not lat_raw or not lng_raw:
        return jsonify({"ok": False, "error": "Campos obrigatórios faltando."}), 400

    try:
        lat = float(lat_raw)
        lng = float(lng_raw)
    except ValueError:
        return jsonify({"ok": False, "error": "Coordenadas inválidas."}), 400

    # Regra: um dispositivo não pode ser lançado duas vezes no mesmo mapa/projeto.
    existing = Record.query.filter(
        Record.map == mp.name,
        Record.device == device_name,
        Record.company == mp.company,
    )
    if mp.project_id is not None:
        existing = existing.filter(Record.project_id == mp.project_id)
    existing = existing.order_by(Record.id.desc()).first()
    if existing is not None:
        return jsonify({
            "ok": False,
            "error": "Este dispositivo já foi lançado neste mapa. Peça ao admin para excluir ou alterar o lançamento anterior.",
        }), 400

    rec = Record(
        map=mp.name,
        company=mp.company,
        project_id=mp.project_id,
        device=device_name,
        type=None,
        splices=0,
        splicer="",
        created_date=None,
        latitude=lat,
        longitude=lng,
        device_info=device_info or None,
    )
    db.session.add(rec)
    db.session.commit()

    return jsonify({"ok": True, "id": rec.id})







@app.route("/admin/restore_rn51e", methods=["GET", "POST"])
@admin_required
def admin_restore_rn51e():
    """Restaura lançamentos do mapa RN51E (LORENZ TECH) a partir de um CSV.

    Regras importantes:
    - Somente ADMIN pode usar.
    - Nunca substitui nada que já exista.
      * Para cada linha, se já houver um Record com mesmo company+map+device,
        a linha é ignorada.
    - Apenas company='LORENZ TECH' e map='RN51E' são aceitos.
    """
    inserted = 0
    skipped = 0
    errors = []

    if request.method == "POST":
        file = request.files.get("csv_file")
        if not file or file.filename == "":
            flash("Selecione um arquivo CSV para importar.", "danger")
            return redirect(url_for("admin_restore_rn51e"))

        try:
            from io import TextIOWrapper
            import csv

            wrapped = TextIOWrapper(file.stream, encoding="utf-8")
            reader = csv.DictReader(wrapped)

            for idx, row in enumerate(reader, start=1):
                try:
                    company = (row.get("company") or "").strip()
                    map_name = (row.get("map") or "").strip()
                    device = (row.get("device") or "").strip()

                    if not company or not map_name or not device:
                        skipped += 1
                        errors.append(f"Linha {idx}: company/map/device vazios.")
                        continue

                    # Garante que só LORENZ TECH / RN51E seja importado
                    if company != "LORENZ TECH" or map_name != "RN51E":
                        skipped += 1
                        continue

                    # Verifica se já existe algum registro para esse company+map+device
                    existing = (
                        Record.query
                        .filter_by(company=company, map=map_name, device=device)
                        .first()
                    )
                    if existing:
                        skipped += 1
                        continue

                    # Converte campos numéricos com segurança
                    def to_int(val):
                        try:
                            return int(str(val).strip())
                        except Exception:
                            return 0

                    def to_float(val):
                        try:
                            return float(str(val).strip())
                        except Exception:
                            return 0.0

                    map_role = (row.get("tipo") or row.get("map_role") or "").strip().upper() or None
                    included_splices = to_int(row.get("incl") or row.get("included_splices_applied") or 0)
                    splices = to_int(row.get("splices") or 0)
                    price_device = to_float(row.get("device_price_usd") or 0.0)
                    total_usd = to_float(row.get("total_usd") or 0.0)
                    splicer = (row.get("splicer") or "DESCONHECIDO").strip()

                    rec = Record(
                        company=company,
                        map=map_name,
                        device=device,
                        splicer=splicer,
                        map_role=map_role,
                        included_splices_applied=included_splices,
                        splices=splices,
                        price_device_usd=price_device,
                        total_usd=total_usd,
                    )
                    db.session.add(rec)
                    inserted += 1

                except Exception as e:
                    skipped += 1
                    errors.append(f"Linha {idx}: erro {e}")

            db.session.commit()
            flash(
                f"Importação concluída. Inseridos: {inserted}, pulados: {skipped}.",
                "success",
            )
        except Exception as e:
            db.session.rollback()
            flash(f"Erro ao processar CSV: {e}", "danger")

        return redirect(url_for("admin_restore_rn51e"))

    return render_template(
        "admin_restore_rn51e.html",
        inserted=inserted,
        skipped=skipped,
        errors=errors,
    )

@app.route("/admin/ribbon-debug")
@admin_required
def admin_ribbon_debug():
    """Mostra estado dos registros ribbon para diagnostico."""
    ribbon_types = DeviceType.query.filter(DeviceType.is_ribbon == True).all()
    ribbon_map = {dt.name.lower(): dt for dt in ribbon_types}

    # Busca TODOS os records com type contendo RIBBON (case insensitive)
    from sqlalchemy import func
    ribbon_records = Record.query.filter(
        func.lower(Record.type).contains('ribbon')
    ).all()

    lines = []
    lines.append(f"=== DeviceTypes ribbon cadastrados ===")
    for dt in ribbon_types:
        lines.append(f"  id={dt.id} name='{dt.name}' is_ribbon={dt.is_ribbon} ribbon_price={dt.ribbon_price_usd} company={dt.company} project_id={dt.project_id}")

    lines.append(f"\n=== Records com type contendo RIBBON ===")
    for rec in ribbon_records:
        lines.append(f"  id={rec.id} type='{rec.type}' splices={rec.splices} ribbon_count={rec.ribbon_count} total={rec.total_usd} company={rec.company} project_id={rec.project_id}")
        rtype = (rec.type or "").lower().strip()
        dt = ribbon_map.get(rtype)
        lines.append(f"    -> DeviceType match: {dt.name if dt else 'NENHUM'} | ribbon_price={dt.ribbon_price_usd if dt else 'N/A'}")

    return "<pre>" + "\n".join(lines) + "</pre>"


@app.route("/admin/fix-ribbon", methods=["GET", "POST"])
@admin_required
def admin_fix_ribbon():
    """Corrige registros ribbon com preco zerado. Aceita record_id + ribbon_count manual."""
    fixed_records = 0
    errors = []

    # 1) Garante FALSE em DeviceTypes com is_ribbon NULL
    try:
        db.session.execute(text("UPDATE device_type SET is_ribbon = FALSE WHERE is_ribbon IS NULL"))
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        errors.append(f"Fix NULL is_ribbon: {e}")

    # 2) Correcao manual: record_id + ribbon_count passados via GET ou POST
    manual_id_raw = (request.values.get("record_id") or "").strip()
    manual_count_raw = (request.values.get("ribbon_count") or "").strip()
    if manual_id_raw.isdigit() and manual_count_raw.isdigit():
        try:
            rec = Record.query.get(int(manual_id_raw))
            if rec:
                ribbon_types = DeviceType.query.filter(DeviceType.is_ribbon == True).all()
                ribbon_map = {dt.name.lower(): dt for dt in ribbon_types}
                rtype = (rec.type or "").lower().strip()
                dt = ribbon_map.get(rtype)
                if dt and dt.ribbon_price_usd:
                    rcount = int(manual_count_raw)
                    price_fitas = float(rcount) * float(dt.ribbon_price_usd)
                    price_enc = float(dt.value_usd or 0.0)
                    rec.ribbon_count = rcount
                    rec.splices = 0
                    rec.price_splices_usd = price_fitas
                    rec.price_device_usd = price_enc
                    rec.total_usd = price_fitas + price_enc
                    db.session.commit()
                    fixed_records += 1
                else:
                    errors.append(f"Record {manual_id_raw}: DeviceType nao encontrado ou sem preco ribbon.")
        except Exception as e:
            db.session.rollback()
            errors.append(f"Correcao manual: {e}")

    # 3) Recalcula automaticamente registros ribbon com ribbon_count preenchido e total zerado
    try:
        ribbon_types = DeviceType.query.filter(DeviceType.is_ribbon == True).all()
        ribbon_map = {dt.name.lower(): dt for dt in ribbon_types}
        if ribbon_map:
            all_records = Record.query.filter(
                Record.total_usd == 0.0,
                Record.ribbon_count.isnot(None),
            ).all()
            for rec in all_records:
                rtype = (rec.type or "").lower().strip()
                dt = ribbon_map.get(rtype)
                if dt and dt.ribbon_price_usd and rec.ribbon_count:
                    price_fitas = float(rec.ribbon_count) * float(dt.ribbon_price_usd)
                    price_enc = float(dt.value_usd or 0.0)
                    rec.price_splices_usd = price_fitas
                    rec.price_device_usd = price_enc
                    rec.total_usd = price_fitas + price_enc
                    rec.splices = 0
                    fixed_records += 1
            db.session.commit()
    except Exception as e:
        db.session.rollback()
        errors.append(f"Fix automatico: {e}")

    msg = f"Corrigidos: {fixed_records} lancamentos ribbon."
    if errors:
        msg += " Erros: " + "; ".join(errors)
    flash(msg, "success" if not errors else "warning")
    return redirect(url_for("index"))


# ═══════════════════════════════════════════════════════
# HORAS TRABALHADAS
# ═══════════════════════════════════════════════════════

@app.route("/hours", methods=["GET", "POST"])
@login_required
def hour_entry():
    """Tela de lancamento de horas trabalhadas."""
    is_admin = bool(getattr(current_user, "is_admin", False))

    # Carrega empresas e projetos para o seletor
    companies = []
    projects_by_company = {}
    if is_admin:
        companies = sorted(set(
            r.company for r in Record.query.with_entities(Record.company).distinct()
            if r.company
        ))
        for proj in Project.query.order_by(Project.name).all():
            projects_by_company.setdefault(proj.company or "", []).append({
                "id": proj.id, "name": proj.name
            })
    else:
        company = getattr(current_user, "company_name", None) or ""
        if company:
            companies = [company]
            for proj in Project.query.filter_by(company=company).order_by(Project.name).all():
                projects_by_company.setdefault(company, []).append({
                    "id": proj.id, "name": proj.name
                })

    # Carrega taxas horárias disponíveis
    hourly_rates = HourlyRate.query.order_by(HourlyRate.company, HourlyRate.description).all()

    # Carrega mapas disponíveis por projeto
    maps_by_project = {}
    for m in CompanyMap.query.order_by(CompanyMap.name).all():
        maps_by_project.setdefault(str(m.project_id) if m.project_id else m.company or "", []).append(m.name)

    errors = []
    if request.method == "POST":
        company = (request.form.get("company") or "").strip() or None
        project_id_raw = (request.form.get("project_id") or "").strip()
        project_id = int(project_id_raw) if project_id_raw.isdigit() else None
        splicer = (request.form.get("splicer") or "").strip() or (
            getattr(current_user, "splicer_name", None) or current_user.username
        )
        hours_raw = (request.form.get("hours") or "0").strip()
        description = (request.form.get("description") or "").strip() or None
        map_name = (request.form.get("map_name") or "").strip() or None
        rate_id_raw = (request.form.get("rate_id") or "").strip()
        created_raw = (request.form.get("created") or "").strip()

        try:
            hours = float(hours_raw or 0)
        except ValueError:
            hours = 0.0
            errors.append("Número de horas inválido.")

        rate = HourlyRate.query.get(int(rate_id_raw)) if rate_id_raw.isdigit() else None
        if not rate:
            errors.append("Selecione uma taxa horária.")

        try:
            created_date = datetime.strptime(created_raw, "%Y-%m-%d") if created_raw else datetime.utcnow()
        except ValueError:
            created_date = datetime.utcnow()

        if not errors and hours > 0 and rate:
            total = hours * rate.rate_usd
            hr = HourRecord(
                company=company,
                project_id=project_id,
                splicer=splicer,
                created_date=created_date,
                hours=hours,
                description=description,
                map_name=map_name,
                rate_usd=rate.rate_usd,
                total_usd=total,
                billing_code=rate.billing_code,
            )
            db.session.add(hr)
            db.session.commit()
            flash(f"Lançamento de {hours}h registrado. Total: ${total:.2f}", "success")
            return redirect(url_for("hour_entry"))
        elif not errors:
            errors.append("Informe um número de horas válido.")

    today = datetime.utcnow().strftime("%Y-%m-%d")
    splicer_options = []
    if is_admin:
        splicer_options = [
            (u.splicer_name or u.username)
            for u in User.query.order_by(User.username).all()
            if (u.splicer_name or u.username)
        ]

    return render_template("hour_entry.html",
        companies=companies,
        projects_by_company=projects_by_company,
        maps_by_project=maps_by_project,
        hourly_rates=hourly_rates,
        splicer_options=splicer_options,
        errors=errors,
        today=today,
        is_admin=is_admin,
    )


@app.route("/hours/list")
@login_required
def hour_list():
    """Lista de lançamentos de horas."""
    is_admin = bool(getattr(current_user, "is_admin", False))
    company_filter = request.args.get("company") or None
    start_raw = request.args.get("start") or None
    end_raw = request.args.get("end") or None

    q = HourRecord.query
    if not is_admin:
        enforced = getattr(current_user, "splicer_name", None) or current_user.username
        q = q.filter(HourRecord.splicer == enforced)
    if company_filter:
        q = q.filter(HourRecord.company == company_filter)
    if start_raw:
        try:
            q = q.filter(HourRecord.created_date >= datetime.fromisoformat(start_raw))
        except ValueError:
            pass
    if end_raw:
        try:
            q = q.filter(HourRecord.created_date <= datetime.fromisoformat(end_raw))
        except ValueError:
            pass

    records = q.order_by(HourRecord.created_date.desc()).all()
    total = sum(r.total_usd or 0 for r in records)
    return render_template("hour_list.html", records=records, total=total,
                           company_filter=company_filter or "",
                           start=start_raw or "", end=end_raw or "",
                           is_admin=is_admin)


@app.route("/hours/<int:hid>/delete", methods=["POST"])
@admin_required
def hour_delete(hid: int):
    hr = HourRecord.query.get_or_404(hid)
    db.session.delete(hr)
    db.session.commit()
    flash("Lançamento de horas removido.", "success")
    return redirect(url_for("hour_list"))


# ── Configuração de taxas horárias ─────────────────────

@app.route("/settings/hourly-rate/add", methods=["POST"])
@admin_required
def settings_hourly_rate_add():
    company = (request.form.get("company") or "").strip() or None
    project_id_raw = (request.form.get("project_id") or "").strip()
    project_id = int(project_id_raw) if project_id_raw.isdigit() else None
    next_url = (request.form.get("next") or "").strip() or None
    try:
        rate = float(request.form.get("rate_usd") or 0)
    except ValueError:
        rate = 0.0
    description = (request.form.get("description") or "").strip() or None
    billing_code = (request.form.get("billing_code") or "").strip() or None

    hr = HourlyRate(
        company=company, project_id=project_id,
        rate_usd=rate, description=description, billing_code=billing_code,
    )
    db.session.add(hr)
    db.session.commit()
    flash("Taxa horária salva.", "success")
    return redirect(next_url or url_for("settings"))


@app.route("/settings/hourly-rate/<int:rid>/delete")
@admin_required
def settings_hourly_rate_delete(rid: int):
    next_url = request.args.get("next") or None
    hr = HourlyRate.query.get_or_404(rid)
    db.session.delete(hr)
    db.session.commit()
    flash("Taxa horária removida.", "success")
    return redirect(next_url or url_for("settings"))


if __name__ == "__main__":
    app.run(debug=True)





# ─────────────────────────────────────────────────────────────
#  AUTO PHOTO LAUNCH — lançamento automático via foto do mapa
# ─────────────────────────────────────────────────────────────
@app.route("/api/maps/<int:map_id>/auto-photo-launch", methods=["POST"])
@login_required
def auto_photo_launch(map_id):
    """Recebe uma ou mais fotos Timemark do mapa e faz lançamento automático.

    Para cada foto:
      1. Claude lê a imagem e extrai: device_name, splices, map_role, ft_in, ft_out, gps, datetime
      2. O sistema localiza o Record no banco pelo device_name + map
      3. Atualiza o record com os dados extraídos (splices, ft_in, ft_out, map_role, preços)
      4. Salva a foto no record
      5. Retorna JSON com resultado por foto

    Não exige que o splicer preencha nada — 100% automático.
    """
    mp = CompanyMap.query.get_or_404(map_id)

    is_owner = bool(getattr(current_user, "is_company_owner", False))
    is_admin = bool(getattr(current_user, "is_admin", False))
    if is_owner and not is_admin:
        return jsonify({"ok": False, "error": "Dono de empresa não pode lançar produção."}), 403

    photos = [p for p in request.files.getlist("photos") if getattr(p, "filename", None)]
    if not photos:
        return jsonify({"ok": False, "error": "Nenhuma foto enviada."}), 400

    results = []
    company = mp.company
    project_id = mp.project_id
    map_name = mp.name
    splicer_name = getattr(current_user, "splicer_name", None) or current_user.username

    for photo_file in photos:
        fname = photo_file.filename or "foto.jpg"
        try:
            raw_bytes = photo_file.read()
        except Exception as e:
            results.append({"file": fname, "ok": False, "error": f"Erro ao ler arquivo: {e}"})
            continue

        # 1. Chama Claude para extrair dados da foto
        parsed, ai_error = extract_timestamp_fields_with_ai(raw_bytes)
        if parsed is None:
            results.append({"file": fname, "ok": False, "error": ai_error or "Claude não conseguiu ler a foto."})
            continue

        device_name = parsed["device_name"]
        splices_val  = parsed["splices"]
        map_role     = parsed["map_role"]   # 'MEIO' ou 'PONTA'
        ft_in        = (parsed.get("ft_in") or "").strip() or None
        ft_out       = (parsed.get("ft_out") or "").strip() or None
        if map_role == "PONTA":
            ft_out = None

        # 2. Localiza o Record no banco
        rec = Record.query.filter(
            Record.map == map_name,
            Record.device == device_name,
            Record.company == company,
        ).order_by(Record.id.asc()).first()

        if not rec:
            results.append({
                "file": fname, "ok": False,
                "error": f"Dispositivo '{device_name}' não encontrado no mapa '{map_name}'."
            })
            continue

        # 3. Resolve preços e billing codes
        type_val = (rec.type or "OTE").strip() or "OTE"
        device_for_price = type_val or device_name

        map_obj = CompanyMap.query.filter(
            CompanyMap.company == company,
            CompanyMap.name == map_name,
            CompanyMap.project_id == project_id,
        ).order_by(CompanyMap.id.asc()).first()

        included_override, included_applied, map_cfg = resolve_included_override(
            company=company,
            project_id=project_id,
            map_obj=map_obj,
            map_val=map_name,
            map_role=map_role,
        )

        is_rib, _ = device_is_ribbon(device_for_price, company, project_id)
        ribbon_count = None

        price_splices, price_device, total = compute_prices(
            splices=splices_val,
            device_name=device_for_price,
            company=company,
            project_id=project_id,
            included_override=included_override,
            map_role=map_role,
            ribbon_count=ribbon_count,
        )
        _bcodes = compute_billing_codes(
            splices_val, device_for_price, company, project_id,
            map_role=map_role, ribbon_count=ribbon_count,
        )

        # 4. Atualiza o Record — nunca modifica valores já salvos se não vieram da foto
        rec.splicer               = splicer_name
        rec.map_role              = map_role
        rec.splices               = 0 if is_rib else splices_val
        rec.ribbon_count          = ribbon_count if is_rib else None
        rec.billing_codes_json    = json.dumps(_bcodes, ensure_ascii=False) if _bcodes else None
        rec.price_splices_usd     = price_splices
        rec.price_device_usd      = price_device
        rec.total_usd             = total
        rec.included_splices_applied = included_applied
        rec.ft_in                 = ft_in
        rec.ft_out                = ft_out

        # 5. Salva a foto no record
        try:
            opt_bytes, opt_ct = optimize_upload_bytes(raw_bytes, photo_file.content_type or "image/jpeg")
            thumb_b, thumb_ct = None, None
            try:
                from PIL import Image as _PILImage
                import io as _io
                img_pil = _PILImage.open(_io.BytesIO(opt_bytes))
                img_pil.thumbnail((480, 480))
                buf = _io.BytesIO()
                img_pil.save(buf, format="JPEG", quality=65)
                thumb_b = buf.getvalue()
                thumb_ct = "image/jpeg"
            except Exception:
                pass

            photo_rec = RecordPhoto(
                record_id=rec.id,
                filename=fname,
                data=opt_bytes,
                content_type=opt_ct,
                thumb_data=thumb_b,
                thumb_content_type=thumb_ct,
            )
            db.session.add(photo_rec)
            db.session.flush()

            # Upload para R2 se configurado
            r2_key = r2_key_for_record_photo(rec.id, fname)
            thumb_key = r2_key.rsplit(".", 1)[0] + "_thumb.jpg" if thumb_b else None
            if os.environ.get("R2_BUCKET"):
                enqueue_r2_upload(photo_rec.id, r2_key, opt_bytes, opt_ct, thumb_key, thumb_b, thumb_ct)
        except Exception as e:
            # Foto falhou mas lançamento continua
            print(f"[AUTO-PHOTO] Erro ao salvar foto: {e}")

        db.session.commit()

        results.append({
            "file": fname,
            "ok": True,
            "device": device_name,
            "map_role": map_role,
            "splices": splices_val,
            "ft_in": ft_in,
            "ft_out": ft_out,
            "record_id": rec.id,
        })

    all_ok = all(r["ok"] for r in results)
    return jsonify({"ok": all_ok, "results": results})


# ─────────────────────────────────────────────────────────────
#  FOTO AUTOMÁTICO GLOBAL — sem precisar abrir mapa
# ─────────────────────────────────────────────────────────────

@app.route("/auto-photo")
@login_required
def auto_photo_page():
    """Tela dedicada de lançamento automático por foto (acessível da navbar)."""
    is_owner = bool(getattr(current_user, "is_company_owner", False))
    is_admin = bool(getattr(current_user, "is_admin", False))
    if is_owner and not is_admin:
        flash("Dono de empresa só pode visualizar.", "danger")
        return redirect(url_for("index"))
    return render_template("auto_photo.html")


@app.route("/api/auto-photo-global-launch", methods=["POST"])
@login_required
def auto_photo_global_launch():
    """Recebe fotos Timemark e lança produção automaticamente.

    Fluxo por foto:
      1. Claude lê: device_name, splices, map_role, ft_in, ft_out, gps
      2. GPS da foto é comparado com latitude/longitude de todos os Records
         para encontrar o dispositivo mais próximo (match por nome + proximidade)
      3. Se o nome do dispositivo bater com algum Record perto do GPS → lança
      4. Fallback: busca só pelo nome do dispositivo sem filtro GPS
    """
    is_owner = bool(getattr(current_user, "is_company_owner", False))
    is_admin = bool(getattr(current_user, "is_admin", False))
    if is_owner and not is_admin:
        return jsonify({"ok": False, "error": "Sem permissão."}), 403

    photos = [p for p in request.files.getlist("photos") if getattr(p, "filename", None)]
    if not photos:
        return jsonify({"ok": False, "error": "Nenhuma foto enviada."}), 400

    splicer_name = getattr(current_user, "splicer_name", None) or current_user.username
    user_company = getattr(current_user, "company_name", None) or getattr(current_user, "default_company", None)

    def haversine_m(lat1, lon1, lat2, lon2):
        """Distância em metros entre dois pontos GPS."""
        import math
        R = 6_371_000
        phi1, phi2 = math.radians(lat1), math.radians(lat2)
        dphi = math.radians(lat2 - lat1)
        dlam = math.radians(lon2 - lon1)
        a = math.sin(dphi/2)**2 + math.cos(phi1)*math.cos(phi2)*math.sin(dlam/2)**2
        return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))

    results = []

    for photo_file in photos:
        fname = photo_file.filename or "foto.jpg"
        try:
            raw_bytes = photo_file.read()
        except Exception as e:
            results.append({"file": fname, "ok": False, "error": f"Erro ao ler arquivo: {e}"})
            continue

        # 1. Claude lê a foto
        parsed, ai_error = extract_timestamp_fields_with_ai(raw_bytes)
        if parsed is None:
            results.append({"file": fname, "ok": False, "error": ai_error or "Claude não conseguiu ler a foto."})
            continue

        device_name  = parsed["device_name"]
        splices_val  = parsed["splices"]
        map_role     = parsed["map_role"]
        ft_in        = (parsed.get("ft_in") or "").strip() or None
        ft_out       = (parsed.get("ft_out") or "").strip() or None
        gps_str      = (parsed.get("gps") or "").strip()
        if map_role == "PONTA":
            ft_out = None

        # 2. Localiza o Record pelo nome + GPS
        rec = None
        gps_distance_m = None
        photo_lat = None
        photo_lon = None

        # Parseia GPS da foto se disponível
        if gps_str:
            try:
                # Suporta: "38.823849,-82.229083" ou "38.823849°N, 82.229083°W"
                gps_clean = gps_str.replace("°N","").replace("°S","-").replace("°W",",-").replace("°E",",").replace(" ","")
                parts = gps_clean.split(",")
                photo_lat = float(parts[0])
                photo_lon = float(parts[1])
            except Exception:
                photo_lat = None
                photo_lon = None

        # Busca candidatos pelo nome do dispositivo
        q = Record.query.filter(Record.device == device_name)
        if user_company:
            q = q.filter(Record.company == user_company)
        candidates = q.all()

        if candidates and photo_lat is not None:
            # Calcula distância para cada candidato com coordenadas
            with_coords = [
                (r, haversine_m(photo_lat, photo_lon, r.latitude, r.longitude))
                for r in candidates
                if r.latitude is not None and r.longitude is not None
            ]
            if with_coords:
                with_coords.sort(key=lambda x: x[1])
                closest_rec, closest_dist = with_coords[0]
                rec = closest_rec
                gps_distance_m = round(closest_dist)
            else:
                rec = candidates[0]
        elif candidates:
            rec = candidates[0]

        # Fallback: se não achou pelo nome, tenta achar o device mais próximo pelo GPS puro
        if not rec and photo_lat is not None:
            all_recs = Record.query.filter(
                Record.latitude.isnot(None),
                Record.longitude.isnot(None),
            )
            if user_company:
                all_recs = all_recs.filter(Record.company == user_company)
            all_recs = all_recs.all()

            if all_recs:
                with_dist = [
                    (r, haversine_m(photo_lat, photo_lon, r.latitude, r.longitude))
                    for r in all_recs
                ]
                with_dist.sort(key=lambda x: x[1])
                closest_rec, closest_dist = with_dist[0]
                # Aceita fallback GPS puro só se estiver a menos de 50 metros
                if closest_dist <= 50:
                    rec = closest_rec
                    gps_distance_m = round(closest_dist)

        if not rec:
            results.append({
                "file": fname, "ok": False,
                "error": f"Dispositivo '{device_name}' não encontrado."
                         + (f" GPS: {gps_str}" if gps_str else " (sem GPS na foto)")
            })
            continue

        # 3. Calcula preços e billing codes
        company     = rec.company
        project_id  = rec.project_id
        map_name    = rec.map
        type_val    = (rec.type or "OTE").strip() or "OTE"
        device_for_price = type_val or device_name

        map_obj = CompanyMap.query.filter(
            CompanyMap.company == company,
            CompanyMap.name == map_name,
            CompanyMap.project_id == project_id,
        ).order_by(CompanyMap.id.asc()).first()

        included_override, included_applied, map_cfg = resolve_included_override(
            company=company, project_id=project_id,
            map_obj=map_obj, map_val=map_name, map_role=map_role,
        )

        is_rib, _ = device_is_ribbon(device_for_price, company, project_id)
        ribbon_count = None

        price_splices, price_device, total = compute_prices(
            splices=splices_val, device_name=device_for_price,
            company=company, project_id=project_id,
            included_override=included_override, map_role=map_role, ribbon_count=ribbon_count,
        )
        _bcodes = compute_billing_codes(
            splices_val, device_for_price, company, project_id,
            map_role=map_role, ribbon_count=ribbon_count,
        )

        # 4. Atualiza Record
        rec.splicer                  = splicer_name
        rec.map_role                 = map_role
        rec.splices                  = 0 if is_rib else splices_val
        rec.ribbon_count             = ribbon_count if is_rib else None
        rec.billing_codes_json       = json.dumps(_bcodes, ensure_ascii=False) if _bcodes else None
        rec.price_splices_usd        = price_splices
        rec.price_device_usd         = price_device
        rec.total_usd                = total
        rec.included_splices_applied = included_applied
        rec.ft_in                    = ft_in
        rec.ft_out                   = ft_out

        # 5. Salva foto
        try:
            opt_bytes, opt_ct = optimize_upload_bytes(raw_bytes, photo_file.content_type or "image/jpeg")
            thumb_b, thumb_ct = None, None
            try:
                from PIL import Image as _PILImage
                import io as _io
                img_pil = _PILImage.open(_io.BytesIO(opt_bytes))
                img_pil.thumbnail((480, 480))
                buf = _io.BytesIO()
                img_pil.save(buf, format="JPEG", quality=65)
                thumb_b  = buf.getvalue()
                thumb_ct = "image/jpeg"
            except Exception:
                pass

            photo_rec = RecordPhoto(
                record_id=rec.id, filename=fname,
                data=opt_bytes, content_type=opt_ct,
                thumb_data=thumb_b, thumb_content_type=thumb_ct,
            )
            db.session.add(photo_rec)
            db.session.flush()

            if os.environ.get("R2_BUCKET"):
                r2_key   = r2_key_for_record_photo(rec.id, fname)
                thumb_key = r2_key.rsplit(".", 1)[0] + "_thumb.jpg" if thumb_b else None
                enqueue_r2_upload(photo_rec.id, r2_key, opt_bytes, opt_ct, thumb_key, thumb_b, thumb_ct)
        except Exception as e:
            print(f"[AUTO-PHOTO-GLOBAL] Erro ao salvar foto: {e}")

        db.session.commit()

        results.append({
            "file":           fname,
            "ok":             True,
            "device":         device_name,
            "map_name":       map_name,
            "map_role":       map_role,
            "splices":        splices_val,
            "ft_in":          ft_in,
            "ft_out":         ft_out,
            "gps_distance_m": gps_distance_m,
            "record_id":      rec.id,
        })

    all_ok = all(r["ok"] for r in results)
    return jsonify({"ok": all_ok, "results": results})

@app.route('/__version')
def __version__():
    return 'PHOTO-REMOVE-V49 2026-02-12'